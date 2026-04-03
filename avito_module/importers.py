from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from .compat import clean_text


@dataclass(slots=True)
class KnowledgeImportDoc:
    title: str
    body_text: str
    kind: str = "faq"
    item_id: str = ""
    item_title: str = ""
    tags: List[str] = field(default_factory=list)
    source_name: str = ""
    source_url: str = ""
    active: bool = True
    meta: Dict[str, Any] = field(default_factory=dict)


@dataclass(slots=True)
class KnowledgeImportResult:
    documents: List[KnowledgeImportDoc]
    errors: List[str] = field(default_factory=list)
    detected_format: str = ""


def _csv_list(value: Any) -> List[str]:
    raw = str(value or "")
    return [clean_text(part) for part in raw.replace(";", ",").split(",") if clean_text(part)]


def _normalize_doc(raw: Dict[str, Any], *, default_kind: str, source_name: str, row_hint: str = "") -> Optional[KnowledgeImportDoc]:
    title = clean_text(raw.get("title") or raw.get("name") or raw.get("question") or raw.get("topic") or row_hint)
    body_text = str(raw.get("body_text") or raw.get("body") or raw.get("text") or raw.get("answer") or raw.get("content") or "").strip()
    if not body_text:
        return None
    return KnowledgeImportDoc(
        title=title or "Без названия",
        body_text=body_text,
        kind=clean_text(raw.get("kind") or default_kind) or default_kind,
        item_id=clean_text(raw.get("item_id") or raw.get("sku") or raw.get("article")),
        item_title=clean_text(raw.get("item_title") or raw.get("item") or raw.get("product") or raw.get("listing")),
        tags=_csv_list(raw.get("tags")),
        source_name=clean_text(raw.get("source_name") or source_name),
        source_url=clean_text(raw.get("source_url") or raw.get("url")),
        active=bool(raw.get("active", True)),
        meta={k: v for k, v in raw.items() if k not in {"title", "name", "question", "topic", "body_text", "body", "text", "answer", "content", "kind", "item_id", "sku", "article", "item_title", "item", "product", "listing", "tags", "source_name", "source_url", "url", "active"}},
    )


def load_knowledge_docs_from_bytes(
    payload: bytes,
    *,
    filename: str = "",
    default_kind: str = "faq",
    source_name: str = "upload",
) -> KnowledgeImportResult:
    name = clean_text(filename).lower()
    suffix = Path(name).suffix.lower()
    if suffix in {".txt", ".md", ".markdown"}:
        text = payload.decode("utf-8", errors="ignore")
        return KnowledgeImportResult(
            documents=[KnowledgeImportDoc(title=Path(name or "document").stem or "Документ", body_text=text.strip(), kind=default_kind, source_name=source_name)],
            detected_format=suffix.lstrip("."),
        )
    if suffix == ".jsonl":
        docs: List[KnowledgeImportDoc] = []
        errors: List[str] = []
        for idx, line in enumerate(payload.decode("utf-8", errors="ignore").splitlines(), start=1):
            line = line.strip()
            if not line:
                continue
            try:
                row = json.loads(line)
            except Exception as exc:
                errors.append(f"jsonl line {idx}: {exc}")
                continue
            if not isinstance(row, dict):
                errors.append(f"jsonl line {idx}: expected object")
                continue
            doc = _normalize_doc(row, default_kind=default_kind, source_name=source_name, row_hint=f"Строка {idx}")
            if doc:
                docs.append(doc)
        return KnowledgeImportResult(documents=docs, errors=errors, detected_format="jsonl")
    if suffix == ".json":
        try:
            data = json.loads(payload.decode("utf-8", errors="ignore"))
        except Exception as exc:
            return KnowledgeImportResult(documents=[], errors=[f"json parse error: {exc}"], detected_format="json")
        docs: List[KnowledgeImportDoc] = []
        items: Iterable[Any]
        if isinstance(data, dict):
            if isinstance(data.get("documents"), list):
                items = data.get("documents") or []
            else:
                items = [data]
        elif isinstance(data, list):
            items = data
        else:
            return KnowledgeImportResult(documents=[], errors=["json root must be object or list"], detected_format="json")
        for idx, row in enumerate(items, start=1):
            if not isinstance(row, dict):
                continue
            doc = _normalize_doc(row, default_kind=default_kind, source_name=source_name, row_hint=f"Запись {idx}")
            if doc:
                docs.append(doc)
        return KnowledgeImportResult(documents=docs, detected_format="json")
    if suffix in {".csv", ".tsv"}:
        dialect = csv.excel_tab if suffix == ".tsv" else csv.excel
        text = payload.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        docs = []
        for idx, row in enumerate(reader, start=2):
            doc = _normalize_doc(dict(row), default_kind=default_kind, source_name=source_name, row_hint=f"Row {idx}")
            if doc:
                docs.append(doc)
        return KnowledgeImportResult(documents=docs, detected_format=suffix.lstrip("."))
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as exc:  # pragma: no cover - optional dependency at runtime
            return KnowledgeImportResult(documents=[], errors=[f"openpyxl unavailable: {exc}"], detected_format="xlsx")
        try:
            wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        except Exception as exc:
            return KnowledgeImportResult(documents=[], errors=[f"xlsx parse error: {exc}"], detected_format="xlsx")
        docs: List[KnowledgeImportDoc] = []
        errors: List[str] = []
        for sheet in wb.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            header = [clean_text(x).lower() for x in rows[0]]
            for idx, values in enumerate(rows[1:], start=2):
                row = {header[i]: values[i] for i in range(min(len(header), len(values)))}
                try:
                    doc = _normalize_doc(row, default_kind=default_kind, source_name=source_name, row_hint=f"{sheet.title}:{idx}")
                    if doc:
                        docs.append(doc)
                except Exception as exc:
                    errors.append(f"{sheet.title}:{idx}: {exc}")
        return KnowledgeImportResult(documents=docs, errors=errors, detected_format="xlsx")
    text = payload.decode("utf-8", errors="ignore")
    return KnowledgeImportResult(
        documents=[KnowledgeImportDoc(title=Path(name or "document").stem or "Документ", body_text=text.strip(), kind=default_kind, source_name=source_name)],
        detected_format=(suffix.lstrip(".") or "text"),
    )
