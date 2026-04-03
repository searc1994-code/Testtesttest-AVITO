from __future__ import annotations

import json
import os
import shutil
import zipfile
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple, Callable

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

import automation_core
import background_jobs
import common
import safe_files
import tenant_manager
from safe_logs import log_event


@dataclass
class PriceBuildResult:
    tenant_id: str
    template_path: Path
    output_path: Optional[Path]
    matched: int
    updated: int
    unchanged: int
    missing_in_master: int
    warnings: List[str]
    missing_articles: List[str]
    large_changes: List[Dict[str, Any]]
    master_rows: int


class PricePipelineError(RuntimeError):
    pass


def _clean(value: Any) -> str:
    return common.clean_text(value)


def _normalize_article(value: Any) -> str:
    text = _clean(value)
    return text.replace(" ", "").upper()


def _as_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except Exception:
            return None
    text = str(value).strip().replace("%", "").replace(" ", "")
    if not text:
        return None
    text = text.replace(",", ".")
    try:
        return float(text)
    except Exception:
        return None


def _coerce_excel_value(value: Optional[float]) -> Any:
    if value is None:
        return None
    try:
        quantized = Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return int(round(float(value)))
    return int(quantized)


def _sheet(workbook, name: str = ""):
    sheet_name = _clean(name)
    if sheet_name and sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    return workbook[workbook.sheetnames[0]]


def _excel_col_index(value: str) -> int:
    try:
        return int(column_index_from_string((_clean(value) or "A").upper()))
    except Exception as exc:
        raise PricePipelineError(f"Некорректная колонка Excel: {value!r}") from exc


def _timestamp_slug() -> str:
    return datetime.now().strftime("%Y%m%d-%H%M%S")


def _copy_master_for_run(master_path: Path, run_dir: Path, recalc_mode: str, warnings: List[str]) -> Path:
    copied = run_dir / master_path.name
    shutil.copy2(master_path, copied)
    mode = _clean(recalc_mode) or "auto"
    if mode in {"auto", "windows_excel"} and os.name == "nt":
        try:
            import win32com.client  # type: ignore

            recalculated = run_dir / f"recalc_{master_path.name}"
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            workbook = excel.Workbooks.Open(str(copied), UpdateLinks=0, ReadOnly=False)
            try:
                workbook.Application.CalculateFullRebuild()
                workbook.SaveCopyAs(str(recalculated))
            finally:
                workbook.Close(SaveChanges=False)
                excel.Quit()
            warnings.append("Master-файл пересчитан через Excel COM.")
            return recalculated
        except Exception as exc:
            warnings.append(f"Не удалось пересчитать master-файл через Excel COM: {exc}. Использую сохранённые значения.")
    elif master_path.suffix.lower() == ".xlsm":
        warnings.append("Master-файл xlsm будет прочитан по сохранённым значениям формул. Для максимальной надёжности сохраняйте master перед ночным запуском.")
    return copied


def _read_master_rows(
    master_path: Path,
    settings: Dict[str, Any],
    warnings: List[str],
    progress_callback: Optional[Callable[[int, int], None]] = None,
) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Any]]:
    values_wb = load_workbook(master_path, data_only=True, read_only=True, keep_vba=master_path.suffix.lower() == ".xlsm")
    try:
        sheet = _sheet(values_wb, (settings.get("prices") or {}).get("master_sheet_name"))
        start_row = int((settings.get("prices") or {}).get("row_start") or 2)
        article_col = _excel_col_index((settings.get("prices") or {}).get("master_article_col") or "D")
        price_col = _excel_col_index((settings.get("prices") or {}).get("master_price_col") or "S")
        discount_col = _excel_col_index((settings.get("prices") or {}).get("master_discount_col") or "U")
        max_col = max(article_col, price_col, discount_col)
        total_rows = max(0, int(sheet.max_row or 0) - start_row + 1)

        rows: Dict[str, Dict[str, Any]] = {}
        duplicates: List[str] = []
        missing_values = 0
        scanned = 0
        processed = 0

        for row_idx, row_values in enumerate(sheet.iter_rows(min_row=start_row, max_col=max_col, values_only=True), start=start_row):
            processed += 1
            if progress_callback and (processed == 1 or processed % 1000 == 0 or processed == total_rows):
                progress_callback(processed, total_rows)
            article_raw = row_values[article_col - 1] if len(row_values) >= article_col else None
            article = _normalize_article(article_raw)
            if not article:
                continue
            scanned += 1
            price_raw = row_values[price_col - 1] if len(row_values) >= price_col else None
            discount_raw = row_values[discount_col - 1] if len(row_values) >= discount_col else None
            price = _as_number(price_raw)
            discount = _as_number(discount_raw)
            if price is None and discount is None:
                missing_values += 1
            if article in rows:
                duplicates.append(article)
            rows[article] = {
                "article": article,
                "price": price,
                "discount": discount,
                "row": row_idx,
            }

        if duplicates:
            warnings.append(f"В master обнаружены дубли артикулов: {len(duplicates)}. Использована последняя встретившаяся строка.")
        if missing_values:
            warnings.append(f"В master найдено строк без цены и скидки: {missing_values}.")

        meta = {
            "master_path": str(master_path),
            "sheet": sheet.title,
            "rows_processed": processed,
            "rows_scanned": scanned,
            "rows_loaded": len(rows),
            "duplicates": duplicates[:50],
            "missing_values": missing_values,
        }
        return rows, meta
    finally:
        try:
            values_wb.close()
        except Exception:
            pass


def _large_change(old_value: Optional[float], new_value: Optional[float], threshold_pct: float) -> Optional[float]:
    if old_value in (None, 0) or new_value is None:
        return None
    try:
        delta = abs((float(new_value) - float(old_value)) / float(old_value) * 100.0)
    except Exception:
        return None
    return round(delta, 2) if delta >= float(threshold_pct) else None


def _build_output_filename(tenant_id: str, template_path: Path, settings: Dict[str, Any]) -> str:
    pattern = _clean((settings.get("prices") or {}).get("output_pattern")) or automation_core.DEFAULT_OUTPUT_PATTERN
    suffix = template_path.suffix or ".xlsx"
    filename = pattern.format(date=datetime.now().strftime("%Y%m%d"), tenant_id=_clean(tenant_id), ext=suffix)
    if not Path(filename).suffix:
        filename += suffix
    return filename


def build_tenant_price_file(
    tenant_id: str,
    master_rows: Dict[str, Dict[str, Any]],
    settings: Dict[str, Any],
    run_dir: Path,
    *,
    progress_start: int = 40,
    progress_end: int = 90,
) -> PriceBuildResult:
    tenant_cfg = (settings.get("tenants") or {}).get(_clean(tenant_id)) or {}
    template_path = automation_core.resolve_template_path(tenant_id, settings)
    warnings: List[str] = []
    missing_articles: List[str] = []
    large_changes: List[Dict[str, Any]] = []

    if not template_path.exists():
        raise PricePipelineError(f"Не найден шаблон цен для кабинета {tenant_id}: {template_path.name}")

    progress_start = max(0, min(100, int(progress_start)))
    progress_end = max(progress_start, min(100, int(progress_end)))

    def _progress(stage: str, message: str, fraction: float, **data: Any) -> None:
        percent = progress_start + int((progress_end - progress_start) * max(0.0, min(1.0, fraction)))
        background_jobs.progress(stage, message, percent=percent, tenant_id=tenant_id, **data)

    _progress("prices_template_open", f"Открываю шаблон цен для кабинета {tenant_id}", 0.0, template_file=template_path.name)

    output_dir = run_dir / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / _build_output_filename(tenant_id, template_path, settings)
    shutil.copy2(template_path, output_path)

    workbook = load_workbook(output_path, keep_vba=output_path.suffix.lower() == ".xlsm")
    sheet = _sheet(workbook, (settings.get("prices") or {}).get("template_sheet_name"))
    _progress("prices_template_loaded", f"Шаблон кабинета {tenant_id} открыт", 0.12, rows=int(sheet.max_row or 0))
    start_row = int((settings.get("prices") or {}).get("row_start") or 2)
    article_col = _excel_col_index((settings.get("prices") or {}).get("template_article_col") or "C")
    price_col = _excel_col_index((settings.get("prices") or {}).get("template_price_col") or "J")
    discount_col = _excel_col_index((settings.get("prices") or {}).get("template_discount_col") or "L")
    warn_change_pct = float((settings.get("prices") or {}).get("warn_change_pct") or 30.0)

    matched = 0
    updated = 0
    unchanged = 0
    total_template_rows = max(0, int(sheet.max_row or 0) - start_row + 1)
    for row_idx in range(start_row, sheet.max_row + 1):
        processed_template_rows = row_idx - start_row + 1
        if total_template_rows and (processed_template_rows == 1 or processed_template_rows % 1000 == 0 or processed_template_rows == total_template_rows):
            _progress(
                "prices_template_rows",
                f"Обновляю строки шаблона для кабинета {tenant_id}: {processed_template_rows}/{total_template_rows}",
                0.12 + 0.68 * (processed_template_rows / max(1, total_template_rows)),
                processed=processed_template_rows,
                total=total_template_rows,
            )
        article = _normalize_article(sheet.cell(row=row_idx, column=article_col).value)
        if not article:
            continue
        master_row = master_rows.get(article)
        if not master_row:
            missing_articles.append(article)
            continue
        matched += 1
        new_price = master_row.get("price")
        new_discount = master_row.get("discount")
        target_price = _coerce_excel_value(new_price) if new_price is not None else None
        target_discount = _coerce_excel_value(new_discount) if new_discount is not None else None
        old_price = _as_number(sheet.cell(row=row_idx, column=price_col).value)
        old_discount = _as_number(sheet.cell(row=row_idx, column=discount_col).value)
        current_price = _coerce_excel_value(old_price) if old_price is not None else None
        current_discount = _coerce_excel_value(old_discount) if old_discount is not None else None

        price_delta = _large_change(current_price, target_price, warn_change_pct)
        if price_delta is not None:
            large_changes.append({
                "article": article,
                "kind": "price",
                "old": current_price,
                "new": target_price,
                "delta_pct": price_delta,
            })
        discount_delta = _large_change(current_discount, target_discount, warn_change_pct)
        if discount_delta is not None:
            large_changes.append({
                "article": article,
                "kind": "discount",
                "old": current_discount,
                "new": target_discount,
                "delta_pct": discount_delta,
            })

        changed = False
        if target_price is not None and current_price != target_price:
            price_cell = sheet.cell(row=row_idx, column=price_col)
            price_cell.value = target_price
            try:
                price_cell.number_format = '0'
            except Exception:
                pass
            changed = True
        if target_discount is not None and current_discount != target_discount:
            discount_cell = sheet.cell(row=row_idx, column=discount_col)
            discount_cell.value = target_discount
            try:
                discount_cell.number_format = '0'
            except Exception:
                pass
            changed = True

        if changed:
            updated += 1
        else:
            unchanged += 1

    _progress("prices_template_save", f"Сохраняю итоговый файл для кабинета {tenant_id}", 0.86, output_file=output_path.name)
    workbook.save(output_path)
    try:
        workbook.close()
    except Exception:
        pass
    shutil.copy2(output_path, automation_core.PRICE_OUTPUT_DIR / output_path.name)
    _progress("prices_template_saved", f"Файл для кабинета {tenant_id} сохранён", 0.97, output_file=output_path.name)

    if missing_articles:
        warnings.append(f"В шаблоне {template_path.name} есть артикулы, которых нет в master: {len(missing_articles)}.")
    if large_changes:
        warnings.append(f"Обнаружены изменения цены/скидки более {warn_change_pct:.0f}%: {len(large_changes)} строк.")

    return PriceBuildResult(
        tenant_id=tenant_id,
        template_path=template_path,
        output_path=output_path,
        matched=matched,
        updated=updated,
        unchanged=unchanged,
        missing_in_master=len(missing_articles),
        warnings=warnings,
        missing_articles=missing_articles[:200],
        large_changes=large_changes[:200],
        master_rows=len(master_rows),
    )


def build_price_files(tenant_ids: Optional[Iterable[str]] = None, run_source: str = "manual") -> Dict[str, Any]:
    settings = automation_core.load_settings()
    automation_core.ensure_dirs()
    background_jobs.progress("prices_init", "Проверяю рабочую папку и master-файл", percent=0, source=run_source)

    master_path = automation_core.resolve_master_path(settings)
    if not master_path.exists():
        raise PricePipelineError(f"Master-файл не найден: {master_path.name}. Положите его в папку {automation_core.PRICE_WORKSPACE_DIR}")

    selected = [
        _clean(tenant_id)
        for tenant_id in (tenant_ids or automation_core.list_enabled_tenant_ids(settings, feature="prices"))
        if _clean(tenant_id)
    ]
    if not selected:
        raise PricePipelineError("Нет кабинетов, включённых для ночной загрузки цен.")

    run_dir = automation_core.create_run_dir("prices_build")
    safe_files.write_text(run_dir / "source.txt", f"run_source={_clean(run_source) or 'manual'}\n", encoding="utf-8")
    warnings: List[str] = []
    copied_master = _copy_master_for_run(master_path, run_dir, (settings.get("prices") or {}).get("recalc_mode") or "auto", warnings)
    background_jobs.progress("prices_master", "Читаю master-файл с ценами", percent=5, master_file=master_path.name)

    def _master_progress(processed: int, total: int) -> None:
        if total <= 0:
            return
        percent = 5 + int(25 * (processed / max(1, total)))
        background_jobs.progress(
            "prices_master_scan",
            f"Читаю master-файл с ценами: {processed}/{total}",
            percent=percent,
            master_file=master_path.name,
            processed=processed,
            total=total,
        )

    master_rows, master_meta = _read_master_rows(copied_master, settings, warnings, progress_callback=_master_progress)
    background_jobs.progress(
        "prices_master_done",
        "Master-файл с ценами прочитан",
        percent=35,
        master_file=master_path.name,
        rows_loaded=len(master_rows),
        rows_scanned=int(master_meta.get("rows_scanned") or 0),
    )

    results: List[Dict[str, Any]] = []
    failures: List[Dict[str, Any]] = []
    for index, tenant_id in enumerate(selected, start=1):
        tenant_progress_start = 35 + int((index - 1) / max(len(selected), 1) * 55)
        tenant_progress_end = 35 + int(index / max(len(selected), 1) * 55)
        background_jobs.progress(
            "prices_tenant_start",
            f"Готовлю файл цен для кабинета {tenant_id}",
            percent=tenant_progress_start,
            tenant_id=tenant_id,
            current=index,
            total=len(selected),
        )
        try:
            result = build_tenant_price_file(tenant_id, master_rows, settings, run_dir, progress_start=tenant_progress_start, progress_end=tenant_progress_end)
            results.append({
                "tenant_id": result.tenant_id,
                "template_path": str(result.template_path),
                "output_path": str(result.output_path) if result.output_path else "",
                "matched": result.matched,
                "updated": result.updated,
                "unchanged": result.unchanged,
                "missing_in_master": result.missing_in_master,
                "warnings": result.warnings,
                "missing_articles": result.missing_articles,
                "large_changes": result.large_changes,
                "master_rows": result.master_rows,
            })
            log_event("automation", "price_file_built", tenant_id=tenant_id, output_path=str(result.output_path), matched=result.matched, updated=result.updated, missing_in_master=result.missing_in_master)
            background_jobs.progress(
                "prices_tenant_done",
                f"Файл для кабинета {tenant_id} подготовлен",
                percent=tenant_progress_end,
                tenant_id=tenant_id,
                matched=result.matched,
                updated=result.updated,
                missing_in_master=result.missing_in_master,
            )
        except Exception as exc:
            error_text = _clean(exc)
            failures.append({"tenant_id": tenant_id, "error": error_text})
            log_event("automation", "price_file_failed", tenant_id=tenant_id, level="error", error=error_text)
            background_jobs.progress(
                "prices_tenant_error",
                f"Не удалось подготовить файл для кабинета {tenant_id}: {error_text}",
                percent=tenant_progress_end,
                tenant_id=tenant_id,
                error=error_text,
            )
            continue

    summary = {
        "run_source": _clean(run_source) or "manual",
        "run_dir": str(run_dir),
        "master": master_meta,
        "warnings": warnings,
        "selected_tenants": selected,
        "results": results,
        "failures": failures,
        "prepared": len(results),
        "failed": len(failures),
    }
    summary_path = run_dir / "summary.json"
    safe_files.write_json(summary_path, summary, ensure_ascii=False, indent=2)
    archive_path = Path(shutil.make_archive(str(run_dir), "zip", root_dir=run_dir))
    summary["archive_path"] = str(archive_path)
    report_path = automation_core.write_report(
        "prices_build",
        status="completed" if not failures else "partial",
        title="Подготовка файлов цен по юрлицам",
        payload=summary,
    )
    summary["report_path"] = str(report_path)
    safe_files.write_json(summary_path, summary, ensure_ascii=False, indent=2)
    background_jobs.progress("prices_done", "Подготовка файлов цен завершена", percent=100, prepared=len(results), failed=len(failures))
    return {
        **summary,
        "message": f"Подготовлено файлов: {len(results)}. Ошибок: {len(failures)}.",
    }


def workspace_health(settings: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    cfg = settings or automation_core.load_settings()
    manifest = automation_core.build_workspace_manifest(cfg)
    status = {
        "master_exists": manifest.get("master_exists", False),
        "templates_missing": [row["tenant_id"] for row in manifest.get("tenants", []) if not row.get("template_exists")],
        "files_total": len(manifest.get("files", [])),
        "workspace_dir": manifest.get("workspace_dir"),
    }
    status["ok"] = bool(status["master_exists"] and not status["templates_missing"])
    return status
