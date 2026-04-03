import csv
import hashlib
import hmac
import io
import json
import os
import subprocess
import sys
import re
import time
import zipfile
from urllib.parse import urlparse
from collections import Counter, defaultdict
from contextlib import contextmanager
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import automation_core
import automation_scheduler
import auth_core
import background_jobs
import common
import config
import history_service
import price_pipeline
import price_uploader
import promo_calendar
import promo_executor
import safe_files
import tenant_manager
from flask import Flask, flash, g, jsonify, redirect, render_template, request, send_file, session, url_for
from safe_logs import log_event, list_channels as safe_list_channels, list_tenants as safe_list_tenants, read_events as safe_read_events, stats as safe_log_stats
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from browser_bot import get_auth_status, process_queue
try:
    from browser_bot import refresh_complaint_outcomes
except Exception:
    def refresh_complaint_outcomes(max_items: int = 0):
        return {"processed": 0, "accepted": 0, "rejected": 0, "pending": 0, "notes": ["Функция обновления исходов жалоб недоступна в текущем browser_bot.py"]}
from common import (
    ARCHIVE_FILE,
    DRAFTS_FILE,
    FLASK_SECRET,
    OPENAI_MODEL,
    PRIVATE_DIR,
    REQUEST_TIMEOUT,
    WB_ANSWER_URL,
    WB_SESSION,
    build_review_text,
    call_ai,
    clean_text,
    fetch_pending_reviews,
    load_rules,
    load_system_prompt,
    normalize_review,
    read_json,
    restore_review_from_form,
    review_signature,
    utc_now_iso,
    write_json,
)
from complaint_core import (
    analyze_complaint_review,
    complaint_dashboard_stats,
    fetch_low_rating_reviews,
    load_complaint_drafts,
    load_complaint_queue,
    queue_complaint_entries,
)
try:
    from complaint_core import build_complaint_effectiveness
except Exception:
    def build_complaint_effectiveness(limit: int = 5000):
        return {
            "total": 0,
            "accepted": 0,
            "rejected": 0,
            "pending": 0,
            "by_category": [],
            "by_product": [],
            "recent": [],
            "notes": ["Функция аналитики результативности жалоб недоступна в текущем complaint_core.py"],
        }
from question_core import (
    QUESTION_PAGE_SIZE_OPTIONS,
    apply_imported_clusters_for_active_tenant,
    auto_split_cluster_by_article,
    build_question_clustering_export_row,
    build_question_rows,
    build_questions_context,
    generate_question_draft,
    get_question_snapshot,
    ignore_question_ids,
    list_question_ids_for_cluster,
    load_question_rules,
    load_question_snapshot,
    parse_rule_form,
    process_auto_question_rules,
    process_question_queue,
    queue_questions_from_form,
    refresh_question_snapshot,
    reassign_cluster_members,
    reset_cluster_assignments,
    save_question_form_edits,
    save_question_prompt as save_question_prompt_text,
    toggle_question_rule,
    upsert_question_rule,
)

try:
    from avito_module.blueprint import register_avito_module
except Exception:
    def register_avito_module(app: Flask, **kwargs):
        return app

app = Flask(__name__)
app.secret_key = config.FLASK_SECRET
app.config.update(
    SECRET_KEY=config.FLASK_SECRET,
    MAX_CONTENT_LENGTH=256 * 1024 * 1024,
    MAX_FORM_MEMORY_SIZE=None,
    MAX_FORM_PARTS=None,
    DEBUG=bool(getattr(config, "FLASK_DEBUG", False)),
    SESSION_COOKIE_HTTPONLY=bool(getattr(config, "SESSION_COOKIE_HTTPONLY", True)),
    SESSION_COOKIE_SAMESITE=getattr(config, "SESSION_COOKIE_SAMESITE", "Lax"),
    SESSION_COOKIE_SECURE=bool(getattr(config, "SESSION_COOKIE_SECURE", False)),
    PERMANENT_SESSION_LIFETIME=timedelta(hours=int(getattr(config, "PERMANENT_SESSION_LIFETIME_HOURS", 12))),
)

try:
    register_avito_module(app)
except Exception:
    pass


def _auth_enabled() -> bool:
    return bool(getattr(config, "APP_AUTH_ENABLED", True))


def _admin_setup_required() -> bool:
    return _auth_enabled() and auth_core.needs_bootstrap()


def _is_logged_in() -> bool:
    return (not _auth_enabled()) or bool(session.get("auth_ok"))


def _current_tenant_for_logs() -> str:
    try:
        return common.clean_text(getattr(g, 'active_tenant_id', '') or session.get('tenant_id') or '')
    except Exception:
        return ''


def _current_user_for_logs() -> str:
    return clean_text(session.get("auth_user"))


def _same_origin_request() -> bool:
    origin = clean_text(request.headers.get("Origin"))
    referer = clean_text(request.headers.get("Referer"))
    host = clean_text(request.host_url).rstrip("/")
    if origin:
        return origin.rstrip("/") == host
    if referer:
        parsed_ref = urlparse(referer)
        parsed_host = urlparse(host)
        return parsed_ref.scheme == parsed_host.scheme and parsed_ref.netloc == parsed_host.netloc
    return True


@contextmanager
def _temporary_tenant_binding(tenant_id: str):
    tenant_id = clean_text(tenant_id)
    if not tenant_id:
        yield None
        return
    tenant = tenant_manager.get_tenant(tenant_id)
    if not tenant:
        raise ValueError(f"Кабинет не найден: {tenant_id}")
    paths = tenant_manager.get_tenant_paths(tenant_id)
    tokens = common.bind_tenant_context(tenant_id, tenant=tenant, paths=paths)
    previous_tenant = getattr(g, 'active_tenant', None)
    previous_tenant_id = clean_text(getattr(g, 'active_tenant_id', ''))
    try:
        g.active_tenant = tenant
        g.active_tenant_id = tenant_id
        yield tenant
    finally:
        common.reset_tenant_context(tokens)
        g.active_tenant = previous_tenant
        g.active_tenant_id = previous_tenant_id


@app.after_request
def _safe_request_logger(response):
    try:
        path = request.path or ''
        if not path.startswith('/static'):
            should_log = request.method == 'POST' or path in {'/', '/complaints', '/analytics', '/tenants', '/diagnostics', '/questions'}
            if should_log:
                started = getattr(g, 'request_started_at', None)
                duration_ms = round((time.monotonic() - started) * 1000, 2) if started else None
                log_event(
                    'app',
                    'http_request',
                    tenant_id=_current_tenant_for_logs(),
                    method=request.method,
                    path=path,
                    endpoint=request.endpoint or '',
                    status_code=getattr(response, 'status_code', 0),
                    duration_ms=duration_ms,
                    query=dict(request.args),
                    user=_current_user_for_logs(),
                )
    except Exception:
        pass
    try:
        response.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
        response.headers.setdefault("X-Content-Type-Options", "nosniff")
        response.headers.setdefault("Referrer-Policy", "same-origin")
        response.headers.setdefault("Content-Security-Policy", "default-src 'self'; style-src 'self' 'unsafe-inline'; script-src 'self' 'unsafe-inline'; img-src 'self' data:; connect-src 'self'; frame-ancestors 'self'; base-uri 'self'; form-action 'self'")
    except Exception:
        pass
    return response


@app.teardown_request
def _safe_request_finalize(exc):
    try:
        if exc:
            log_event('app', 'http_error', tenant_id=_current_tenant_for_logs(), level='error', path=request.path, endpoint=request.endpoint or '', error=str(exc), user=_current_user_for_logs())
    except Exception:
        pass
    try:
        common.reset_tenant_context(getattr(g, 'tenant_context_tokens', None))
    except Exception:
        pass
    return None


@app.before_request
def _bind_active_tenant() -> Any:
    g.request_started_at = time.monotonic()
    g.tenant_context_tokens = None
    g.tenants = tenant_manager.load_tenants()
    g.active_tenant = None
    g.active_tenant_id = ''
    endpoint = request.endpoint or ''
    is_webhook_endpoint = endpoint == 'avito_module.avito_webhook'

    public_auth_endpoints = {'login', 'logout', 'setup_admin', 'static', 'avito_module.avito_webhook'}
    if _admin_setup_required() and endpoint not in {'setup_admin', 'static', 'avito_module.avito_webhook'} and not endpoint.startswith('static'):
        return redirect(url_for('setup_admin', next=request.full_path if request.query_string else request.path))
    if _auth_enabled() and endpoint not in public_auth_endpoints and not endpoint.startswith('static') and not _is_logged_in():
        return redirect(url_for('login', next=request.full_path if request.query_string else request.path))

    if request.method == 'POST' and endpoint not in {'static', 'avito_module.avito_webhook'} and not endpoint.startswith('static') and not _same_origin_request():
        log_event('security', 'cross_origin_post_blocked', tenant_id=_current_tenant_for_logs(), level='warning', path=request.path, endpoint=endpoint, origin=clean_text(request.headers.get('Origin')), referer=clean_text(request.headers.get('Referer')), user=_current_user_for_logs())
        return ("Forbidden", 403)

    allowed_without_tenants = {'tenants', 'tenants_add', 'tenant_select', 'login', 'logout', 'setup_admin', 'static', 'avito_module.avito_webhook'}
    if not g.tenants:
        if is_webhook_endpoint:
            return ("Tenant not found", 404)
        if endpoint not in allowed_without_tenants and not endpoint.startswith('static'):
            return redirect(url_for('tenants'))
        return None

    requested_id = clean_text(
        (request.view_args or {}).get('tenant_id')
        or request.values.get('tenant_id')
        or request.args.get('tenant_id')
    )
    session_id = clean_text(session.get('tenant_id'))
    active_id = requested_id or session_id or clean_text(g.tenants[0].get('id'))
    valid_ids = {clean_text(t.get('id')) for t in g.tenants}
    if active_id not in valid_ids:
        if is_webhook_endpoint:
            return ("Tenant not found", 404)
        active_id = clean_text(g.tenants[0].get('id'))

    session['tenant_id'] = active_id
    tenant = tenant_manager.get_tenant(active_id) or {}
    paths = tenant_manager.ensure_tenant_dirs(active_id)
    g.tenant_context_tokens = common.bind_tenant_context(active_id, tenant=tenant, paths=paths)
    g.active_tenant = tenant
    g.active_tenant_id = active_id
    return None


@app.context_processor
def _inject_tenant_nav() -> Dict[str, Any]:
    active = getattr(g, 'active_tenant', None)
    active_id = clean_text(active.get('id')) if isinstance(active, dict) else clean_text(getattr(g, 'active_tenant_id', ''))
    latest_jobs = background_jobs.list_latest_jobs_by_kind(active_id, limit=8) if active_id else []
    return {
        'tenants_nav': tenant_manager.collect_tenant_summaries(),
        'active_tenant': active,
        'active_tenant_id': active_id,
        'background_jobs_latest': latest_jobs,
        'app_auth_enabled': _auth_enabled(),
        'current_user': clean_text(session.get('auth_user')) if _is_logged_in() else '',
    }


def _tenant_query_arg() -> Dict[str, Any]:
    active_id = clean_text(getattr(g, 'active_tenant_id', '') or session.get('tenant_id'))
    return {'tenant_id': active_id} if active_id else {}


REPLY_QUEUE_FILE = common.REPLY_QUEUE_FILE
REPLY_SNAPSHOT_FILE = common.REPLY_SNAPSHOT_FILE
SYSTEM_PROMPT_FILE = common.SYSTEM_PROMPT_FILE
REPLY_SNAPSHOT_TTL_SECONDS = int(getattr(common, "REPLY_SNAPSHOT_TTL_SECONDS", 90))
REPLY_RAW_BATCH_SIZE = int(getattr(common, "REPLY_RAW_BATCH_SIZE", 100))
REPLY_MAX_RAW_BATCHES = int(getattr(common, "REPLY_MAX_RAW_BATCHES", 20))
PAGE_SIZE_OPTIONS = [10, 20, 100, 300, 1000]
REPLY_DEFAULT_PAGE_SIZE = 100
COMPLAINT_DEFAULT_PAGE_SIZE = 100
API_SEND_DELAY_SECONDS = float(getattr(common, "API_SEND_DELAY_SECONDS", 0.4))
MAX_REPLY_LENGTH = 4500
MIN_REPLY_LENGTH = 1500
REPLY_SENT_STATUSES = {"sent", "success", "submitted"}
BASE_SYSTEM_PROMPT_FILE = Path(__file__).with_name("system_prompt.txt")


def _submit_background_task(kind: str, label: str, target, *args: Any, unique_key: str = "", **kwargs: Any) -> Dict[str, Any]:
    tenant_id = clean_text(getattr(g, "active_tenant_id", "") or session.get("tenant_id"))
    job, created = background_jobs.submit_job(
        kind=kind,
        tenant_id=tenant_id,
        label=label,
        target=target,
        args=args,
        kwargs=kwargs,
        unique_key=unique_key,
    )
    job_id = clean_text(job.get("job_id"))
    if created:
        flash(f"{label}: задача поставлена в фон. Job ID: {job_id}.", "success")
    else:
        flash(f"{label}: похожая задача уже выполняется. Job ID: {job_id}.", "success")
    return job


def _job_prepare_replies(reviews: List[Dict[str, Any]], force: bool = False) -> Dict[str, Any]:
    generated = 0
    errors: List[str] = []
    for review in reviews:
        try:
            generate_reply_for_review(review, force=force)
            generated += 1
        except Exception as exc:
            errors.append(f"{clean_text(review.get('id'))}: {exc}")
    return {"generated": generated, "failed": len(errors), "errors": errors[:20], "message": f"Подготовлено черновиков: {generated}. Ошибок: {len(errors)}."}


def _job_prepare_complaints(reviews: List[Dict[str, Any]], force: bool = False) -> Dict[str, Any]:
    prepared = 0
    errors: List[str] = []
    for review in reviews:
        try:
            analyze_complaint_review(review, force=force)
            prepared += 1
        except Exception as exc:
            errors.append(f"{clean_text(review.get('id'))}: {exc}")
    return {"prepared": prepared, "failed": len(errors), "errors": errors[:20], "message": f"Подготовлено жалоб: {prepared}. Ошибок: {len(errors)}."}


def _job_refresh_questions() -> Dict[str, Any]:
    snapshot = refresh_question_snapshot()
    return {
        "count_unanswered": int(snapshot.get("count_unanswered") or 0),
        "questions": len(snapshot.get("questions") or []),
        "message": f"Снимок вопросов обновлён. В кэше неотвеченных: {int(snapshot.get('count_unanswered') or 0)}.",
    }


def _job_prepare_questions(selected_ids: List[str], force: bool = False) -> Dict[str, Any]:
    snapshot = get_question_snapshot(force_refresh=False)
    snapshot_map = {
        clean_text(item.get("id")): common.normalize_question(item)
        for item in snapshot.get("questions") or []
        if clean_text(item.get("id"))
    }
    generated = 0
    missing = 0
    for question_id in selected_ids:
        question = snapshot_map.get(question_id)
        if not question:
            missing += 1
            continue
        generate_question_draft(question, force=force)
        generated += 1
    return {
        "generated": generated,
        "missing": missing,
        "message": f"Подготовлено черновиков вопросов: {generated}. Не найдено в снимке: {missing}.",
    }


@app.route("/setup-admin", methods=["GET", "POST"])
def setup_admin() -> Any:
    if not _auth_enabled():
        return redirect(url_for("index"))
    if not _admin_setup_required():
        return redirect(url_for("login"))

    auth_state = auth_core.describe_auth_state()
    next_url = clean_text(request.values.get("next") or url_for("index"))
    if request.method == "POST":
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")
        try:
            record = auth_core.bootstrap_admin_password(
                password,
                confirm_password=confirm_password,
                username=auth_state.get("username") or getattr(config, "APP_ADMIN_USERNAME", "admin"),
            )
            session.clear()
            session["auth_ok"] = True
            session["auth_user"] = clean_text(record.get("username") or auth_state.get("username") or "admin")
            session.permanent = True
            flash("Пароль администратора сохранён. Вход выполнен.", "success")
            if next_url.startswith("/"):
                return redirect(next_url)
            return redirect(url_for("index"))
        except Exception as exc:
            flash(f"Не удалось сохранить пароль администратора: {exc}", "error")
    return render_template(
        "setup_admin.html",
        next=next_url,
        auth_state=auth_state,
        min_password_length=auth_core.MIN_PASSWORD_LENGTH,
    )


@app.route("/login", methods=["GET", "POST"])
def login() -> Any:
    if not _auth_enabled():
        session["auth_ok"] = True
        session["auth_user"] = clean_text(getattr(config, "APP_ADMIN_USERNAME", "admin")) or "admin"
        return redirect(request.args.get("next") or url_for("index"))
    if _admin_setup_required():
        return redirect(url_for("setup_admin", next=request.args.get("next") or request.form.get("next") or url_for("index")))

    auth_state = auth_core.describe_auth_state()
    expected_user = clean_text(auth_state.get("username") or getattr(config, "APP_ADMIN_USERNAME", "admin")) or "admin"
    if request.method == "POST":
        username = clean_text(request.form.get("username"))
        password = request.form.get("password", "")
        if auth_core.verify_credentials(username, password):
            session.clear()
            session["auth_ok"] = True
            session["auth_user"] = expected_user
            session.permanent = True
            flash("Вход выполнен.", "success")
            next_url = clean_text(request.form.get("next") or request.args.get("next"))
            if next_url and next_url.startswith("/"):
                return redirect(next_url)
            return redirect(url_for("index"))
        flash("Неверный логин или пароль.", "error")
    return render_template(
        "login.html",
        next=request.args.get("next") or request.form.get("next") or url_for("index"),
        auth_enabled=_auth_enabled(),
        admin_username=expected_user,
        auth_state=auth_state,
    )


@app.route("/logout", methods=["GET", "POST"])
def logout() -> Any:
    session.clear()
    flash("Сессия завершена.", "success")
    return redirect(url_for("login"))


@app.route("/jobs/status")
def jobs_status() -> Any:
    tenant_id = clean_text(request.args.get("tenant_id")) or clean_text(getattr(g, "active_tenant_id", "") or session.get("tenant_id"))
    job_id = clean_text(request.args.get("job_id"))
    if job_id:
        payload = background_jobs.get_job(job_id) or {}
    else:
        payload = {"jobs": background_jobs.list_jobs(tenant_id=tenant_id, limit=max(1, min(50, _safe_int(request.args.get("limit", 10), 10))))}
    return jsonify(payload)


@app.route("/replies/history-sync/start", methods=["POST"])
def replies_history_sync_start() -> Any:
    params = _reviews_params_from_source(request.form)
    tenant_id = clean_text(getattr(g, "active_tenant_id", "") or session.get("tenant_id"))
    try:
        started = history_service.start_sync(tenant_id)
        if started:
            flash("Историческая синхронизация запущена во внешнем worker-процессе.", "success")
        else:
            flash("Историческая синхронизация уже выполняется для этого кабинета.", "success")
    except Exception as exc:
        flash(f"Не удалось запустить историческую синхронизацию: {exc}", "error")
    return redirect(url_for("index", **params))


@app.route("/replies/history-sync/stop", methods=["POST"])
def replies_history_sync_stop() -> Any:
    params = _reviews_params_from_source(request.form)
    tenant_id = clean_text(getattr(g, "active_tenant_id", "") or session.get("tenant_id"))
    try:
        history_service.stop_sync(tenant_id)
        flash("Запрошена остановка исторической синхронизации.", "success")
    except Exception as exc:
        flash(f"Не удалось остановить историческую синхронизацию: {exc}", "error")
    return redirect(url_for("index", **params))


@app.route("/replies/history-sync/check", methods=["POST"])
def replies_history_sync_check() -> Any:
    params = _reviews_params_from_source(request.form)
    tenant_id = clean_text(getattr(g, "active_tenant_id", "") or session.get("tenant_id"))
    try:
        counts = history_service.get_counts(tenant_id)
        meta = history_service.effective_meta(tenant_id)
        flash(
            f"Историческая БД: total={counts.get('total', 0)}, active={counts.get('active', 0)}, archive={counts.get('archive', 0)}, needs_reply={counts.get('needs_reply', 0)}. Статус: {meta.get('status') or 'idle'}.",
            "success",
        )
    except Exception as exc:
        flash(f"Не удалось проверить состояние исторической синхронизации: {exc}", "error")
    return redirect(url_for("index", **params))


@app.route("/replies/history-sync/status")
def replies_history_sync_status() -> Any:
    tenant_id = clean_text(getattr(g, "active_tenant_id", "") or session.get("tenant_id"))
    return jsonify(history_service.job_payload(tenant_id))


def load_archive() -> List[Dict[str, Any]]:
    data = read_json(ARCHIVE_FILE, [])
    return data if isinstance(data, list) else []


def save_archive(data: List[Dict[str, Any]]) -> None:
    write_json(ARCHIVE_FILE, data)


def load_drafts() -> Dict[str, Dict[str, Any]]:
    data = read_json(DRAFTS_FILE, {})
    return data if isinstance(data, dict) else {}


def save_drafts(data: Dict[str, Dict[str, Any]]) -> None:
    write_json(DRAFTS_FILE, data)


def load_reply_queue() -> List[Dict[str, Any]]:
    data = read_json(REPLY_QUEUE_FILE, [])
    return data if isinstance(data, list) else []


def save_reply_queue(data: List[Dict[str, Any]]) -> None:
    write_json(REPLY_QUEUE_FILE, data)


def load_reply_snapshot() -> Dict[str, Any]:
    data = read_json(REPLY_SNAPSHOT_FILE, {})
    return data if isinstance(data, dict) else {}


def save_reply_snapshot(data: Dict[str, Any]) -> None:
    write_json(REPLY_SNAPSHOT_FILE, data)


def get_locally_sent_ids() -> set[str]:
    return {clean_text(item.get("id")) for item in load_archive() if clean_text(item.get("status")) in REPLY_SENT_STATUSES}


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(value or default)
    except Exception:
        return default


def _safe_bool_form(value: Any) -> bool:
    return clean_text(value).lower() in {"1", "true", "yes", "on"}


def _submit_system_background_task(kind: str, label: str, target, *args: Any, unique_key: str = "", **kwargs: Any) -> Dict[str, Any]:
    job, created = background_jobs.submit_job(
        kind=kind,
        tenant_id="_system",
        label=label,
        target=target,
        args=args,
        kwargs=kwargs,
        unique_key=unique_key,
    )
    job_id = clean_text(job.get("job_id"))
    if created:
        flash(f"{label}: задача поставлена в фон. Job ID: {job_id}.", "success")
    else:
        flash(f"{label}: похожая задача уже выполняется. Job ID: {job_id}.", "success")
    return job


def _selected_automation_tenants_from_form(feature: str) -> List[str]:
    settings = automation_core.load_settings()
    requested = [clean_text(item) for item in request.form.getlist("tenant_ids") if clean_text(item)]
    if clean_text(request.form.get("tenant_scope")) == "all" or not requested:
        return automation_core.list_enabled_tenant_ids(settings, feature=feature)
    enabled = set(automation_core.list_enabled_tenant_ids(settings, feature=feature))
    return [tenant_id for tenant_id in requested if tenant_id in enabled]


def _parse_created_date(value: Any) -> datetime:
    text = clean_text(value)
    if not text:
        return datetime.min.replace(tzinfo=timezone.utc)
    for candidate in [text, text.replace("Z", "+00:00")]:
        try:
            dt = datetime.fromisoformat(candidate)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt
        except Exception:
            continue
    return datetime.min.replace(tzinfo=timezone.utc)


def _snapshot_is_fresh(snapshot: Dict[str, Any]) -> bool:
    fetched_at = _parse_created_date(snapshot.get("fetched_at"))
    age = datetime.now(timezone.utc) - fetched_at
    return age.total_seconds() <= REPLY_SNAPSHOT_TTL_SECONDS


def _fetch_reply_snapshot_from_wb() -> Dict[str, Any]:
    all_feedbacks: List[Dict[str, Any]] = []
    raw_skip = 0
    count_unanswered = 0
    count_archive = 0
    for batch_index in range(REPLY_MAX_RAW_BATCHES):
        feedbacks, count_unanswered, count_archive = fetch_pending_reviews(skip=raw_skip, take=REPLY_RAW_BATCH_SIZE)
        if not feedbacks:
            break
        for review in feedbacks:
            all_feedbacks.append(normalize_review(review))
        raw_skip += REPLY_RAW_BATCH_SIZE
        if len(feedbacks) < REPLY_RAW_BATCH_SIZE:
            break
        if batch_index < REPLY_MAX_RAW_BATCHES - 1:
            time.sleep(0.36)
    snapshot = {
        "fetched_at": utc_now_iso(),
        "count_unanswered": count_unanswered,
        "count_archive": count_archive,
        "raw_scanned": raw_skip,
        "feedbacks": all_feedbacks,
    }
    save_reply_snapshot(snapshot)
    try:
        history_service.upsert_active_snapshot(snapshot, getattr(g, "active_tenant_id", "") or common.get_active_tenant_id())
    except Exception as exc:
        log_event("replies", "history_snapshot_upsert_failed", tenant_id=_current_tenant_for_logs(), level="error", error=str(exc))
    return snapshot


def get_reply_snapshot(force_refresh: bool = False) -> Dict[str, Any]:
    snapshot = load_reply_snapshot()
    if snapshot and not force_refresh and _snapshot_is_fresh(snapshot):
        return snapshot
    return _fetch_reply_snapshot_from_wb()


def _reply_snapshot_map(force_refresh: bool = False) -> Dict[str, Dict[str, Any]]:
    snapshot = get_reply_snapshot(force_refresh=force_refresh)
    result: Dict[str, Dict[str, Any]] = {}
    for review in snapshot.get("feedbacks", []):
        nr = normalize_review(review)
        review_id = clean_text(nr.get("id"))
        if review_id:
            result[review_id] = nr
    return result


def _history_row_to_review(row: Dict[str, Any]) -> Dict[str, Any]:
    return normalize_review({
        "id": clean_text(row.get("review_id") or row.get("id")),
        "text": clean_text(row.get("text")),
        "pros": clean_text(row.get("pros")),
        "cons": clean_text(row.get("cons")),
        "productValuation": _safe_int(row.get("stars")),
        "createdDate": clean_text(row.get("created_date")),
        "userName": clean_text(row.get("user_name")),
        "subjectName": clean_text(row.get("subject_name")),
        "answer": {
            "text": clean_text(row.get("answer_text")),
            "createDate": clean_text(row.get("answer_create_date")),
        },
        "productDetails": {
            "productName": clean_text(row.get("product_name")),
            "supplierArticle": clean_text(row.get("supplier_article")),
            "brandName": clean_text(row.get("brand_name")),
            "nmId": _safe_int(row.get("nm_id")),
        },
    })


def _get_snapshot_reviews_by_ids(selected_ids: List[str], force_refresh: bool = False) -> Tuple[List[Dict[str, Any]], List[str]]:
    review_map = _reply_snapshot_map(force_refresh=force_refresh)
    found: List[Dict[str, Any]] = []
    missing: List[str] = []
    for review_id in [clean_text(x) for x in selected_ids if clean_text(x)]:
        review = review_map.get(review_id)
        if review is None:
            history_row = history_service.get_row_by_id(review_id, getattr(g, "active_tenant_id", "") or common.get_active_tenant_id())
            if history_row:
                review = _history_row_to_review(history_row)
        if review:
            found.append(review)
        else:
            missing.append(review_id)
    return found, missing


def _save_server_reply_drafts(items: List[Dict[str, Any]]) -> Tuple[int, List[str]]:
    drafts = load_drafts()
    review_ids = [clean_text(item.get("review_id")) for item in items if clean_text(item.get("review_id"))]
    reviews, missing = _get_snapshot_reviews_by_ids(review_ids, force_refresh=False)
    review_map = {clean_text(review.get("id")): normalize_review(review) for review in reviews}
    saved = 0
    notes: List[str] = [f"{review_id}: отзыв не найден в локальном снимке. Обновите отзывы из WB и повторите." for review_id in missing]

    for payload in items:
        review_id = clean_text(payload.get("review_id"))
        if not review_id:
            continue
        review = review_map.get(review_id)
        if not review:
            continue
        reply_text = trim_reply(clean_text(payload.get("reply")), limit=MAX_REPLY_LENGTH)
        if len(reply_text) < MIN_REPLY_LENGTH:
            notes.append(f"{review_id}: текст ответа короче {MIN_REPLY_LENGTH} символов.")
            continue

        current = drafts.get(review_id, {})
        current.update(
            {
                "reply": reply_text,
                "signature": review_signature(review),
                "generated_at": utc_now_iso(),
                "source": clean_text(current.get("source") or "manual_edit"),
                "rule_ids": current.get("rule_ids", []),
                "cross_sell_items": current.get("cross_sell_items", []),
                "prompt_signature": current.get("prompt_signature", ""),
            }
        )
        drafts[review_id] = current
        saved += 1

    save_drafts(drafts)
    return saved, notes


def _queue_server_reply_drafts_by_ids(selected_ids: List[str]) -> Tuple[int, List[str]]:
    ids = [clean_text(x) for x in selected_ids if clean_text(x)]
    reviews, missing = _get_snapshot_reviews_by_ids(ids, force_refresh=False)
    drafts = load_drafts()
    reply_text_map: Dict[str, str] = {}
    notes: List[str] = [f"{review_id}: отзыв не найден в локальном снимке. Обновите отзывы из WB и повторите." for review_id in missing]

    valid_reviews: List[Dict[str, Any]] = []
    for review in reviews:
        review_id = clean_text(review.get("id"))
        draft = drafts.get(review_id)
        if not draft:
            notes.append(f"{review_id}: на сервере нет сохранённого черновика. Сначала подготовьте AI-черновик.")
            continue
        reply_text = trim_reply(clean_text(draft.get("reply")), limit=MAX_REPLY_LENGTH)
        if len(reply_text) < MIN_REPLY_LENGTH:
            notes.append(f"{review_id}: серверный черновик слишком короткий.")
            continue
        reply_text_map[review_id] = reply_text
        valid_reviews.append(review)

    if not valid_reviews:
        return 0, notes

    class _FormShim(dict):
        def getlist(self, key: str):
            value = self.get(key, [])
            return value if isinstance(value, list) else [value]

    shim = _FormShim()
    shim["selected_ids"] = [clean_text(r.get("id")) for r in valid_reviews]
    for review in valid_reviews:
        review_id = clean_text(review.get("id"))
        prefix = f"review__{review_id}__"
        shim[prefix + "text"] = clean_text(review.get("text"))
        shim[prefix + "pros"] = clean_text(review.get("pros"))
        shim[prefix + "cons"] = clean_text(review.get("cons"))
        shim[prefix + "stars"] = str(int(review.get("productValuation", 0) or 0))
        shim[prefix + "created_date"] = clean_text(review.get("createdDate"))
        shim[prefix + "user_name"] = clean_text(review.get("userName"))
        shim[prefix + "subject_name"] = clean_text(review.get("subjectName"))
        shim[prefix + "product_name"] = clean_text(review.get("productDetails", {}).get("productName"))
        shim[prefix + "supplier_article"] = clean_text(review.get("productDetails", {}).get("supplierArticle"))
        shim[prefix + "brand_name"] = clean_text(review.get("productDetails", {}).get("brandName"))
        shim[prefix + "nm_id"] = str(_safe_int(review.get("productDetails", {}).get("nmId")))
        shim[f"reply__{review_id}"] = reply_text_map[review_id]

    added, queue_notes = queue_selected_replies(shim)
    notes.extend(queue_notes)
    return added, notes


def get_reply_queue_index() -> Dict[str, Dict[str, Any]]:
    result: Dict[str, Dict[str, Any]] = {}
    for item in load_reply_queue():
        review_id = clean_text(item.get("review_id"))
        if review_id:
            result[review_id] = item
    return result


def find_matching_rules(review: Dict[str, Any], rules: Dict[str, Any]) -> List[Dict[str, Any]]:
    product_name = clean_text(review.get("productDetails", {}).get("productName")).lower()
    article = clean_text(review.get("productDetails", {}).get("supplierArticle")).lower()
    combined_text = " ".join(
        [
            clean_text(review.get("text")).lower(),
            clean_text(review.get("pros")).lower(),
            clean_text(review.get("cons")).lower(),
        ]
    )
    matched: List[Dict[str, Any]] = []
    for case in rules.get("special_cases", []):
        keywords = [clean_text(x).lower() for x in case.get("keywords", []) if clean_text(x)]
        product_keywords = [clean_text(x).lower() for x in case.get("product_keywords", []) if clean_text(x)]
        article_keywords = [clean_text(x).lower() for x in case.get("article_keywords", []) if clean_text(x)]
        keyword_ok = True if not keywords else any(k in combined_text for k in keywords)
        product_ok = True if not product_keywords else any(k in product_name for k in product_keywords)
        article_ok = True if not article_keywords else any(k in article for k in article_keywords)
        if keyword_ok and product_ok and article_ok:
            matched.append(case)
    return matched


def choose_cross_sell_items(review: Dict[str, Any], rules: Dict[str, Any], matched_rules: List[Dict[str, Any]]) -> List[str]:
    stars = int(review.get("productValuation", 0) or 0)
    combined = " ".join(
        [
            clean_text(review.get("productDetails", {}).get("productName")).lower(),
            clean_text(review.get("text")).lower(),
            clean_text(review.get("pros")).lower(),
            clean_text(review.get("cons")).lower(),
        ]
    )
    suggestions: List[str] = []
    for rule in matched_rules:
        for item in rule.get("cross_sell_articles", []):
            val = clean_text(item)
            if val and "ЗАМЕНИТЬ" not in val and val not in suggestions:
                suggestions.append(val)
    if stars >= 4:
        for item in rules.get("cross_sell_catalog", []):
            title = clean_text(item.get("title"))
            article = clean_text(item.get("article"))
            if not title or not article or "ЗАМЕНИТЬ" in article:
                continue
            tags = [clean_text(tag).lower() for tag in item.get("tags", []) if clean_text(tag)]
            if not tags or any(tag in combined for tag in tags):
                candidate = f"арт. {article} — {title}"
                if candidate not in suggestions:
                    suggestions.append(candidate)
    return suggestions[:8]


def trim_reply(text: str, limit: int = MAX_REPLY_LENGTH) -> str:
    text = clean_text(text)
    if len(text) <= limit:
        return text
    chunks = re.split(r"(?<=[.!?])\s+", text)
    result = ""
    for chunk in chunks:
        candidate = f"{result} {chunk}".strip()
        if len(candidate) > limit:
            break
        result = candidate
    if result:
        return result
    return text[: limit - 1].rstrip() + "…"


def normalize_page_size(value: Any, default: int = REPLY_DEFAULT_PAGE_SIZE) -> int:
    try:
        page_size = int(value or default)
    except Exception:
        page_size = default
    return page_size if page_size in PAGE_SIZE_OPTIONS else default


def _augment_system_prompt(prompt_text: str) -> str:
    base = common.clean_text_preserve_lines(prompt_text).strip() or "Пиши вежливо."
    return (
        base
        + "\n\nОбязательные технические требования для ответа на отзыв:"
        + f"\n- Ответ должен быть только на русском языке."
        + f"\n- Длина ответа должна быть строго от {MIN_REPLY_LENGTH} до {MAX_REPLY_LENGTH} символов."
        + "\n- Ответ должен быть персональным, учитывать товар, оценку и содержание отзыва."
        + "\n- Нельзя выдавать один и тот же шаблонный ответ для разных отзывов."
        + "\n- Если отзыв без текста, всё равно напиши содержательный персональный ответ на основе товара, бренда, категории и оценки."
        + "\n- Можно мягко рекомендовать связанные товары, если это предусмотрено инструкциями."
    )


def _expand_reply_if_needed(review: Dict[str, Any], system_prompt: str, user_prompt: str, draft_text: str) -> str:
    reply_text = trim_reply(draft_text, limit=MAX_REPLY_LENGTH)
    if len(reply_text) >= MIN_REPLY_LENGTH:
        return reply_text

    expanded = call_ai(
        [
            {"role": "system", "content": system_prompt},
            {
                "role": "user",
                "content": (
                    user_prompt
                    + "\n\nНиже черновик ответа, который получился слишком коротким:"
                    + f"\n{reply_text}"
                    + f"\n\nПерепиши и разверни ответ так, чтобы итоговая длина была от {MIN_REPLY_LENGTH} до {MAX_REPLY_LENGTH} символов. "
                    + "Ответ должен оставаться естественным, вежливым, персональным и полезным для покупателя."
                ),
            },
        ],
        model=OPENAI_MODEL,
        temperature=0.55,
    )
    return trim_reply(expanded, limit=MAX_REPLY_LENGTH)


def prompt_signature(text: str) -> str:
    return hashlib.sha256(clean_text(text).encode('utf-8')).hexdigest()


def is_draft_compatible(draft_entry: Dict[str, Any] | None, review_sig: str, current_prompt_sig: str) -> bool:
    if not draft_entry:
        return False
    if clean_text(draft_entry.get('signature')) != clean_text(review_sig):
        return False
    source = clean_text(draft_entry.get('source')).lower()
    if source == 'shortcut':
        return False
    if source == 'ai':
        return clean_text(draft_entry.get('prompt_signature')) == clean_text(current_prompt_sig)
    return source in {'manual', 'manual_edit', 'queued'}


def build_prompt_context(review: Dict[str, Any], rules: Dict[str, Any], matched_rules: List[Dict[str, Any]], cross_sell_items: List[str]) -> str:
    review_text = build_review_text(review)
    product = review.get("productDetails", {})
    business_notes = [f"- {clean_text(x)}" for x in rules.get("default_instructions", []) if clean_text(x)]
    case_notes: List[str] = []
    for case in matched_rules:
        title = clean_text(case.get("title") or case.get("id"))
        instruction = clean_text(case.get("instruction"))
        if instruction:
            case_notes.append(f"- {title}: {instruction}")
    cross_sell_text = "; ".join(cross_sell_items) if cross_sell_items else "нет"
    return f"""
Данные по отзыву:
- Товар: {clean_text(product.get('productName'))}
- Артикул продавца: {clean_text(product.get('supplierArticle'))}
- Бренд: {clean_text(product.get('brandName'))}
- Оценка: {int(review.get('productValuation', 0) or 0)}
- Покупатель: {clean_text(review.get('userName'))}
- Дата: {clean_text(review.get('createdDate'))}

Отзыв покупателя:
{review_text}

Базовые инструкции бизнеса:
{os.linesep.join(business_notes) if business_notes else '- Нет дополнительных базовых инструкций'}

Специальные инструкции для этого случая:
{os.linesep.join(case_notes) if case_notes else '- Для этого отзыва спец-инструкции не сработали'}

Артикулы для мягкой рекомендации:
{cross_sell_text}

Сформируй только готовый ответ покупателю без пояснений для менеджера.
Важно: даже если покупатель поставил только оценку без текста, ответ должен быть персональным, учитывать товар и не повторять дословно один и тот же шаблон для разных товаров.
Длина ответа должна быть от 1500 до 4500 символов.
Если в системном промпте указаны рекомендации по связанным товарам, встрои их органично в основной текст ответа.
""".strip()


def generate_shortcut_reply(review: Dict[str, Any], matched_rules: List[Dict[str, Any]], cross_sell_items: List[str]) -> str | None:
    stars = int(review.get("productValuation", 0) or 0)
    review_text = build_review_text(review)
    if review_text == "Покупатель поставил оценку без текста." and stars >= 5:
        reply = "Спасибо за высокую оценку! Очень рады, что товар вам понравился. Будем рады видеть вас снова."
        if cross_sell_items:
            reply += f" Также вам может понравиться {cross_sell_items[0]}."
        return trim_reply(reply)
    if review_text == "Покупатель поставил оценку без текста." and stars <= 3:
        return trim_reply(
            "Спасибо за оценку. Нам жаль, что впечатление оказалось неидеальным. Если напишете, что именно не понравилось, мы обязательно поможем и всё подскажем."
        )
    for rule in matched_rules:
        fixed_reply = clean_text(rule.get("fixed_reply"))
        if fixed_reply:
            reply = fixed_reply
            if stars >= 4 and cross_sell_items:
                reply += f" Также вам может понравиться {cross_sell_items[0]}."
            return trim_reply(reply)
    return None


def generate_reply_for_review(review: Dict[str, Any], force: bool = False) -> Dict[str, Any]:
    review = normalize_review(review)
    review_id = clean_text(review.get("id"))
    signature = review_signature(review)
    drafts = load_drafts()
    prompt_text = get_prompt_text()
    current_prompt_sig = prompt_signature(prompt_text)
    cached = drafts.get(review_id)
    if not force and is_draft_compatible(cached, signature, current_prompt_sig):
        return cached

    rules = load_rules()
    matched_rules = find_matching_rules(review, rules)
    cross_sell_items = choose_cross_sell_items(review, rules, matched_rules)
    user_prompt = build_prompt_context(review, rules, matched_rules, cross_sell_items)
    system_prompt = _augment_system_prompt(prompt_text)
    raw_reply = call_ai(
        [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        model=OPENAI_MODEL,
        temperature=0.6,
    )
    reply_text = _expand_reply_if_needed(review, system_prompt, user_prompt, raw_reply)
    reply_text = trim_reply(reply_text, limit=MAX_REPLY_LENGTH)
    entry = {
        "reply": reply_text,
        "signature": signature,
        "generated_at": utc_now_iso(),
        "rule_ids": [clean_text(x.get("id")) for x in matched_rules],
        "cross_sell_items": cross_sell_items,
        "source": "ai",
        "prompt_signature": current_prompt_sig,
    }
    drafts[review_id] = entry
    save_drafts(drafts)
    return entry


def send_reply_to_wb(review_id: str, reply_text: str) -> None:
    payload = {"id": review_id, "text": reply_text}
    response = WB_SESSION.post(WB_ANSWER_URL, json=payload, timeout=REQUEST_TIMEOUT)
    if response.status_code == 204:
        return
    detail = ""
    try:
        payload = response.json()
        detail = clean_text(payload.get("detail") or payload.get("errorText") or payload)
    except Exception:
        detail = clean_text(response.text)
    raise RuntimeError(f"WB API не принял ответ: HTTP {response.status_code}. {detail}")


def upsert_archive_record(record: Dict[str, Any]) -> None:
    archive = load_archive()
    record_id = clean_text(record.get("id"))
    updated = False
    for idx, item in enumerate(archive):
        if clean_text(item.get("id")) == record_id:
            archive[idx] = {**item, **record}
            updated = True
            break
    if not updated:
        archive.append(record)
    save_archive(archive)


def remove_draft(review_id: str) -> None:
    drafts = load_drafts()
    if review_id in drafts:
        del drafts[review_id]
        save_drafts(drafts)


def get_reply_queue_status_map() -> Dict[str, Dict[str, Any]]:
    mapping: Dict[str, Dict[str, Any]] = {}
    for item in load_reply_queue():
        review_id = clean_text(item.get("review_id"))
        if review_id:
            mapping[review_id] = item
    return mapping


def _reply_matches_text(row: Dict[str, Any], query: str) -> bool:
    query = clean_text(query).lower()
    if not query:
        return True
    haystack = " ".join(
        [
            clean_text(row.get("id")),
            clean_text(row.get("product_name")),
            clean_text(row.get("supplier_article")),
            clean_text(row.get("brand_name")),
            clean_text(row.get("user_name")),
            clean_text(row.get("subject_name")),
            clean_text(row.get("review_text")),
            clean_text(row.get("reply")),
            clean_text(row.get("text")),
            clean_text(row.get("pros")),
            clean_text(row.get("cons")),
            str(row.get("nm_id") or ""),
        ]
    ).lower()
    return query in haystack


def _row_reply_is_sent(row: Dict[str, Any]) -> bool:
    queue_status = clean_text(row.get("queue_status"))
    archive_status = clean_text(row.get("archive_status"))
    return queue_status in REPLY_SENT_STATUSES or archive_status in REPLY_SENT_STATUSES


def _sort_reply_rows(rows: List[Dict[str, Any]], sort_by: str) -> List[Dict[str, Any]]:
    sort_by = clean_text(sort_by) or "newest"
    if sort_by == "oldest":
        return sorted(rows, key=lambda r: _parse_created_date(r.get("created_date")))
    if sort_by == "stars_low":
        return sorted(rows, key=lambda r: (int(r.get("stars", 0) or 0), _parse_created_date(r.get("created_date"))))
    if sort_by == "stars_high":
        return sorted(rows, key=lambda r: (-int(r.get("stars", 0) or 0), _parse_created_date(r.get("created_date"))))
    return sorted(rows, key=lambda r: _parse_created_date(r.get("created_date")), reverse=True)


def _row_reply_matches_filters(
    row: Dict[str, Any],
    stars_filter: str = "all",
    draft_filter: str = "all",
    queue_filter: str = "all",
    search_query: str = "",
    hide_sent: bool = True,
) -> bool:
    stars = int(row.get("stars", 0) or 0)
    if stars_filter in {"1", "2", "3", "4", "5"} and stars != int(stars_filter):
        return False
    has_draft = bool(clean_text(row.get("reply")))
    if draft_filter == "with_draft" and not has_draft:
        return False
    if draft_filter == "without_draft" and has_draft:
        return False
    queue_status = clean_text(row.get("queue_status"))
    if queue_filter == "ready" and (not has_draft or queue_status in {"queued", "processing"}):
        return False
    if queue_filter == "queued" and queue_status != "queued":
        return False
    if queue_filter == "processing" and queue_status != "processing":
        return False
    if queue_filter == "failed" and queue_status != "failed":
        return False
    if queue_filter == "sent" and not _row_reply_is_sent(row):
        return False
    if queue_filter == "not_queued" and queue_status in {"queued", "processing"}:
        return False
    if hide_sent and _row_reply_is_sent(row):
        return False
    if not _reply_matches_text(row, search_query):
        return False
    return True


def build_reply_rows(
    page: int = 1,
    page_size: int = REPLY_DEFAULT_PAGE_SIZE,
    sort_by: str = "newest",
    stars_filter: str = "all",
    draft_filter: str = "all",
    queue_filter: str = "all",
    search_query: str = "",
    hide_sent: bool = True,
    source_filter: str = "all",
    content_filter: str = "all",
    answer_state: str = "needs_reply",
    force_refresh: bool = False,
) -> Dict[str, Any]:
    page = max(page, 1)
    if force_refresh:
        try:
            get_reply_snapshot(force_refresh=True)
        except Exception:
            pass

    drafts = load_drafts()
    current_prompt_sig = prompt_signature(get_prompt_text())
    queue_map = get_reply_queue_status_map()
    archive = load_archive()
    archive_map = {clean_text(item.get("id")): item for item in archive}

    if history_service.db_has_data(getattr(g, "active_tenant_id", "") or common.get_active_tenant_id()):
        rows_db = history_service.list_rows(
            tenant_id=getattr(g, "active_tenant_id", "") or common.get_active_tenant_id(),
            sort_by=sort_by,
            stars_filter=stars_filter,
            search_query=search_query,
            source_filter=source_filter,
            content_filter=content_filter,
            answer_state=answer_state,
        )
        all_rows: List[Dict[str, Any]] = []
        submitted_hidden = 0
        for db_row in rows_db:
            review_id = clean_text(db_row.get("review_id"))
            review_stub = _history_row_to_review(db_row)
            signature = review_signature(review_stub)
            draft_entry = drafts.get(review_id)
            reply_text = ""
            if is_draft_compatible(draft_entry, signature, current_prompt_sig):
                reply_text = clean_text(draft_entry.get("reply"))
            queue_entry = queue_map.get(review_id, {})
            archive_entry = archive_map.get(review_id, {})
            row = {
                "id": review_id,
                "product_name": clean_text(db_row.get("product_name")),
                "supplier_article": clean_text(db_row.get("supplier_article")),
                "brand_name": clean_text(db_row.get("brand_name")),
                "nm_id": _safe_int(db_row.get("nm_id")),
                "stars": _safe_int(db_row.get("stars")),
                "stars_view": "⭐" * _safe_int(db_row.get("stars")),
                "review_text": clean_text(db_row.get("review_text")),
                "reply": reply_text,
                "has_draft": bool(reply_text),
                "user_name": clean_text(db_row.get("user_name")),
                "created_date": clean_text(db_row.get("created_date")),
                "subject_name": clean_text(db_row.get("subject_name")),
                "text": clean_text(db_row.get("text")),
                "pros": clean_text(db_row.get("pros")),
                "cons": clean_text(db_row.get("cons")),
                "queue_status": clean_text(queue_entry.get("status")),
                "queue_error": clean_text(queue_entry.get("error")),
                "queue_sent_at": clean_text(queue_entry.get("sent_at")),
                "archive_status": clean_text(archive_entry.get("status")),
                "archive_sent_at": clean_text(archive_entry.get("sent_at")),
                "source": clean_text(db_row.get("source")),
                "is_empty_rating_only": bool(db_row.get("is_empty_rating_only")),
                "has_wb_answer": bool(clean_text(db_row.get("answer_text"))),
            }
            if hide_sent and _row_reply_is_sent(row):
                submitted_hidden += 1
                continue
            if _row_reply_matches_filters(
                row,
                stars_filter=stars_filter,
                draft_filter=draft_filter,
                queue_filter=queue_filter,
                search_query=search_query,
                hide_sent=False,
            ):
                all_rows.append(row)
        total_filtered = len(all_rows)
        page_count = max(1, (total_filtered + page_size - 1) // page_size) if total_filtered else 1
        if page > page_count:
            page = page_count
        start = (page - 1) * page_size
        end = start + page_size
        page_rows = all_rows[start:end]
        counts = history_service.get_counts(getattr(g, "active_tenant_id", "") or common.get_active_tenant_id())
        meta = history_service.effective_meta(getattr(g, "active_tenant_id", "") or common.get_active_tenant_id())
        return {
            "rows": page_rows,
            "page": page,
            "page_size": page_size,
            "page_count": page_count,
            "has_prev": page > 1,
            "has_next": page < page_count,
            "total_filtered": total_filtered,
            "submitted_hidden": submitted_hidden,
            "raw_scanned": _safe_int(counts.get("total")),
            "count_unanswered": _safe_int(counts.get("needs_reply")),
            "count_archive": _safe_int(counts.get("archive")),
            "draft_count": sum(1 for row in all_rows if row.get("has_draft")),
            "queue_total": sum(1 for item in queue_map.values() if clean_text(item.get("status")) in {"queued", "processing", "failed"}),
            "filters": {
                "sort": sort_by or "newest",
                "stars": stars_filter or "all",
                "draft": draft_filter or "all",
                "queue": queue_filter or "all",
                "q": search_query or "",
                "hide_sent": bool(hide_sent),
                "source": source_filter or "all",
                "content": content_filter or "all",
                "answer_state": answer_state or "needs_reply",
                "page_size": page_size,
            },
            "snapshot_fetched_at": clean_text(meta.get("finished_at") or meta.get("last_active_snapshot_at")),
            "auth_required": False,
            "history_counts": counts,
            "history_meta": meta,
            "using_history_db": True,
        }

    snapshot = get_reply_snapshot(force_refresh=force_refresh)
    feedbacks = snapshot.get("feedbacks", []) if isinstance(snapshot.get("feedbacks", []), list) else []

    all_rows: List[Dict[str, Any]] = []
    submitted_hidden = 0
    for review in feedbacks:
        review = normalize_review(review)
        review_id = clean_text(review.get("id"))
        signature = review_signature(review)
        draft_entry = drafts.get(review_id)
        reply_text = ""
        if is_draft_compatible(draft_entry, signature, current_prompt_sig):
            reply_text = clean_text(draft_entry.get("reply"))
        product = review.get("productDetails", {}) or {}
        queue_entry = queue_map.get(review_id, {})
        archive_entry = archive_map.get(review_id, {})
        row = {
            "id": review_id,
            "product_name": clean_text(product.get("productName")),
            "supplier_article": clean_text(product.get("supplierArticle")),
            "brand_name": clean_text(product.get("brandName")),
            "nm_id": _safe_int(product.get("nmId")),
            "stars": _safe_int(review.get("productValuation")),
            "stars_view": "⭐" * _safe_int(review.get("productValuation")),
            "review_text": build_review_text(review),
            "reply": reply_text,
            "has_draft": bool(reply_text),
            "user_name": clean_text(review.get("userName")),
            "created_date": clean_text(review.get("createdDate")),
            "subject_name": clean_text(review.get("subjectName")),
            "text": clean_text(review.get("text")),
            "pros": clean_text(review.get("pros")),
            "cons": clean_text(review.get("cons")),
            "queue_status": clean_text(queue_entry.get("status")),
            "queue_error": clean_text(queue_entry.get("error")),
            "queue_sent_at": clean_text(queue_entry.get("sent_at")),
            "archive_status": clean_text(archive_entry.get("status")),
            "archive_sent_at": clean_text(archive_entry.get("sent_at")),
            "source": "active",
            "is_empty_rating_only": False,
            "has_wb_answer": False,
        }
        if hide_sent and _row_reply_is_sent(row):
            submitted_hidden += 1
        if _row_reply_matches_filters(
            row,
            stars_filter=stars_filter,
            draft_filter=draft_filter,
            queue_filter=queue_filter,
            search_query=search_query,
            hide_sent=hide_sent,
        ):
            all_rows.append(row)

    filtered_rows = _sort_reply_rows(all_rows, sort_by)
    total_filtered = len(filtered_rows)
    page_count = max(1, (total_filtered + page_size - 1) // page_size) if total_filtered else 1
    if page > page_count:
        page = page_count
    start = (page - 1) * page_size
    end = start + page_size
    page_rows = filtered_rows[start:end]
    draft_count = sum(1 for row in filtered_rows if row.get("has_draft"))
    queue_total = sum(1 for item in queue_map.values() if clean_text(item.get("status")) in {"queued", "processing", "failed"})
    return {
        "rows": page_rows,
        "page": page,
        "page_size": page_size,
        "page_count": page_count,
        "has_prev": page > 1,
        "has_next": page < page_count,
        "total_filtered": total_filtered,
        "submitted_hidden": submitted_hidden,
        "raw_scanned": _safe_int(snapshot.get("raw_scanned")),
        "count_unanswered": _safe_int(snapshot.get("count_unanswered")),
        "count_archive": _safe_int(snapshot.get("count_archive")),
        "draft_count": draft_count,
        "queue_total": queue_total,
        "filters": {
            "sort": sort_by or "newest",
            "stars": stars_filter or "all",
            "draft": draft_filter or "all",
            "queue": queue_filter or "all",
            "q": search_query or "",
            "hide_sent": bool(hide_sent),
            "source": source_filter or "all",
            "content": content_filter or "all",
            "answer_state": answer_state or "needs_reply",
            "page_size": page_size,
        },
        "snapshot_fetched_at": clean_text(snapshot.get("fetched_at")),
        "auth_required": False,
        "history_counts": history_service.get_counts(getattr(g, "active_tenant_id", "") or common.get_active_tenant_id()),
        "history_meta": history_service.effective_meta(getattr(g, "active_tenant_id", "") or common.get_active_tenant_id()),
        "using_history_db": False,
    }


def build_analytics() -> Dict[str, Any]:
    archive = load_archive()
    rules = load_rules()
    total_sent = sum(1 for item in archive if clean_text(item.get("status")) in REPLY_SENT_STATUSES)
    stars_counter = Counter(int(item.get("stars", 0) or 0) for item in archive if item.get("stars") is not None)
    products_counter = Counter(clean_text(item.get("product_name")) for item in archive if clean_text(item.get("product_name")))
    source_counter = Counter(clean_text(item.get("reply_source")) for item in archive if clean_text(item.get("reply_source")))
    article_counter = Counter(clean_text(item.get("supplier_article")) for item in archive if clean_text(item.get("supplier_article")))
    issue_counter: Counter[str] = Counter()
    for item in archive:
        stars = int(item.get("stars", 0) or 0)
        if stars > 3:
            continue
        fake_review = {
            "text": item.get("raw_text", ""),
            "pros": item.get("raw_pros", ""),
            "cons": item.get("raw_cons", ""),
            "productValuation": stars,
            "productDetails": {
                "productName": item.get("product_name", ""),
                "supplierArticle": item.get("supplier_article", ""),
            },
        }
        matched = find_matching_rules(fake_review, rules)
        for rule in matched:
            issue_counter[clean_text(rule.get("title") or rule.get("id"))] += 1
    recent = sorted(archive, key=lambda x: clean_text(x.get("sent_at")), reverse=True)[:20]
    star_values = [int(item.get("stars", 0) or 0) for item in archive if item.get("stars") is not None]
    avg_stars = round(sum(star_values) / len(star_values), 2) if star_values else 0.0
    return {
        "total_sent": total_sent,
        "avg_stars": avg_stars,
        "stars": dict(sorted(stars_counter.items())),
        "top_products": products_counter.most_common(10),
        "top_articles": article_counter.most_common(10),
        "top_issues": issue_counter.most_common(10),
        "reply_sources": source_counter.most_common(),
        "recent": recent,
    }


def _collect_selected_reviews_from_form(form: Any) -> List[Dict[str, Any]]:
    selected_ids = [clean_text(x) for x in form.getlist("selected_ids") if clean_text(x)]
    reviews: List[Dict[str, Any]] = []
    for review_id in selected_ids:
        prefix = f"review__{review_id}__"
        reviews.append(
            normalize_review(
                {
                    "id": review_id,
                    "text": form.get(prefix + "text", ""),
                    "pros": form.get(prefix + "pros", ""),
                    "cons": form.get(prefix + "cons", ""),
                    "productValuation": int(form.get(prefix + "stars", 0) or 0),
                    "createdDate": form.get(prefix + "created_date", ""),
                    "userName": form.get(prefix + "user_name", ""),
                    "subjectName": form.get(prefix + "subject_name", ""),
                    "productDetails": {
                        "productName": form.get(prefix + "product_name", ""),
                        "supplierArticle": form.get(prefix + "supplier_article", ""),
                        "brandName": form.get(prefix + "brand_name", ""),
                        "nmId": int(form.get(prefix + "nm_id", 0) or 0),
                    },
                }
            )
        )
    return reviews


def _reviews_params_from_source(src: Any) -> Dict[str, Any]:
    def _as_bool(value: Any) -> bool:
        return str(value).lower() in {"1", "true", "on", "yes"}

    params = {
        "page": max(1, _safe_int(src.get("page", 1), 1)),
        "page_size": normalize_page_size(src.get("page_size", REPLY_DEFAULT_PAGE_SIZE), REPLY_DEFAULT_PAGE_SIZE),
        "sort": clean_text(src.get("sort")) or "newest",
        "stars": clean_text(src.get("stars")) or "all",
        "draft": clean_text(src.get("draft")) or "all",
        "queue": clean_text(src.get("queue")) or "all",
        "q": clean_text(src.get("q")),
        "hide_sent": 1 if _as_bool(src.get("hide_sent", 1)) else 0,
        "source": clean_text(src.get("source")) or "all",
        "content": clean_text(src.get("content")) or "all",
        "answer_state": clean_text(src.get("answer_state")) or "needs_reply",
    }
    params.update(_tenant_query_arg())
    return params


def _complaints_params_from_source(src: Any) -> Dict[str, Any]:
    def _as_bool(value: Any) -> bool:
        return str(value).lower() in {"1", "true", "on", "yes"}

    params = {
        "page": max(1, int(src.get("page", 1) or 1)),
        "page_size": normalize_page_size(src.get("page_size", COMPLAINT_DEFAULT_PAGE_SIZE), COMPLAINT_DEFAULT_PAGE_SIZE),
        "sort": clean_text(src.get("sort")) or "newest",
        "stars": clean_text(src.get("stars")) or "all",
        "draft": clean_text(src.get("draft")) or "all",
        "queue": clean_text(src.get("queue")) or "all",
        "q": clean_text(src.get("q")),
        "hide_submitted": 1 if _as_bool(src.get("hide_submitted", 1)) else 0,
    }
    params.update(_tenant_query_arg())
    return params


def _questions_params_from_source(src: Any) -> Dict[str, Any]:
    def _as_bool(value: Any) -> bool:
        return str(value).lower() in {"1", "true", "on", "yes"}

    page_size = _safe_int(src.get("page_size", 100), 100)
    if page_size not in QUESTION_PAGE_SIZE_OPTIONS:
        page_size = 100 if 100 in QUESTION_PAGE_SIZE_OPTIONS else QUESTION_PAGE_SIZE_OPTIONS[0]
    cluster = clean_text(src.get("cluster") or src.get("cluster_key"))
    mode = clean_text(src.get("mode")) or "questions"
    if mode not in {"clusters", "questions"}:
        mode = "questions"
    params = {
        "page": max(1, _safe_int(src.get("page", 1), 1)),
        "page_size": page_size,
        "sort": clean_text(src.get("sort")) or "cluster",
        "draft": clean_text(src.get("draft")) or "all",
        "queue": clean_text(src.get("queue")) or "all",
        "q": clean_text(src.get("q")),
        "hide_submitted": 1 if _as_bool(src.get("hide_submitted", 1)) else 0,
        "cluster": cluster,
        "mode": mode,
    }
    params.update(_tenant_query_arg())
    return params


def _resolve_question_ids(form: Any, allow_cluster_scope: bool = False) -> List[str]:
    selected_ids = [clean_text(x) for x in form.getlist("selected_ids") if clean_text(x)]
    if selected_ids:
        return selected_ids
    if not allow_cluster_scope:
        return []
    scope = clean_text(form.get("scope")).lower()
    if scope not in {"cluster", "all_cluster"}:
        return []
    cluster_key = clean_text(form.get("cluster_key") or form.get("cluster"))
    if not cluster_key:
        return []
    return list_question_ids_for_cluster(cluster_key, hide_submitted=False)


QUESTION_CLUSTERING_EXPORT_COLUMNS: List[Tuple[str, str, str]] = [
    ("tenant_id", "tenant_id", "ID кабинета, чтобы потом вернуть кластер в нужный tenant."),
    ("tenant_name", "tenant_name", "Человеческое название кабинета."),
    ("tenant_question_key", "tenant_question_key", "Стабильный внешний ключ строки tenant_id::question_id."),
    ("question_id", "question_id", "ID вопроса в WB внутри кабинета."),
    ("created_at", "created_at", "Дата создания вопроса в WB (ISO)."),
    ("current_status", "current_status", "Текущий статус выгрузки. Сейчас всегда unanswered."),
    ("question_text", "question_text", "Исходный текст вопроса покупателя без сокращений."),
    ("normalized_question", "normalized_question", "Нормализованный текст вопроса для будущей кластеризации."),
    ("product_group_hint", "product_group_hint", "Нормализованная товарная группа. Строится в первую очередь по названию товара, а затем по артикулу."),
    ("article_group_hint", "article_group_hint", "Очищенная группа артикула продавца без цифр и лишних разделителей."),
    ("product_name", "product_name", "Название товара из WB."),
    ("supplier_article_raw", "supplier_article_raw", "Исходный артикул продавца как есть."),
    ("supplier_article_norm", "supplier_article_norm", "Нормализованный артикул продавца для технических сверок."),
    ("nm_id", "nm_id", "nmID товара WB."),
    ("brand_name", "brand_name", "Бренд товара."),
    ("subject_name", "subject_name", "Категория / subject товара, если WB отдал это поле."),
    ("size", "size", "Размер / вариант, если WB отдал значение."),
    ("wb_state", "wb_state", "Состояние вопроса по данным WB."),
    ("was_viewed", "was_viewed", "Был ли вопрос помечен WB как просмотренный."),
    ("is_warned", "is_warned", "Есть ли предупреждение от WB по вопросу."),
    ("snapshot_fetched_at", "snapshot_fetched_at", "Когда был получен снапшот, из которого выгрузили строку."),
]


def _question_clustering_sort_key(row: Dict[str, Any]) -> Tuple[str, str, str, str, str]:
    return (
        clean_text(row.get("product_group_hint")),
        clean_text(row.get("normalized_question")) or clean_text(row.get("question_text")),
        clean_text(row.get("tenant_id")),
        clean_text(row.get("created_at")),
        clean_text(row.get("question_id")),
    )


def _collect_question_clustering_rows(refresh_all: bool = False) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    tenants = [tenant for tenant in tenant_manager.load_tenants() if tenant.get("enabled", True)]
    rows: List[Dict[str, Any]] = []
    summary_rows: List[Dict[str, Any]] = []
    warnings: List[str] = []
    seen_keys: set[str] = set()
    original_tenant_id = clean_text(getattr(g, "active_tenant_id", "") or session.get("tenant_id"))

    try:
        for tenant in tenants:
            tenant_id = clean_text(tenant.get("id"))
            tenant_name = clean_text(tenant.get("name"))
            if not tenant_id:
                continue
            try:
                with _temporary_tenant_binding(tenant_id):
                    snapshot = load_question_snapshot()
                    if refresh_all or not isinstance(snapshot, dict) or not (snapshot.get("questions") or []):
                        snapshot = get_question_snapshot(force_refresh=True)
                fetched_at = clean_text((snapshot or {}).get("fetched_at"))
                raw_questions = (snapshot or {}).get("questions") or []
                tenant_rows: List[Dict[str, Any]] = []
                for item in raw_questions:
                    row = build_question_clustering_export_row(
                        item,
                        tenant_id=tenant_id,
                        tenant_name=tenant_name,
                        snapshot_fetched_at=fetched_at,
                    )
                    stable_key = clean_text(row.get("tenant_question_key"))
                    if not clean_text(row.get("question_id")) or not stable_key or stable_key in seen_keys:
                        continue
                    seen_keys.add(stable_key)
                    tenant_rows.append(row)
                tenant_rows.sort(key=_question_clustering_sort_key)
                rows.extend(tenant_rows)
                summary_rows.append(
                    {
                        "tenant_id": tenant_id,
                        "tenant_name": tenant_name,
                        "question_count": len(tenant_rows),
                        "count_unanswered": int((snapshot or {}).get("count_unanswered") or len(tenant_rows)),
                        "raw_scanned": int((snapshot or {}).get("raw_scanned") or len(tenant_rows)),
                        "snapshot_fetched_at": fetched_at,
                        "truncated": bool((snapshot or {}).get("truncated")),
                        "has_new_questions": bool((snapshot or {}).get("has_new_questions")),
                        "status": "ok",
                        "error": "",
                    }
                )
            except Exception as exc:
                warning = f"{tenant_name or tenant_id}: {exc}"
                warnings.append(warning)
                summary_rows.append(
                    {
                        "tenant_id": tenant_id,
                        "tenant_name": tenant_name,
                        "question_count": 0,
                        "count_unanswered": 0,
                        "raw_scanned": 0,
                        "snapshot_fetched_at": "",
                        "truncated": False,
                        "has_new_questions": False,
                        "status": "error",
                        "error": str(exc),
                    }
                )
    finally:
        restore_id = original_tenant_id or clean_text((tenants[0] if tenants else {}).get("id"))
        if restore_id and restore_id != clean_text(getattr(g, "active_tenant_id", "")):
            try:
                tenant = tenant_manager.get_tenant(restore_id)
                g.active_tenant = tenant
                g.active_tenant_id = restore_id
            except Exception:
                pass

    rows.sort(key=_question_clustering_sort_key)
    meta = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "refresh_all": bool(refresh_all),
        "tenant_count": len([row for row in summary_rows if clean_text(row.get("status")) == "ok"]),
        "question_count": len(rows),
        "truncated_tenants": len([row for row in summary_rows if row.get("truncated")]),
        "warnings": warnings,
    }
    return rows, summary_rows, meta


def _style_export_header(cell: Any) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _fit_export_columns(ws: Any, max_width: int = 64) -> None:
    long_columns = {"G", "H", "I", "J", "K", "L", "M"}
    for idx, column_cells in enumerate(ws.columns, start=1):
        letter = get_column_letter(idx)
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if value:
                max_len = max(max_len, len(value))
            cell.alignment = Alignment(vertical="top", wrap_text=letter in long_columns)
        width = min(max(max_len + 2, 12), max_width if letter in long_columns else 28)
        ws.column_dimensions[letter].width = width


def _build_question_clustering_workbook(rows: List[Dict[str, Any]], summary_rows: List[Dict[str, Any]], meta: Dict[str, Any]) -> bytes:
    wb = Workbook()
    raw_ws = wb.active
    raw_ws.title = "questions_raw"
    raw_headers = [item[0] for item in QUESTION_CLUSTERING_EXPORT_COLUMNS]
    raw_ws.append(raw_headers)
    for cell in raw_ws[1]:
        _style_export_header(cell)
    raw_ws.freeze_panes = "A2"
    for row in rows:
        raw_ws.append([row.get(key) for key, _, _ in QUESTION_CLUSTERING_EXPORT_COLUMNS])
    raw_ws.auto_filter.ref = raw_ws.dimensions
    _fit_export_columns(raw_ws, max_width=72)
    for cell in raw_ws[1]:
        _style_export_header(cell)

    summary_ws = wb.create_sheet("summary")
    summary_ws.append(["metric", "value"])
    for cell in summary_ws[1]:
        _style_export_header(cell)
    summary_pairs = [
        ("generated_at", meta.get("generated_at")),
        ("refresh_all", "yes" if meta.get("refresh_all") else "no"),
        ("tenant_count", meta.get("tenant_count")),
        ("question_count", meta.get("question_count")),
        ("truncated_tenants", meta.get("truncated_tenants")),
        ("warnings", " | ".join(meta.get("warnings") or [])),
    ]
    for metric, value in summary_pairs:
        summary_ws.append([metric, value])
    summary_ws.append([])
    summary_headers = ["tenant_id", "tenant_name", "question_count", "count_unanswered", "raw_scanned", "snapshot_fetched_at", "truncated", "has_new_questions", "status", "error"]
    summary_ws.append(summary_headers)
    for cell in summary_ws[8]:
        _style_export_header(cell)
    for row in summary_rows:
        summary_ws.append([row.get(key) for key in summary_headers])
    summary_ws.freeze_panes = "A2"
    _fit_export_columns(summary_ws, max_width=60)
    for cell in summary_ws[1]:
        _style_export_header(cell)
    for cell in summary_ws[8]:
        _style_export_header(cell)

    guide_ws = wb.create_sheet("column_guide")
    guide_ws.append(["column_name", "label", "meaning"])
    for cell in guide_ws[1]:
        _style_export_header(cell)
    for column_name, label, meaning in QUESTION_CLUSTERING_EXPORT_COLUMNS:
        guide_ws.append([column_name, label, meaning])
    guide_ws.freeze_panes = "A2"
    _fit_export_columns(guide_ws, max_width=80)
    for cell in guide_ws[1]:
        _style_export_header(cell)

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _build_question_clustering_csv(rows: List[Dict[str, Any]]) -> bytes:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=[item[0] for item in QUESTION_CLUSTERING_EXPORT_COLUMNS])
    writer.writeheader()
    for row in rows:
        writer.writerow({key: row.get(key) for key, _, _ in QUESTION_CLUSTERING_EXPORT_COLUMNS})
    return output.getvalue().encode("utf-8-sig")


def _build_question_clustering_manifest(summary_rows: List[Dict[str, Any]], meta: Dict[str, Any]) -> bytes:
    payload = {
        **meta,
        "columns": [item[0] for item in QUESTION_CLUSTERING_EXPORT_COLUMNS],
        "tenants": summary_rows,
    }
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def _build_question_clustering_zip(rows: List[Dict[str, Any]], summary_rows: List[Dict[str, Any]], meta: Dict[str, Any]) -> Tuple[io.BytesIO, str]:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"questions_for_clustering_all_tenants_{stamp}"
    zip_stream = io.BytesIO()
    with zipfile.ZipFile(zip_stream, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"{base_name}.xlsx", _build_question_clustering_workbook(rows, summary_rows, meta))
        zf.writestr(f"{base_name}.csv", _build_question_clustering_csv(rows))
        zf.writestr(f"{base_name}_manifest.json", _build_question_clustering_manifest(summary_rows, meta))
    zip_stream.seek(0)
    return zip_stream, f"{base_name}.zip"


QUESTION_CLUSTER_IMPORT_ALIASES: Dict[str, str] = {
    "tenant_id": "tenant_id",
    "tenant": "tenant_id",
    "cabinet": "tenant_id",
    "cabinet_id": "tenant_id",
    "tenant_name": "tenant_name",
    "cabinet_name": "tenant_name",
    "tenant_question_key": "tenant_question_key",
    "stable_key": "tenant_question_key",
    "external_key": "tenant_question_key",
    "question_id": "question_id",
    "wb_question_id": "question_id",
    "questionid": "question_id",
    "cluster_key": "cluster_key",
    "cluster_id": "cluster_key",
    "cluster_code": "cluster_key",
    "cluster_title": "cluster_title",
    "cluster_name": "cluster_title",
    "cluster": "cluster_title",
    "cluster_order": "cluster_order",
    "sort_order": "cluster_order",
    "cluster_rank": "cluster_order",
    "cluster_position": "cluster_order",
    "question_text": "question_text",
    "normalized_question": "normalized_question",
    "product_group_hint": "product_group_hint",
    "article_group_hint": "article_group_hint",
    "product_name": "product_name",
    "supplier_article_raw": "supplier_article_raw",
    "nm_id": "nm_id",
}



def _normalize_cluster_import_header(value: Any) -> str:
    text = clean_text(value).lower().replace("ё", "е")
    text = re.sub(r"[^0-9a-zа-я]+", "_", text, flags=re.I).strip("_")
    return text



def _worksheet_dict_rows(ws: Any) -> Tuple[List[str], List[Dict[str, Any]]]:
    headers: List[str] = []
    dict_rows: List[Dict[str, Any]] = []
    for raw_row in ws.iter_rows(values_only=True):
        row_values = list(raw_row)
        if not any(value not in (None, "") for value in row_values):
            continue
        if not headers:
            headers = ["" if value is None else str(value).strip() for value in row_values]
            continue
        row_dict = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row_dict[header] = row_values[idx] if idx < len(row_values) else None
        dict_rows.append(row_dict)
    return headers, dict_rows



def _canonicalize_cluster_import_row(raw_row: Dict[str, Any], cluster_lookup: Optional[Dict[str, Dict[str, Any]]] = None) -> Optional[Dict[str, Any]]:
    canonical: Dict[str, Any] = {}
    for key, value in (raw_row or {}).items():
        normalized_key = _normalize_cluster_import_header(key)
        canonical_key = QUESTION_CLUSTER_IMPORT_ALIASES.get(normalized_key)
        if canonical_key:
            canonical[canonical_key] = value

    tenant_question_key = clean_text(canonical.get("tenant_question_key"))
    tenant_id = clean_text(canonical.get("tenant_id"))
    question_id = clean_text(canonical.get("question_id"))
    if tenant_question_key and "::" in tenant_question_key:
        inferred_tenant, inferred_question = tenant_question_key.split("::", 1)
        tenant_id = tenant_id or clean_text(inferred_tenant)
        question_id = question_id or clean_text(inferred_question)
    if tenant_id and question_id and not tenant_question_key:
        tenant_question_key = f"{tenant_id}::{question_id}"

    cluster_key = clean_text(canonical.get("cluster_key"))
    cluster_title = clean_text(canonical.get("cluster_title"))
    cluster_order = canonical.get("cluster_order")
    cluster_lookup = cluster_lookup or {}
    if cluster_key and cluster_key in cluster_lookup:
        lookup_row = cluster_lookup[cluster_key]
        cluster_title = cluster_title or clean_text(lookup_row.get("cluster_title"))
        if cluster_order in (None, ""):
            cluster_order = lookup_row.get("cluster_order")

    if not question_id or (not cluster_key and not cluster_title):
        return None
    return {
        "tenant_id": tenant_id,
        "tenant_name": clean_text(canonical.get("tenant_name")),
        "tenant_question_key": tenant_question_key or question_id,
        "question_id": question_id,
        "cluster_key": cluster_key,
        "cluster_title": cluster_title,
        "cluster_order": cluster_order,
        "question_text": clean_text(canonical.get("question_text")),
        "normalized_question": clean_text(canonical.get("normalized_question")),
        "product_group_hint": clean_text(canonical.get("product_group_hint")),
        "article_group_hint": clean_text(canonical.get("article_group_hint")),
        "product_name": clean_text(canonical.get("product_name")),
        "supplier_article_raw": clean_text(canonical.get("supplier_article_raw")),
        "nm_id": canonical.get("nm_id"),
    }



def _parse_clusters_lookup(rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    lookup: Dict[str, Dict[str, Any]] = {}
    for raw_row in rows:
        canonical: Dict[str, Any] = {}
        for key, value in (raw_row or {}).items():
            normalized_key = _normalize_cluster_import_header(key)
            canonical_key = QUESTION_CLUSTER_IMPORT_ALIASES.get(normalized_key)
            if canonical_key:
                canonical[canonical_key] = value
        cluster_key = clean_text(canonical.get("cluster_key"))
        cluster_title = clean_text(canonical.get("cluster_title"))
        cluster_order = canonical.get("cluster_order")
        if not cluster_key and cluster_title:
            cluster_key = cluster_title
        if not cluster_key and not cluster_title:
            continue
        lookup[cluster_key] = {
            "cluster_key": cluster_key,
            "cluster_title": cluster_title,
            "cluster_order": cluster_order,
        }
    return lookup



def _sheet_looks_like_cluster_map(headers: List[str], title: str = "") -> bool:
    canon = {QUESTION_CLUSTER_IMPORT_ALIASES.get(_normalize_cluster_import_header(item)) for item in headers if QUESTION_CLUSTER_IMPORT_ALIASES.get(_normalize_cluster_import_header(item))}
    if {"question_id", "cluster_key"}.issubset(canon) or {"tenant_question_key", "cluster_key"}.issubset(canon):
        return True
    if {"question_id", "cluster_title"}.issubset(canon) or {"tenant_question_key", "cluster_title"}.issubset(canon):
        return True
    normalized_title = _normalize_cluster_import_header(title)
    return normalized_title in {"cluster_map", "clusters_map", "questions_cluster_map"}



def _sheet_looks_like_clusters(headers: List[str], title: str = "") -> bool:
    canon = {QUESTION_CLUSTER_IMPORT_ALIASES.get(_normalize_cluster_import_header(item)) for item in headers if QUESTION_CLUSTER_IMPORT_ALIASES.get(_normalize_cluster_import_header(item))}
    normalized_title = _normalize_cluster_import_header(title)
    if normalized_title in {"clusters", "cluster_summary", "cluster_meta"}:
        return True
    return ("cluster_key" in canon or "cluster_title" in canon) and "question_id" not in canon and "tenant_question_key" not in canon



def _deduplicate_cluster_import_rows(rows: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], int]:
    deduped: Dict[str, Dict[str, Any]] = {}
    duplicates = 0
    for row in rows:
        tenant_id = clean_text(row.get("tenant_id"))
        question_id = clean_text(row.get("question_id"))
        stable_key = clean_text(row.get("tenant_question_key"))
        stable_key = stable_key or (f"{tenant_id}::{question_id}" if tenant_id else question_id)
        if not stable_key:
            continue
        if stable_key in deduped:
            duplicates += 1
        deduped[stable_key] = row
    return list(deduped.values()), duplicates



def _read_cluster_import_from_workbook(binary: bytes, source_name: str = "") -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(binary), read_only=True, data_only=True)
    sheet_payloads: List[Dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        headers, dict_rows = _worksheet_dict_rows(worksheet)
        if not headers and not dict_rows:
            continue
        sheet_payloads.append({"title": worksheet.title, "headers": headers, "rows": dict_rows})

    cluster_lookup: Dict[str, Dict[str, Any]] = {}
    for payload in sheet_payloads:
        if _sheet_looks_like_clusters(payload["headers"], payload["title"]):
            cluster_lookup.update(_parse_clusters_lookup(payload["rows"]))

    candidate_rows: List[Dict[str, Any]] = []
    raw_row_count = 0
    candidate_sheets: List[str] = []
    for payload in sheet_payloads:
        if not _sheet_looks_like_cluster_map(payload["headers"], payload["title"]):
            continue
        candidate_sheets.append(payload["title"])
        raw_row_count += len(payload["rows"])
        for raw_row in payload["rows"]:
            normalized = _canonicalize_cluster_import_row(raw_row, cluster_lookup=cluster_lookup)
            if normalized:
                candidate_rows.append(normalized)

    rows, duplicates = _deduplicate_cluster_import_rows(candidate_rows)
    if not rows:
        raise ValueError("Не найден лист cluster_map с колонками question_id / tenant_question_key и cluster_key / cluster_title.")
    return rows, {
        "source_name": source_name,
        "source_type": "xlsx",
        "candidate_sheets": candidate_sheets,
        "clusters_hint": len(cluster_lookup),
        "raw_rows": raw_row_count,
        "duplicates": duplicates,
    }



def _decode_tabular_bytes(binary: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251", "latin-1"):
        try:
            return binary.decode(encoding)
        except Exception:
            continue
    return binary.decode("utf-8", errors="ignore")



def _read_cluster_import_from_csv(binary: bytes, source_name: str = "") -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    text = _decode_tabular_bytes(binary)
    sample = text[:4096]
    delimiter = ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;	")
        delimiter = dialect.delimiter
    except Exception:
        pass
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    candidate_rows = []
    raw_row_count = 0
    for raw_row in reader:
        raw_row_count += 1
        normalized = _canonicalize_cluster_import_row(raw_row, cluster_lookup={})
        if normalized:
            candidate_rows.append(normalized)
    rows, duplicates = _deduplicate_cluster_import_rows(candidate_rows)
    if not rows:
        raise ValueError("CSV не содержит строк cluster_map с question_id / tenant_question_key и cluster_key / cluster_title.")
    return rows, {
        "source_name": source_name,
        "source_type": "csv",
        "raw_rows": raw_row_count,
        "duplicates": duplicates,
        "delimiter": delimiter,
    }



def _read_cluster_import_from_zip(binary: bytes, source_name: str = "") -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    errors: List[str] = []
    with zipfile.ZipFile(io.BytesIO(binary), "r") as zf:
        names = [name for name in zf.namelist() if not name.endswith("/")]
        candidate_names = sorted(
            names,
            key=lambda name: (
                0 if name.lower().endswith((".xlsx", ".xlsm")) and "cluster" in name.lower() else
                1 if name.lower().endswith((".xlsx", ".xlsm")) else
                2 if name.lower().endswith(".csv") and "cluster" in name.lower() else
                3 if name.lower().endswith(".csv") else
                9,
                name.lower(),
            ),
        )
        for name in candidate_names:
            try:
                payload = zf.read(name)
                lower_name = name.lower()
                if lower_name.endswith((".xlsx", ".xlsm")):
                    rows, meta = _read_cluster_import_from_workbook(payload, source_name=name)
                elif lower_name.endswith(".csv"):
                    rows, meta = _read_cluster_import_from_csv(payload, source_name=name)
                else:
                    continue
                meta["source_type"] = f"zip>{meta.get('source_type', 'file')}"
                meta["outer_source_name"] = source_name
                meta["inner_member"] = name
                return rows, meta
            except Exception as exc:
                errors.append(f"{name}: {exc}")
    details = " | ".join(errors[:8])
    raise ValueError(f"В ZIP не найден подходящий cluster_map (.xlsx/.csv). {details}")



def _parse_cluster_import_payload(filename: str, binary: bytes) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    lower_name = (filename or "").lower()
    if lower_name.endswith(".zip"):
        return _read_cluster_import_from_zip(binary, source_name=filename)
    if lower_name.endswith((".xlsx", ".xlsm")):
        return _read_cluster_import_from_workbook(binary, source_name=filename)
    if lower_name.endswith(".csv"):
        return _read_cluster_import_from_csv(binary, source_name=filename)
    raise ValueError("Поддерживаются только .zip, .xlsx, .xlsm или .csv с cluster_map.")



def _apply_cluster_import_rows(
    rows: List[Dict[str, Any]],
    *,
    source_name: str,
    overwrite_manual: bool = False,
    clear_previous_imported: bool = True,
) -> Dict[str, Any]:
    tenants = tenant_manager.load_tenants()
    tenant_map = {clean_text(item.get("id")): item for item in tenants if clean_text(item.get("id"))}
    grouped_rows: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    unknown_tenant_rows = 0
    for row in rows:
        tenant_id = clean_text(row.get("tenant_id"))
        stable_key = clean_text(row.get("tenant_question_key"))
        if not tenant_id and stable_key and "::" in stable_key:
            tenant_id = clean_text(stable_key.split("::", 1)[0])
            row["tenant_id"] = tenant_id
        if not tenant_id or tenant_id not in tenant_map:
            unknown_tenant_rows += 1
            continue
        grouped_rows[tenant_id].append(row)

    original_tenant_id = clean_text(getattr(g, "active_tenant_id", "") or session.get("tenant_id"))
    summary_rows: List[Dict[str, Any]] = []
    totals = {
        "tenant_count": 0,
        "tenant_errors": 0,
        "rows_in_file": len(rows),
        "unknown_tenant_rows": unknown_tenant_rows,
        "applied": 0,
        "preserved_manual": 0,
        "invalid_rows": 0,
        "missing_in_snapshot": 0,
        "cluster_count": 0,
        "cleaned_ai_drafts": 0,
    }

    try:
        for tenant_id, tenant_rows in grouped_rows.items():
            tenant_name = clean_text((tenant_map.get(tenant_id) or {}).get("name"))
            try:
                with _temporary_tenant_binding(tenant_id):
                    result = apply_imported_clusters_for_active_tenant(
                        tenant_rows,
                        source_name=source_name,
                        overwrite_manual=overwrite_manual,
                        clear_previous_imported=clear_previous_imported,
                    )
                totals["tenant_count"] += 1
                totals["applied"] += int(result.get("applied") or 0)
                totals["preserved_manual"] += int(result.get("preserved_manual") or 0)
                totals["invalid_rows"] += int(result.get("invalid_rows") or 0)
                totals["missing_in_snapshot"] += int(result.get("missing_in_snapshot") or 0)
                totals["cluster_count"] += int(result.get("cluster_count") or 0)
                totals["cleaned_ai_drafts"] += int(result.get("cleaned_ai_drafts") or 0)
                summary_rows.append(
                    {
                        "tenant_id": tenant_id,
                        "tenant_name": tenant_name,
                        "row_count": len(tenant_rows),
                        "applied": int(result.get("applied") or 0),
                        "preserved_manual": int(result.get("preserved_manual") or 0),
                        "invalid_rows": int(result.get("invalid_rows") or 0),
                        "missing_in_snapshot": int(result.get("missing_in_snapshot") or 0),
                        "cluster_count": int(result.get("cluster_count") or 0),
                        "cleaned_ai_drafts": int(result.get("cleaned_ai_drafts") or 0),
                        "status": "ok",
                        "error": "",
                    }
                )
            except Exception as exc:
                totals["tenant_errors"] += 1
                summary_rows.append(
                    {
                        "tenant_id": tenant_id,
                        "tenant_name": tenant_name,
                        "row_count": len(tenant_rows),
                        "applied": 0,
                        "preserved_manual": 0,
                        "invalid_rows": 0,
                        "missing_in_snapshot": 0,
                        "cluster_count": 0,
                        "cleaned_ai_drafts": 0,
                        "status": "error",
                        "error": str(exc),
                    }
                )
    finally:
        restore_id = original_tenant_id or clean_text((tenants[0] if tenants else {}).get("id"))
        if restore_id and restore_id != clean_text(getattr(g, "active_tenant_id", "")):
            try:
                tenant = tenant_manager.get_tenant(restore_id)
                g.active_tenant = tenant
                g.active_tenant_id = restore_id
            except Exception:
                pass

    totals["tenants"] = summary_rows
    return totals


def _ensure_prompt_file() -> None:
    SYSTEM_PROMPT_FILE.parent.mkdir(parents=True, exist_ok=True)
    BASE_SYSTEM_PROMPT_FILE.parent.mkdir(parents=True, exist_ok=True)
    if not BASE_SYSTEM_PROMPT_FILE.exists():
        safe_files.write_text(BASE_SYSTEM_PROMPT_FILE, "Пиши вежливо.", encoding="utf-8")
    if not SYSTEM_PROMPT_FILE.exists():
        safe_files.write_text(Path(SYSTEM_PROMPT_FILE), BASE_SYSTEM_PROMPT_FILE.read_text(encoding="utf-8"), encoding="utf-8")


def get_prompt_text() -> str:
    _ensure_prompt_file()
    root_text = BASE_SYSTEM_PROMPT_FILE.read_text(encoding="utf-8") if BASE_SYSTEM_PROMPT_FILE.exists() else ""
    tenant_text = SYSTEM_PROMPT_FILE.read_text(encoding="utf-8") if SYSTEM_PROMPT_FILE.exists() else ""
    root_clean = root_text.strip()
    tenant_clean = tenant_text.strip()
    if root_clean and root_clean != tenant_clean:
        safe_files.write_text(Path(SYSTEM_PROMPT_FILE), root_clean, encoding="utf-8")
        return root_clean
    return tenant_clean or root_clean or "Пиши вежливо."


def save_prompt_text(text: str) -> None:
    _ensure_prompt_file()
    clean = text.strip() or "Пиши вежливо."
    safe_files.write_text(BASE_SYSTEM_PROMPT_FILE, clean, encoding="utf-8")
    safe_files.write_text(Path(SYSTEM_PROMPT_FILE), clean, encoding="utf-8")


def queue_selected_replies(form: Any) -> Tuple[int, List[str]]:
    reviews = _collect_selected_reviews_from_form(form)
    selected_ids = [clean_text(x) for x in form.getlist("selected_ids") if clean_text(x)]
    queue = load_reply_queue()
    queue_open = {clean_text(item.get("review_id")) for item in queue if clean_text(item.get("status")) in {"queued", "processing"}}
    sent_ids = get_locally_sent_ids()
    drafts = load_drafts()
    added = 0
    notes: List[str] = []

    reviews_map = {clean_text(review.get("id")): review for review in reviews}
    for review_id in selected_ids:
        review = reviews_map.get(review_id)
        if not review:
            notes.append(f"{review_id}: отзыв не найден в форме.")
            continue
        reply_text = trim_reply(clean_text(form.get(f"reply__{review_id}")), limit=MAX_REPLY_LENGTH)
        if len(reply_text) < MIN_REPLY_LENGTH:
            notes.append(f"{review_id}: сначала сгенерируйте или введите текст ответа.")
            continue
        if review_id in sent_ids:
            notes.append(f"{review_id}: отзыв уже отмечен как отправленный локально.")
            continue
        # обновляем/создаём локальный draft из формы, чтобы редактура не потерялась
        draft_entry = drafts.get(review_id, {})
        draft_entry.update(
            {
                "reply": reply_text,
                "signature": review_signature(review),
                "generated_at": utc_now_iso(),
                "source": clean_text(draft_entry.get("source") or "manual_edit"),
                "rule_ids": draft_entry.get("rule_ids", []),
                "cross_sell_items": draft_entry.get("cross_sell_items", []),
            }
        )
        drafts[review_id] = draft_entry
        if review_id in queue_open:
            for item in queue:
                if clean_text(item.get("review_id")) == review_id and clean_text(item.get("status")) in {"queued", "processing", "failed"}:
                    item.update(
                        {
                            "reply": reply_text,
                            "signature": review_signature(review),
                            "review": {
                                "product_name": clean_text(review.get("productDetails", {}).get("productName")),
                                "supplier_article": clean_text(review.get("productDetails", {}).get("supplierArticle")),
                                "brand_name": clean_text(review.get("productDetails", {}).get("brandName")),
                                "nm_id": _safe_int(review.get("productDetails", {}).get("nmId")),
                                "stars": _safe_int(review.get("productValuation")),
                                "user_name": clean_text(review.get("userName")),
                                "created_date": clean_text(review.get("createdDate")),
                                "subject_name": clean_text(review.get("subjectName")),
                                "text": clean_text(review.get("text")),
                                "pros": clean_text(review.get("pros")),
                                "cons": clean_text(review.get("cons")),
                                "review_text": build_review_text(review),
                            },
                            "status": "queued" if clean_text(item.get("status")) != "processing" else clean_text(item.get("status")),
                            "error": "",
                        }
                    )
                    notes.append(f"{review_id}: черновик в очереди обновлён.")
                    break
            continue
        queue.append(
            {
                "review_id": review_id,
                "signature": review_signature(review),
                "status": "queued",
                "queued_at": utc_now_iso(),
                "reply": reply_text,
                "reply_source": clean_text(draft_entry.get("source") or "manual_edit"),
                "review": {
                    "product_name": clean_text(review.get("productDetails", {}).get("productName")),
                    "supplier_article": clean_text(review.get("productDetails", {}).get("supplierArticle")),
                    "brand_name": clean_text(review.get("productDetails", {}).get("brandName")),
                    "nm_id": _safe_int(review.get("productDetails", {}).get("nmId")),
                    "stars": _safe_int(review.get("productValuation")),
                    "user_name": clean_text(review.get("userName")),
                    "created_date": clean_text(review.get("createdDate")),
                    "subject_name": clean_text(review.get("subjectName")),
                    "text": clean_text(review.get("text")),
                    "pros": clean_text(review.get("pros")),
                    "cons": clean_text(review.get("cons")),
                    "review_text": build_review_text(review),
                },
            }
        )
        added += 1
    save_reply_queue(queue)
    save_drafts(drafts)
    return added, notes


def process_reply_queue(max_items: int = 0) -> Dict[str, Any]:
    queue = load_reply_queue()
    pending_indexes = [idx for idx, item in enumerate(queue) if clean_text(item.get("status")) in {"queued", "failed"}]
    if max_items > 0:
        pending_indexes = pending_indexes[:max_items]
    sent = 0
    failed = 0
    processed = 0
    for idx in pending_indexes:
        item = queue[idx]
        review_id = clean_text(item.get("review_id"))
        reply_text = trim_reply(clean_text(item.get("reply")), limit=MAX_REPLY_LENGTH)
        if len(reply_text) < MIN_REPLY_LENGTH:
            item["status"] = "failed"
            item["error"] = "Текст ответа пустой или слишком короткий."
            failed += 1
            processed += 1
            continue
        item["status"] = "processing"
        item["error"] = ""
        save_reply_queue(queue)
        try:
            send_reply_to_wb(review_id, reply_text)
            item["status"] = "sent"
            item["sent_at"] = utc_now_iso()
            item["error"] = ""
            review_meta = item.get("review", {}) or {}
            draft_entry = load_drafts().get(review_id, {})
            upsert_archive_record(
                {
                    "id": review_id,
                    "status": "sent",
                    "sent_at": item["sent_at"],
                    "product_name": clean_text(review_meta.get("product_name")),
                    "supplier_article": clean_text(review_meta.get("supplier_article")),
                    "brand_name": clean_text(review_meta.get("brand_name")),
                    "nm_id": _safe_int(review_meta.get("nm_id")),
                    "stars": _safe_int(review_meta.get("stars")),
                    "user_name": clean_text(review_meta.get("user_name")),
                    "created_date": clean_text(review_meta.get("created_date")),
                    "subject_name": clean_text(review_meta.get("subject_name")),
                    "review_text": clean_text(review_meta.get("review_text")),
                    "raw_text": clean_text(review_meta.get("text")),
                    "raw_pros": clean_text(review_meta.get("pros")),
                    "raw_cons": clean_text(review_meta.get("cons")),
                    "reply": reply_text,
                    "reply_source": clean_text(item.get("reply_source") or draft_entry.get("source") or "queued"),
                    "rule_ids": draft_entry.get("rule_ids", []),
                    "cross_sell_items": draft_entry.get("cross_sell_items", []),
                }
            )
            remove_draft(review_id)
            try:
                history_service.mark_replied(review_id, reply_text, getattr(g, "active_tenant_id", "") or common.get_active_tenant_id())
            except Exception as exc:
                log_event("replies", "mark_history_replied_failed", tenant_id=_current_tenant_for_logs(), level="error", review_id=review_id, error=str(exc))
            sent += 1
            processed += 1
            save_reply_queue(queue)
            time.sleep(API_SEND_DELAY_SECONDS)
        except Exception as exc:
            item["status"] = "failed"
            item["error"] = str(exc)
            failed += 1
            processed += 1
            save_reply_queue(queue)
            if "429" in str(exc):
                time.sleep(max(2.0, API_SEND_DELAY_SECONDS * 3))
    message = f"Очередь ответов обработана. Успешно: {sent}. С ошибкой: {failed}. Всего: {processed}."
    return {"message": message, "sent": sent, "failed": failed, "processed": processed}


@app.route("/")
def index() -> str:
    params = _reviews_params_from_source(request.args)
    page = int(params["page"])
    force_refresh = str(request.args.get("refresh", "")).lower() in {"1", "true", "yes"}
    try:
        context = build_reply_rows(
            page=page,
            page_size=params["page_size"],
            sort_by=params["sort"],
            stars_filter=params["stars"],
            draft_filter=params["draft"],
            queue_filter=params["queue"],
            search_query=params["q"],
            hide_sent=bool(params["hide_sent"]),
            source_filter=params["source"],
            content_filter=params["content"],
            answer_state=params["answer_state"],
            force_refresh=force_refresh,
        )
        context["private_dir"] = str(PRIVATE_DIR)
        context["prompt_text"] = get_prompt_text()
        context["prompt_file_path"] = str(SYSTEM_PROMPT_FILE)
        context["page_size_options"] = PAGE_SIZE_OPTIONS
        context["reply_queue_open"] = sum(1 for item in load_reply_queue() if clean_text(item.get("status")) in {"queued", "processing", "failed"})
        return render_template("index.html", **context)
    except Exception as exc:
        return render_template(
            "index.html",
            rows=[],
            page=page,
            page_size=params["page_size"],
            page_count=1,
            has_prev=False,
            has_next=False,
            total_filtered=0,
            submitted_hidden=0,
            raw_scanned=0,
            count_unanswered=0,
            count_archive=0,
            draft_count=0,
            reply_queue_open=0,
            filters={
                "sort": params["sort"],
                "stars": params["stars"],
                "draft": params["draft"],
                "queue": params["queue"],
                "q": params["q"],
                "hide_sent": bool(params["hide_sent"]),
                "source": params["source"],
                "content": params["content"],
                "answer_state": params["answer_state"],
                "page_size": params["page_size"],
            },
            snapshot_fetched_at="",
            history_counts=history_service.get_counts(getattr(g, "active_tenant_id", "") or common.get_active_tenant_id()),
            history_meta=history_service.effective_meta(getattr(g, "active_tenant_id", "") or common.get_active_tenant_id()),
            using_history_db=history_service.db_has_data(getattr(g, "active_tenant_id", "") or common.get_active_tenant_id()),
            private_dir=str(PRIVATE_DIR),
            prompt_text=get_prompt_text(),
            prompt_file_path=str(SYSTEM_PROMPT_FILE),
            page_size_options=PAGE_SIZE_OPTIONS,
            load_error=str(exc),
        )


@app.route("/replies/prepare", methods=["POST"])
def replies_prepare() -> Any:
    reviews = _collect_selected_reviews_from_form(request.form)
    force = bool(request.form.get("force"))
    params = _reviews_params_from_source(request.form)
    if not reviews:
        flash("Выберите хотя бы один отзыв для AI-генерации ответа.", "error")
        return redirect(url_for("index", **params))
    try:
        _submit_background_task("replies_prepare", "Подготовка AI-черновиков ответов", _job_prepare_replies, reviews, force=force)
    except Exception as exc:
        flash(f"Не удалось поставить подготовку черновиков в фон: {exc}", "error")
    return redirect(url_for("index", **params))


@app.route("/replies/prepare_json", methods=["POST"])
def replies_prepare_json() -> Any:
    data = request.get_json(silent=True) or {}
    selected_ids = [clean_text(x) for x in data.get("selected_ids", []) if clean_text(x)]
    force = bool(data.get("force"))
    if not selected_ids:
        return jsonify({"ok": False, "message": "Не выбраны отзывы."}), 400
    reviews, missing = _get_snapshot_reviews_by_ids(selected_ids, force_refresh=False)
    generated = 0
    errors: List[str] = []
    for review in reviews:
        try:
            generate_reply_for_review(review, force=force)
            generated += 1
        except Exception as exc:
            errors.append(f"{clean_text(review.get('id'))}: {exc}")
    return jsonify({
        "ok": generated > 0 and not errors,
        "generated": generated,
        "missing": missing,
        "errors": errors,
        "message": f"AI подготовил или обновил черновики ответов: {generated}."
    })


@app.route("/replies/save_drafts_json", methods=["POST"])
def replies_save_drafts_json() -> Any:
    data = request.get_json(silent=True) or {}
    items = data.get("items", [])
    if not isinstance(items, list):
        items = []
    if not items:
        return jsonify({"ok": False, "message": "Нет данных для сохранения черновиков."}), 400
    saved, notes = _save_server_reply_drafts(items)
    return jsonify({
        "ok": True,
        "saved": saved,
        "notes": notes,
        "message": f"Серверных черновиков сохранено: {saved}."
    })


@app.route("/replies/queue_ids_json", methods=["POST"])
def replies_queue_ids_json() -> Any:
    data = request.get_json(silent=True) or {}
    selected_ids = [clean_text(x) for x in data.get("selected_ids", []) if clean_text(x)]
    if not selected_ids:
        return jsonify({"ok": False, "message": "Не выбраны отзывы для очереди."}), 400
    added, notes = _queue_server_reply_drafts_by_ids(selected_ids)
    ok = added > 0
    return jsonify({
        "ok": ok,
        "added": added,
        "notes": notes,
        "message": f"В очередь ответов добавлено: {added}." if ok else "Ни один ответ не был добавлен в очередь. Проверьте, что у выбранных отзывов есть текст ответа."
    })


@app.route("/replies/queue_json", methods=["POST"])
def replies_queue_json() -> Any:
    data = request.get_json(silent=True) or {}
    items = data.get("items", [])
    if not isinstance(items, list):
        items = []
    if not items:
        return jsonify({"ok": False, "message": "Не выбраны ответы для очереди."}), 400
    saved, save_notes = _save_server_reply_drafts(items)
    selected_ids = [clean_text(item.get("review_id")) for item in items if clean_text(item.get("review_id"))]
    added, queue_notes = _queue_server_reply_drafts_by_ids(selected_ids)
    ok = added > 0
    return jsonify({
        "ok": ok,
        "saved": saved,
        "added": added,
        "notes": [*save_notes, *queue_notes],
        "message": f"В очередь ответов добавлено: {added}." if ok else "Ни один ответ не был добавлен в очередь. Проверьте, что у выбранных отзывов есть текст ответа."
    })


@app.route("/replies/queue", methods=["POST"])
def replies_queue() -> Any:
    params = _reviews_params_from_source(request.form)
    selected_ids = [clean_text(x) for x in request.form.getlist("selected_ids") if clean_text(x)]
    if not selected_ids:
        flash("Выберите хотя бы один отзыв для постановки в очередь отправки.", "error")
        return redirect(url_for("index", **params))
    added, notes = queue_selected_replies(request.form)
    if added:
        flash(f"В очередь ответов добавлено: {added}.", "success")
    else:
        flash("Ни один ответ не был добавлен в очередь. Проверьте, что у выбранных отзывов есть текст ответа.", "error")
    for note in notes[:20]:
        flash(note, "success")
    return redirect(url_for("index", **params))


@app.route("/replies/process", methods=["POST"])
def replies_process() -> Any:
    params = _reviews_params_from_source(request.form)
    try:
        max_items = _safe_int(request.form.get("max_items", 0), 0)
    except Exception:
        max_items = 0
    try:
        _submit_background_task("replies_process", "Отправка очереди ответов", process_reply_queue, max(0, max_items), unique_key="replies_process")
    except Exception as exc:
        flash(f"Не удалось поставить отправку очереди ответов в фон: {exc}", "error")
    return redirect(url_for("index", **params))


@app.route("/replies/prompt", methods=["POST"])
def replies_prompt() -> Any:
    params = _reviews_params_from_source(request.form)
    try:
        save_prompt_text(request.form.get("prompt_text", ""))
        flash("Промпт для генерации ответов обновлён и сохранён в текстовый файл.", "success")
    except Exception as exc:
        flash(f"Не удалось сохранить промпт: {exc}", "error")
    return redirect(url_for("index", **params))


@app.route("/generate", methods=["POST"])
def generate() -> Any:
    review = restore_review_from_form(request.form)
    try:
        entry = generate_reply_for_review(review, force=bool(request.form.get("force")))
        flash(f"Черновик для отзыва {review['id']} обновлён через AI.", "success")
    except Exception as exc:
        flash(f"Не удалось сгенерировать ответ: {exc}", "error")
    return redirect(url_for("index"))


@app.route("/generate-page", methods=["POST"])
def generate_page() -> Any:
    try:
        snapshot = get_reply_snapshot(force_refresh=bool(request.form.get("force_refresh")))
        drafts = load_drafts()
        generated = 0
        skipped = 0
        for review in snapshot.get("feedbacks", []):
            review = normalize_review(review)
            review_id = clean_text(review.get("id"))
            signature = review_signature(review)
            cached = drafts.get(review_id)
            if is_draft_compatible(cached, signature, prompt_signature(get_prompt_text())):
                skipped += 1
                continue
            generate_reply_for_review(review, force=False)
            generated += 1
        flash(f"Черновики обновлены: {generated}. Уже были готовы: {skipped}.", "success")
    except Exception as exc:
        flash(f"Не удалось сгенерировать черновики для страницы: {exc}", "error")
    return redirect(url_for("index"))


@app.route("/reply", methods=["POST"])
def reply() -> Any:
    review = restore_review_from_form(request.form)
    review_id = clean_text(review.get("id"))
    reply_text = trim_reply(clean_text(request.form.get("reply")), limit=MAX_REPLY_LENGTH)
    if len(reply_text) < MIN_REPLY_LENGTH:
        flash("Ответ слишком короткий. Введите текст ответа или сначала сгенерируйте черновик.", "error")
        return redirect(url_for("index"))
    already_sent = get_locally_sent_ids()
    if review_id in already_sent:
        flash("Этот отзыв уже отмечен как отправленный локально. Повторная отправка заблокирована.", "error")
        return redirect(url_for("index"))
    drafts = load_drafts()
    draft_entry = drafts.get(review_id, {})
    try:
        send_reply_to_wb(review_id, reply_text)
        upsert_archive_record(
            {
                "id": review_id,
                "status": "sent",
                "sent_at": utc_now_iso(),
                "product_name": clean_text(review.get("productDetails", {}).get("productName")),
                "supplier_article": clean_text(review.get("productDetails", {}).get("supplierArticle")),
                "brand_name": clean_text(review.get("productDetails", {}).get("brandName")),
                "nm_id": int(review.get("productDetails", {}).get("nmId", 0) or 0),
                "stars": int(review.get("productValuation", 0) or 0),
                "user_name": clean_text(review.get("userName")),
                "created_date": clean_text(review.get("createdDate")),
                "subject_name": clean_text(review.get("subjectName")),
                "review_text": build_review_text(review),
                "raw_text": clean_text(review.get("text")),
                "raw_pros": clean_text(review.get("pros")),
                "raw_cons": clean_text(review.get("cons")),
                "reply": reply_text,
                "reply_source": clean_text(draft_entry.get("source") or "manual"),
                "rule_ids": draft_entry.get("rule_ids", []),
                "cross_sell_items": draft_entry.get("cross_sell_items", []),
            }
        )
        remove_draft(review_id)
        flash(f"Ответ по отзыву {review_id} успешно отправлен в Wildberries и сохранён в локальный архив.", "success")
    except Exception as exc:
        flash(f"Не удалось отправить ответ в WB: {exc}", "error")
    return redirect(url_for("index"))


@app.route("/analytics")
def analytics() -> str:
    stats = build_analytics()
    return render_template("analytics.html", stats=stats)


@app.route("/complaints")
def complaints() -> str:
    params = _complaints_params_from_source(request.args)
    page = int(params["page"])
    try:
        context = fetch_low_rating_reviews(
            page=page,
            page_size=params["page_size"],
            sort_by=params["sort"],
            stars_filter=params["stars"],
            draft_filter=params["draft"],
            queue_filter=params["queue"],
            search_query=params["q"],
            hide_submitted=bool(params["hide_submitted"]),
        )
        context["auth"] = get_auth_status()
        context["stats"] = complaint_dashboard_stats()
        context["private_dir"] = str(PRIVATE_DIR)
        context["page_size_options"] = PAGE_SIZE_OPTIONS
        return render_template("complaints.html", **context)
    except Exception as exc:
        return render_template(
            "complaints.html",
            rows=[],
            page=page,
            page_size=params["page_size"],
            page_count=1,
            has_prev=False,
            has_next=False,
            total_filtered=0,
            submitted_hidden=0,
            raw_scanned=0,
            low_total_scanned=0,
            count_unanswered=0,
            count_archive=0,
            queue_total=len(load_complaint_queue()),
            draft_total=len(load_complaint_drafts()),
            auth=get_auth_status(),
            stats=complaint_dashboard_stats(),
            private_dir=str(PRIVATE_DIR),
            page_size_options=PAGE_SIZE_OPTIONS,
            filters={
                "sort": params["sort"],
                "stars": params["stars"],
                "draft": params["draft"],
                "queue": params["queue"],
                "q": params["q"],
                "hide_submitted": bool(params["hide_submitted"]),
            },
            load_error=str(exc),
        )


@app.route("/complaints/prepare", methods=["POST"])
def complaints_prepare() -> Any:
    reviews = _collect_selected_reviews_from_form(request.form)
    force = bool(request.form.get("force"))
    params = _complaints_params_from_source(request.form)
    if not reviews:
        flash("Выберите хотя бы один отзыв для AI-анализа жалобы.", "error")
        return redirect(url_for("complaints", **params))
    try:
        _submit_background_task("complaints_prepare", "Подготовка жалоб", _job_prepare_complaints, reviews, force=force)
    except Exception as exc:
        flash(f"Не удалось поставить подготовку жалоб в фон: {exc}", "error")
    return redirect(url_for("complaints", **params))


@app.route("/complaints/queue", methods=["POST"])
def complaints_queue() -> Any:
    selected_ids = [clean_text(x) for x in request.form.getlist("selected_ids") if clean_text(x)]
    params = _complaints_params_from_source(request.form)
    if not selected_ids:
        flash("Выберите хотя бы один отзыв для постановки в очередь жалоб.", "error")
        return redirect(url_for("complaints", **params))
    drafts = load_complaint_drafts()
    entries = [drafts.get(review_id, {}) for review_id in selected_ids]
    added, notes = queue_complaint_entries(entries)
    if added:
        flash(f"В очередь добавлено жалоб: {added}.", "success")
    else:
        flash("Ни одна жалоба не попала в очередь. Проверьте, что для выбранных отзывов уже подготовлены AI-черновики.", "error")
    for note in notes[:15]:
        flash(note, "error")
    return redirect(url_for("complaints", **params))


@app.route("/complaints/process", methods=["POST"])
def complaints_process() -> Any:
    params = _complaints_params_from_source(request.form)
    try:
        max_items = int(request.form.get("max_items", 0) or 0)
    except Exception:
        max_items = 0
    dry_run = bool(request.form.get("dry_run"))
    try:
        _submit_background_task("complaints_process", "Отправка очереди жалоб", process_queue, max(0, max_items), dry_run=dry_run, unique_key="complaints_process")
    except Exception as exc:
        flash(f"Не удалось поставить браузерную отправку жалоб в фон: {exc}", "error")
    return redirect(url_for("complaints", **params))


@app.route("/complaints/export.xlsx")
def complaints_export_xlsx() -> Any:
    params = _complaints_params_from_source(request.args)
    context = fetch_low_rating_reviews(
        page=1,
        page_size=100000,
        sort_by=params["sort"],
        stars_filter=params["stars"],
        draft_filter=params["draft"],
        queue_filter=params["queue"],
        search_query=params["q"],
        hide_submitted=bool(params["hide_submitted"]),
        force_refresh=bool(request.args.get("refresh")),
    )
    rows = context.get("rows", []) if context else []
    wb = Workbook()
    ws = wb.active
    ws.title = "Негативные отзывы"
    headers = ["ID отзыва", "Оценка", "Дата", "Покупатель", "Товар", "Артикул", "nmID", "Текст", "Плюсы", "Минусы", "Сводный текст отзыва"]
    ws.append(headers)
    for row in rows:
        ws.append([
            clean_text(row.get("id")),
            int(row.get("stars", 0) or 0),
            clean_text(row.get("created_date")),
            clean_text(row.get("user_name")),
            clean_text(row.get("product_name")),
            clean_text(row.get("supplier_article")),
            clean_text(row.get("nm_id")),
            clean_text(row.get("text")),
            clean_text(row.get("pros")),
            clean_text(row.get("cons")),
            clean_text(row.get("review_text")),
        ])
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 60)
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"negative_reviews_{clean_text(getattr(g, 'active_tenant_id', 'tenant') or 'tenant')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(stream, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/complaints/effectiveness")
def complaints_effectiveness() -> Any:
    stats = build_complaint_effectiveness(limit=5000)
    return render_template("complaint_effectiveness.html", stats=stats)


@app.route("/complaints/effectiveness/refresh", methods=["POST"])
def complaints_effectiveness_refresh() -> Any:
    try:
        max_items = int(request.form.get("max_items", 0) or 0)
    except Exception:
        max_items = 0
    try:
        _submit_background_task("complaints_refresh_outcomes", "Обновление исходов жалоб", refresh_complaint_outcomes, max(0, max_items), unique_key="complaints_refresh_outcomes")
    except Exception as exc:
        flash(f"Не удалось поставить обновление статусов жалоб в фон: {exc}", "error")
    return redirect(url_for("complaints_effectiveness", **_tenant_query_arg()))


@app.route("/questions")
def questions() -> str:
    log_event('questions', 'page_open', tenant_id=_current_tenant_for_logs(), mode=clean_text(request.args.get('mode') or 'questions'))
    params = _questions_params_from_source(request.args)
    page = int(params["page"])
    force_refresh = str(request.args.get("refresh", "")).lower() in {"1", "true", "yes"}
    try:
        context = build_questions_context(
            page=page,
            page_size=params["page_size"],
            sort_by=params["sort"],
            draft_filter=params["draft"],
            queue_filter=params["queue"],
            search_query=params["q"],
            hide_submitted=bool(params["hide_submitted"]),
            cluster_filter=params["cluster"],
            mode=params["mode"],
            force_refresh=force_refresh,
        )
        context["private_dir"] = str(PRIVATE_DIR)
        return render_template("questions.html", **context)
    except Exception as exc:
        return render_template(
            "questions.html",
            rows=[],
            all_rows=[],
            clusters=[],
            page=page,
            page_size=params["page_size"],
            page_count=1,
            has_prev=False,
            has_next=False,
            total_filtered=0,
            submitted_hidden=0,
            raw_scanned=0,
            count_unanswered=0,
            count_archive=0,
            draft_total=0,
            queue_total=0,
            queue_open=0,
            cluster_total=0,
            draft_ready_total=0,
            archive_total=0,
            snapshot_fetched_at="",
            truncated=False,
            has_new_questions=False,
            selected_cluster=None,
            selected_cluster_rule={},
            filters={
                "sort": params["sort"],
                "draft": params["draft"],
                "queue": params["queue"],
                "q": params["q"],
                "hide_submitted": bool(params["hide_submitted"]),
                "cluster": params["cluster"],
                "mode": params["mode"],
                "page_size": params["page_size"],
            },
            question_prompt_text=common.load_question_prompt(),
            question_rules=load_question_rules(),
            private_dir=str(PRIVATE_DIR),
            load_error=str(exc),
            page_size_options=QUESTION_PAGE_SIZE_OPTIONS,
        )


@app.route("/questions/refresh", methods=["POST"])
def questions_refresh() -> Any:
    log_event('questions', 'refresh_requested', tenant_id=_current_tenant_for_logs())
    params = _questions_params_from_source(request.form)
    try:
        _submit_background_task("questions_refresh", "Обновление снимка вопросов", _job_refresh_questions, unique_key="questions_refresh")
    except Exception as exc:
        flash(f"Не удалось поставить обновление вопросов в фон: {exc}", "error")
    return redirect(url_for("questions", **params))


@app.route("/questions/save", methods=["POST"])
def questions_save() -> Any:
    log_event('questions', 'form_save_requested', tenant_id=_current_tenant_for_logs())
    params = _questions_params_from_source(request.form)
    try:
        updated = save_question_form_edits(request.form)
        flash(f"Сохранено комментариев и правок: {updated}.", "success")
    except Exception as exc:
        flash(f"Не удалось сохранить комментарии и правки: {exc}", "error")
    return redirect(url_for("questions", **params))


@app.route("/questions/ignore", methods=["POST"])
def questions_ignore() -> Any:
    log_event('questions', 'ignore_requested', tenant_id=_current_tenant_for_logs())
    params = _questions_params_from_source(request.form)
    selected_ids = _resolve_question_ids(request.form, allow_cluster_scope=False)
    if not selected_ids:
        question_id = clean_text(request.form.get("question_id"))
        if question_id:
            selected_ids = [question_id]
    if not selected_ids:
        flash("Не выбрано ни одного вопроса для удаления из рабочей ленты.", "error")
        return redirect(url_for("questions", **params))
    try:
        removed = ignore_question_ids(selected_ids)
        flash(f"Удалено из рабочей ленты вопросов: {removed}.", "success")
    except Exception as exc:
        flash(f"Не удалось удалить вопросы из рабочей ленты: {exc}", "error")
    return redirect(url_for("questions", **params))


@app.route("/questions/prepare", methods=["POST"])
def questions_prepare() -> Any:
    log_event('questions', 'prepare_requested', tenant_id=_current_tenant_for_logs())
    params = _questions_params_from_source(request.form)
    selected_ids = _resolve_question_ids(request.form, allow_cluster_scope=True)
    if not selected_ids:
        flash("Выберите хотя бы один вопрос или укажите генерацию по кластеру.", "error")
        return redirect(url_for("questions", **params))
    force = bool(request.form.get("force"))
    try:
        save_question_form_edits(request.form)
        _submit_background_task("questions_prepare", "Подготовка черновиков вопросов", _job_prepare_questions, selected_ids, force=force)
    except Exception as exc:
        flash(f"Не удалось поставить подготовку черновиков вопросов в фон: {exc}", "error")
    return redirect(url_for("questions", **params))


@app.route("/questions/queue", methods=["POST"])
def questions_queue() -> Any:
    log_event('questions', 'queue_requested', tenant_id=_current_tenant_for_logs())
    params = _questions_params_from_source(request.form)
    selected_ids = _resolve_question_ids(request.form, allow_cluster_scope=True)
    if not selected_ids:
        flash("Выберите хотя бы один вопрос или весь кластер для постановки в очередь.", "error")
        return redirect(url_for("questions", **params))

    class _ProxyForm:
        def __init__(self, source_form: Any, ids: List[str]):
            self._source_form = source_form
            self._ids = ids
        def getlist(self, name: str) -> List[str]:
            if name == "selected_ids":
                return list(self._ids)
            return self._source_form.getlist(name)
        def get(self, name: str, default: Any = None) -> Any:
            return self._source_form.get(name, default)

    try:
        save_question_form_edits(request.form)
        added, notes = queue_questions_from_form(_ProxyForm(request.form, selected_ids))
        if added:
            flash(f"В очередь вопросов добавлено: {added}.", "success")
        else:
            flash("Ни один вопрос не был поставлен в очередь. Проверьте тексты ответов и действие по строкам.", "error")
        for note in notes[:20]:
            flash(note, "success")
    except Exception as exc:
        flash(f"Не удалось поставить вопросы в очередь: {exc}", "error")
    return redirect(url_for("questions", **params))


@app.route("/questions/export-clustering", methods=["POST"])
def questions_export_clustering() -> Any:
    log_event('questions', 'export_clustering_requested', tenant_id=_current_tenant_for_logs())
    params = _questions_params_from_source(request.form)
    refresh_all = str(request.form.get("refresh_all", "")).lower() in {"1", "true", "on", "yes"}
    try:
        rows, summary_rows, meta = _collect_question_clustering_rows(refresh_all=refresh_all)
        zip_stream, filename = _build_question_clustering_zip(rows, summary_rows, meta)
        return send_file(zip_stream, as_attachment=True, download_name=filename, mimetype="application/zip")
    except Exception as exc:
        flash(f"Не удалось подготовить выгрузку для кластеризации: {exc}", "error")
        return redirect(url_for("questions", **params))


@app.route("/questions/import-clustering", methods=["POST"])
def questions_import_clustering() -> Any:
    log_event('questions', 'import_clustering_requested', tenant_id=_current_tenant_for_logs())
    params = _questions_params_from_source(request.form)
    uploaded = request.files.get("cluster_import_file")
    if uploaded is None or not clean_text(getattr(uploaded, "filename", "")):
        flash("Выберите файл cluster_map (.zip / .xlsx / .csv) для обратного импорта кластеров.", "error")
        return redirect(url_for("questions", **params))

    overwrite_manual = bool(request.form.get("overwrite_manual"))
    clear_previous_imported = str(request.form.get("clear_previous_imported", "1")).lower() in {"1", "true", "on", "yes"}
    try:
        binary = uploaded.read()
        rows, parser_meta = _parse_cluster_import_payload(uploaded.filename, binary)
        result = _apply_cluster_import_rows(
            rows,
            source_name=clean_text(parser_meta.get("outer_source_name") or parser_meta.get("source_name") or uploaded.filename),
            overwrite_manual=overwrite_manual,
            clear_previous_imported=clear_previous_imported,
        )
        flash(
            (
                "Импорт кластеров завершён. "
                f"Строк в файле: {result.get('rows_in_file', 0)}. "
                f"Назначений применено: {result.get('applied', 0)}. "
                f"Кластеров: {result.get('cluster_count', 0)}. "
                f"Кабинетов обновлено: {result.get('tenant_count', 0)}."
            ),
            "success" if result.get("tenant_errors", 0) == 0 else "error",
        )
        if result.get("preserved_manual"):
            flash(f"Сохранено ручных подкластеров менеджера без перезаписи: {result.get('preserved_manual')}", "success")
        if result.get("cleaned_ai_drafts"):
            flash(f"Удалено устаревших AI-черновиков после смены кластеров: {result.get('cleaned_ai_drafts')}", "success")
        if result.get("unknown_tenant_rows"):
            flash(f"Строк с неизвестным tenant_id пропущено: {result.get('unknown_tenant_rows')}", "error")
        if parser_meta.get("duplicates"):
            flash(f"В импортируемом файле были дубли строк; использована последняя версия для {parser_meta.get('duplicates')} ключей.", "success")
        for tenant_row in (result.get("tenants") or [])[:10]:
            if clean_text(tenant_row.get("status")) == "ok":
                flash(
                    f"{tenant_row.get('tenant_name') or tenant_row.get('tenant_id')}: применено {tenant_row.get('applied', 0)} строк, кластеров {tenant_row.get('cluster_count', 0)}.",
                    "success",
                )
            else:
                flash(
                    f"{tenant_row.get('tenant_name') or tenant_row.get('tenant_id')}: ошибка импорта — {tenant_row.get('error')}",
                    "error",
                )
        params["mode"] = "questions"
        params["sort"] = "cluster"
        params["cluster"] = ""
        params["page"] = 1
    except Exception as exc:
        flash(f"Не удалось импортировать cluster_map обратно в программу: {exc}", "error")
    return redirect(url_for("questions", **params))


@app.route("/questions/process", methods=["POST"])
def questions_process() -> Any:
    log_event('questions', 'process_queue_requested', tenant_id=_current_tenant_for_logs())
    params = _questions_params_from_source(request.form)
    try:
        max_items = _safe_int(request.form.get("max_items", 0), 0)
        auto_only = bool(request.form.get("auto_only"))
        _submit_background_task("questions_process", "Обработка очереди вопросов", process_question_queue, max(0, max_items), auto_only=auto_only, unique_key="questions_process")
    except Exception as exc:
        flash(f"Не удалось поставить обработку очереди вопросов в фон: {exc}", "error")
    return redirect(url_for("questions", **params))


@app.route("/questions/auto-run", methods=["POST"])
def questions_auto_run() -> Any:
    log_event('questions', 'auto_run_requested', tenant_id=_current_tenant_for_logs())
    params = _questions_params_from_source(request.form)
    try:
        limit = _safe_int(request.form.get("max_items", 0), 0)
        send_now = bool(request.form.get("send_now"))
        _submit_background_task("questions_auto_run", "Автообработка вопросов", process_auto_question_rules, limit=max(0, limit), send_now=send_now, unique_key="questions_auto_run")
    except Exception as exc:
        flash(f"Не удалось поставить автообработку вопросов в фон: {exc}", "error")
    return redirect(url_for("questions", **params))


@app.route("/questions/prompt", methods=["POST"])
def questions_prompt() -> Any:
    log_event('questions', 'prompt_save_requested', tenant_id=_current_tenant_for_logs())
    params = _questions_params_from_source(request.form)
    try:
        save_question_prompt_text(request.form.get("question_prompt_text", ""))
        flash("Общий промпт для ответов на вопросы сохранён.", "success")
    except Exception as exc:
        flash(f"Не удалось сохранить промпт вопросов: {exc}", "error")
    return redirect(url_for("questions", **params))


@app.route("/questions/rule/save", methods=["POST"])
def questions_rule_save() -> Any:
    log_event('questions', 'rule_save_requested', tenant_id=_current_tenant_for_logs())
    params = _questions_params_from_source(request.form)
    try:
        payload = parse_rule_form(request.form, request.files)
        if not clean_text(payload.get("cluster_key")):
            raise ValueError("Не указан кластер для правила.")
        rule = upsert_question_rule(**payload)
        flash(f"Правило «{rule.get('title') or rule.get('id')}» сохранено.", "success")
        if clean_text(request.form.get("after_save")) == "generate_cluster":
            selected_ids = list_question_ids_for_cluster(clean_text(payload.get("cluster_key")), hide_submitted=False)
            snapshot = get_question_snapshot(force_refresh=False)
            snapshot_map = {
                clean_text(item.get("id")): common.normalize_question(item)
                for item in snapshot.get("questions") or []
                if clean_text(item.get("id"))
            }
            generated = 0
            for question_id in selected_ids:
                question = snapshot_map.get(question_id)
                if not question:
                    continue
                generate_question_draft(question, force=True)
                generated += 1
            flash(f"Черновики по кластеру пересобраны: {generated}.", "success")
    except Exception as exc:
        flash(f"Не удалось сохранить правило: {exc}", "error")
    return redirect(url_for("questions", **params))


@app.route("/questions/rule/toggle", methods=["POST"])
def questions_rule_toggle() -> Any:
    log_event('questions', 'rule_toggle_requested', tenant_id=_current_tenant_for_logs())
    params = _questions_params_from_source(request.form)
    rule_id = clean_text(request.form.get("rule_id"))
    enabled = bool(request.form.get("enabled"))
    if not rule_id:
        flash("Не указан идентификатор правила.", "error")
        return redirect(url_for("questions", **params))
    changed = toggle_question_rule(rule_id, enabled)
    if changed:
        flash("Статус правила обновлён.", "success")
    else:
        flash("Правило не найдено.", "error")
    return redirect(url_for("questions", **params))


@app.route("/questions/cluster/reassign", methods=["POST"])
def questions_cluster_reassign() -> Any:
    log_event('questions', 'cluster_reassign_requested', tenant_id=_current_tenant_for_logs())
    params = _questions_params_from_source(request.form)
    selected_ids = _resolve_question_ids(request.form, allow_cluster_scope=False)
    target_title = clean_text(request.form.get("new_cluster_title"))
    if not selected_ids:
        flash("Выберите вопросы для переноса в подкластер.", "error")
        return redirect(url_for("questions", **params))
    if not target_title:
        flash("Укажите название нового подкластера.", "error")
        return redirect(url_for("questions", **params))
    try:
        cluster_key = reassign_cluster_members(selected_ids, target_title)
        flash(f"Вопросы перенесены в подкластер «{target_title}».", "success")
        params["cluster"] = cluster_key
        params["mode"] = "questions"
    except Exception as exc:
        flash(f"Не удалось перенести вопросы: {exc}", "error")
    return redirect(url_for("questions", **params))


@app.route("/questions/cluster/reset", methods=["POST"])
def questions_cluster_reset() -> Any:
    log_event('questions', 'cluster_reset_requested', tenant_id=_current_tenant_for_logs())
    params = _questions_params_from_source(request.form)
    selected_ids = _resolve_question_ids(request.form, allow_cluster_scope=False)
    if not selected_ids:
        flash("Выберите вопросы, для которых нужно снять ручную группировку.", "error")
        return redirect(url_for("questions", **params))
    try:
        removed = reset_cluster_assignments(selected_ids)
        flash(f"Сброшено ручных назначений: {removed}.", "success")
    except Exception as exc:
        flash(f"Не удалось сбросить ручные назначения: {exc}", "error")
    return redirect(url_for("questions", **params))


@app.route("/questions/cluster/auto-split", methods=["POST"])
def questions_cluster_auto_split() -> Any:
    log_event('questions', 'cluster_auto_split_requested', tenant_id=_current_tenant_for_logs())
    params = _questions_params_from_source(request.form)
    cluster_key = clean_text(request.form.get("cluster_key") or request.form.get("cluster"))
    if not cluster_key:
        flash("Не указан кластер для авторазбиения.", "error")
        return redirect(url_for("questions", **params))
    try:
        moved = auto_split_cluster_by_article(cluster_key)
        flash(f"Авторазбиение выполнено. Перенесено вопросов: {moved}.", "success")
    except Exception as exc:
        flash(f"Не удалось автоматически разбить кластер: {exc}", "error")
    return redirect(url_for("questions", **params))


@app.route("/tenants")
def tenants() -> Any:
    summaries = tenant_manager.collect_tenant_summaries()
    active_id = clean_text(getattr(g, "active_tenant_id", "") or session.get("tenant_id"))
    return render_template(
        "tenants.html",
        tenants=summaries,
        active_tenant_id=active_id,
        has_any=bool(summaries),
        private_root=str(getattr(common, "PRIVATE_ROOT", Path(str(common.PRIVATE_DIR)).parent)),
    )


@app.route("/tenants/add", methods=["POST"])
def tenants_add() -> Any:
    try:
        tenant = tenant_manager.create_tenant(
            name=request.form.get("name", ""),
            phone=request.form.get("phone", ""),
            wb_api_key=request.form.get("wb_api_key", ""),
            tenant_slug=request.form.get("tenant_slug", ""),
            emoji=request.form.get("emoji", ""),
        )
        session["tenant_id"] = tenant["id"]
        flash(f"Кабинет «{tenant['name']}» добавлен. Сохраните авторизацию WB для этого кабинета.", "success")
    except Exception as exc:
        flash(f"Не удалось добавить кабинет: {exc}", "error")
    return redirect(url_for("tenants", **_tenant_query_arg()))


@app.route("/tenants/select/<tenant_id>")
def tenant_select(tenant_id: str) -> Any:
    if not tenant_manager.get_tenant(tenant_id):
        flash("Кабинет не найден.", "error")
        return redirect(url_for("tenants"))
    session["tenant_id"] = tenant_id
    next_url = request.args.get("next") or url_for("index", tenant_id=tenant_id)
    return redirect(next_url)


@app.route("/tenants/<tenant_id>/update", methods=["POST"])
def tenant_update(tenant_id: str) -> Any:
    try:
        tenant_manager.update_tenant(
            tenant_id,
            name=clean_text(request.form.get("name")),
            phone=clean_text(request.form.get("phone")),
            emoji=clean_text(request.form.get("emoji")) or "🏢",
            wb_api_key=clean_text(request.form.get("wb_api_key")),
            enabled=bool(request.form.get("enabled")),
            notes=clean_text(request.form.get("notes")),
        )
        flash("Настройки кабинета сохранены.", "success")
    except Exception as exc:
        flash(f"Не удалось сохранить кабинет: {exc}", "error")
    return redirect(url_for("tenants", tenant_id=tenant_id))


@app.route("/tenants/<tenant_id>/health", methods=["POST"])
def tenant_health(tenant_id: str) -> Any:
    try:
        tenant_manager.refresh_tenant_health(tenant_id, force=True)
        flash("Проверка здоровья кабинета обновлена.", "success")
    except Exception as exc:
        flash(f"Не удалось проверить здоровье кабинета: {exc}", "error")
    return redirect(url_for("tenants", tenant_id=tenant_id))


@app.route("/tenants/health-all", methods=["POST"])
def tenant_health_all() -> Any:
    errors = 0
    for tenant in tenant_manager.load_tenants():
        try:
            tenant_manager.refresh_tenant_health(clean_text(tenant.get("id")), force=True)
        except Exception:
            errors += 1
    if errors:
        flash(f"Проверка здоровья выполнена, ошибок: {errors}.", "error")
    else:
        flash("Проверка здоровья всех кабинетов обновлена.", "success")
    return redirect(url_for("tenants", **_tenant_query_arg()))


@app.route("/tenants/<tenant_id>/login", methods=["POST"])
def tenant_login(tenant_id: str) -> Any:
    try:
        tenant_manager.spawn_login_for_tenant(tenant_id)
        flash("Открыт браузер для входа в кабинет WB. После логина вернитесь в панель и обновите страницу.", "success")
    except Exception as exc:
        flash(f"Не удалось запустить вход: {exc}", "error")
    return redirect(url_for("tenants", tenant_id=tenant_id))


@app.route("/tenants/<tenant_id>/reset", methods=["POST"])
def tenant_reset(tenant_id: str) -> Any:
    keep_auth = bool(request.form.get("keep_auth"))
    try:
        tenant_manager.delete_tenant_runtime_data(tenant_id, keep_auth=keep_auth)
        flash("Данные кабинета очищены.", "success")
    except Exception as exc:
        flash(f"Не удалось очистить данные кабинета: {exc}", "error")
    return redirect(url_for("tenants", tenant_id=tenant_id))


@app.route("/tenants/<tenant_id>/backup", methods=["POST"])
def tenant_backup(tenant_id: str) -> Any:
    try:
        zip_path = tenant_manager.backup_tenant(tenant_id)
        return send_file(zip_path, as_attachment=True, download_name=zip_path.name)
    except Exception as exc:
        flash(f"Не удалось создать backup: {exc}", "error")
        return redirect(url_for("tenants", tenant_id=tenant_id))


@app.route("/automation")
def automation() -> Any:
    settings = automation_core.load_settings()
    manifest = automation_core.build_workspace_manifest(settings)
    scheduler = automation_scheduler.scheduler_status()
    reports = automation_core.list_recent_reports(limit=30)
    system_jobs = background_jobs.list_jobs(tenant_id="_system", limit=30)
    workspace_health = price_pipeline.workspace_health(settings)
    next_runs = automation_core.next_runs(settings)
    tenant_summaries = tenant_manager.collect_tenant_summaries()
    return render_template(
        "automation.html",
        settings=settings,
        manifest=manifest,
        scheduler=scheduler,
        reports=reports,
        system_jobs=system_jobs,
        workspace_health=workspace_health,
        next_runs=next_runs,
        tenant_summaries=tenant_summaries,
    )


@app.route("/automation/settings", methods=["POST"])
def automation_settings_save() -> Any:
    settings = automation_core.load_settings()
    settings["schedule_enabled"] = _safe_bool_form(request.form.get("schedule_enabled"))
    settings["archive_runs"] = _safe_bool_form(request.form.get("archive_runs"))

    promo = settings.get("promo") or {}
    promo["enabled"] = _safe_bool_form(request.form.get("promo_enabled"))
    promo["mode"] = clean_text(request.form.get("promo_mode")) or promo.get("mode") or "manual"
    promo["schedule_time"] = clean_text(request.form.get("promo_schedule_time")) or promo.get("schedule_time") or "23:00"
    promo["window_days"] = _safe_int(request.form.get("promo_window_days"), int(promo.get("window_days") or 7))
    promo["future_only"] = _safe_bool_form(request.form.get("promo_future_only"))
    promo["all_promotions"] = _safe_bool_form(request.form.get("promo_all_promotions"))
    promo["include_auto"] = _safe_bool_form(request.form.get("promo_include_auto"))
    promo["include_regular"] = _safe_bool_form(request.form.get("promo_include_regular"))
    promo["max_retries"] = _safe_int(request.form.get("promo_max_retries"), int(promo.get("max_retries") or 3))
    promo["strategy"] = clean_text(request.form.get("promo_strategy")) or promo.get("strategy") or "api_then_browser"
    promo["verify_after_action"] = _safe_bool_form(request.form.get("promo_verify_after_action"))
    settings["promo"] = promo

    prices = settings.get("prices") or {}
    prices["enabled"] = _safe_bool_form(request.form.get("prices_enabled"))
    prices["mode"] = clean_text(request.form.get("prices_mode")) or prices.get("mode") or "manual"
    prices["schedule_time"] = clean_text(request.form.get("prices_schedule_time")) or prices.get("schedule_time") or "01:00"
    prices["master_filename"] = clean_text(request.form.get("master_filename")) or prices.get("master_filename") or "master_prices.xlsm"
    prices["master_sheet_name"] = clean_text(request.form.get("master_sheet_name"))
    prices["template_sheet_name"] = clean_text(request.form.get("template_sheet_name"))
    prices["template_pattern"] = clean_text(request.form.get("template_pattern")) or prices.get("template_pattern") or "price_template__{tenant_id}.xlsx"
    prices["output_pattern"] = clean_text(request.form.get("output_pattern")) or prices.get("output_pattern") or "{date}__{tenant_id}__prices{ext}"
    prices["warn_change_pct"] = request.form.get("warn_change_pct") or prices.get("warn_change_pct") or 30
    prices["row_start"] = _safe_int(request.form.get("row_start"), int(prices.get("row_start") or 2))
    for field in ["master_article_col", "master_price_col", "master_discount_col", "template_article_col", "template_price_col", "template_discount_col"]:
        value = clean_text(request.form.get(field))
        if value:
            prices[field] = value.upper()
    prices["recalc_mode"] = clean_text(request.form.get("recalc_mode")) or prices.get("recalc_mode") or "auto"
    prices["verify_via_api"] = _safe_bool_form(request.form.get("verify_via_api"))
    prices["upload_via_browser"] = _safe_bool_form(request.form.get("upload_via_browser"))
    settings["prices"] = prices

    tenants_cfg = settings.get("tenants") or {}
    for tenant in tenant_manager.load_tenants():
        tenant_id = clean_text(tenant.get("id"))
        if not tenant_id:
            continue
        entry = tenants_cfg.get(tenant_id) if isinstance(tenants_cfg.get(tenant_id), dict) else {}
        entry["enabled"] = _safe_bool_form(request.form.get(f"tenant__{tenant_id}__enabled"))
        entry["promo_enabled"] = _safe_bool_form(request.form.get(f"tenant__{tenant_id}__promo_enabled"))
        entry["price_enabled"] = _safe_bool_form(request.form.get(f"tenant__{tenant_id}__price_enabled"))
        entry["template_filename"] = clean_text(request.form.get(f"tenant__{tenant_id}__template_filename")) or entry.get("template_filename") or automation_core.expected_template_filename(tenant_id, settings)
        entry["notes"] = clean_text(request.form.get(f"tenant__{tenant_id}__notes"))
        tenants_cfg[tenant_id] = entry
    settings["tenants"] = tenants_cfg

    try:
        automation_core.save_settings(settings)
        flash("Настройки ночной автоматизации сохранены.", "success")
    except Exception as exc:
        flash(f"Не удалось сохранить настройки автоматизации: {exc}", "error")
    return redirect(url_for("automation", **_tenant_query_arg()))


@app.route("/automation/promo/scan", methods=["POST"])
def automation_promo_scan() -> Any:
    tenant_ids = _selected_automation_tenants_from_form("promo")
    if not tenant_ids:
        flash("Нет кабинетов для сканирования будущих акций.", "error")
        return redirect(url_for("automation", **_tenant_query_arg()))
    _submit_system_background_task(
        kind="promo_scan",
        label="Скан будущих акций",
        target=promo_calendar.scan_future_promotions,
        unique_key="promo_scan",
        tenant_ids=tenant_ids,
        run_source="manual",
    )
    return redirect(url_for("automation", **_tenant_query_arg()))


@app.route("/automation/prices/build", methods=["POST"])
def automation_prices_build() -> Any:
    tenant_ids = _selected_automation_tenants_from_form("prices")
    if not tenant_ids:
        flash("Нет кабинетов для подготовки файлов цен.", "error")
        return redirect(url_for("automation", **_tenant_query_arg()))
    _submit_system_background_task(
        kind="prices_build",
        label="Подготовка файлов цен",
        target=price_pipeline.build_price_files,
        unique_key="prices_build",
        tenant_ids=tenant_ids,
        run_source="manual",
    )
    return redirect(url_for("automation", **_tenant_query_arg()))


@app.route("/automation/promo/execute", methods=["POST"])
def automation_promo_execute() -> Any:
    tenant_ids = _selected_automation_tenants_from_form("promo")
    if not tenant_ids:
        flash("Нет кабинетов для снятия будущих акций.", "error")
        return redirect(url_for("automation", **_tenant_query_arg()))
    _submit_system_background_task(
        kind="promo_execute",
        label="Снятие будущих акций",
        target=promo_executor.execute_future_promotions,
        unique_key="promo_execute",
        tenant_ids=tenant_ids,
        run_source="manual",
    )
    return redirect(url_for("automation", **_tenant_query_arg()))


@app.route("/automation/prices/upload", methods=["POST"])
def automation_prices_upload() -> Any:
    tenant_ids = _selected_automation_tenants_from_form("prices")
    if not tenant_ids:
        flash("Нет кабинетов для загрузки цен.", "error")
        return redirect(url_for("automation", **_tenant_query_arg()))
    released = background_jobs.abandon_running_jobs(
        "prices_upload",
        tenant_id="_system",
        unique_key="prices_upload",
        older_than_seconds=300,
        progress_at_least=95.0,
        message="Предыдущая загрузка цен была автоматически освобождена после зависания на 95% перед новым ручным запуском",
    )
    if released:
        flash(f"Освобождено зависших задач загрузки цен: {len(released)}.", "warning")
    _submit_system_background_task(
        kind="prices_upload",
        label="Загрузка цен в WB",
        target=price_uploader.run_price_upload_cycle,
        unique_key="prices_upload",
        tenant_ids=tenant_ids,
        run_source="manual",
        rebuild=True,
    )
    return redirect(url_for("automation", **_tenant_query_arg()))


@app.route("/automation/download")
def automation_download() -> Any:
    raw_path = clean_text(request.args.get("path"))
    if not raw_path:
        flash("Не указан путь к файлу отчёта.", "error")
        return redirect(url_for("automation", **_tenant_query_arg()))
    try:
        target = Path(raw_path).expanduser().resolve()
        root = automation_core.AUTOMATION_ROOT.resolve()
        target.relative_to(root)
    except Exception:
        flash("Запрошен недопустимый путь к файлу автоматизации.", "error")
        return redirect(url_for("automation", **_tenant_query_arg()))
    if not target.exists() or not target.is_file():
        flash("Файл не найден.", "error")
        return redirect(url_for("automation", **_tenant_query_arg()))
    return send_file(target, as_attachment=True, download_name=target.name)


@app.route("/diagnostics")
def diagnostics() -> Any:
    tenant_filter = clean_text(request.args.get("tenant")) or (clean_text(getattr(g, "active_tenant_id", "")) or "all")
    channel_filter = clean_text(request.args.get("channel")) or "all"
    limit = max(50, min(5000, _safe_int(request.args.get("limit", 300), 300)))
    logs_tenant_arg = None if tenant_filter == "all" else tenant_filter
    entries = list(reversed(safe_read_events(logs_tenant_arg, channel_filter, limit)))
    stats = safe_log_stats(logs_tenant_arg, channel_filter, min(limit, 1000))
    channel_options = ["all"] + safe_list_channels(logs_tenant_arg)
    tenant_options = ["all"] + sorted({clean_text(t.get("id")) for t in tenant_manager.load_tenants()} | set(safe_list_tenants()))
    return render_template(
        "diagnostics.html",
        entries=entries,
        stats=stats,
        channel_filter=channel_filter,
        tenant_filter=tenant_filter,
        channel_options=channel_options,
        tenant_options=tenant_options,
        limit=limit,
    )


if __name__ == "__main__":
    started = automation_scheduler.start_scheduler()
    if started:
        import atexit
        atexit.register(automation_scheduler.stop_scheduler)
    app.run(debug=bool(getattr(config, "FLASK_DEBUG", False)))
