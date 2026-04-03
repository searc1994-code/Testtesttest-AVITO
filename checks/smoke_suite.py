from __future__ import annotations

import json
import os
import shutil
import sys
import tempfile
import threading
import time
import types
from datetime import timedelta, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

ROOT = Path(__file__).resolve().parents[1]
TEMP_ROOT = Path(tempfile.mkdtemp(prefix='wb-merged-smoke-'))
os.environ['WB_PRIVATE_DIR'] = str(TEMP_ROOT / 'private')
os.environ.pop('APP_ADMIN_PASSWORD', None)
os.environ.pop('WB_API_KEY', None)
(TEMP_ROOT / 'private' / 'security').mkdir(parents=True, exist_ok=True)
(TEMP_ROOT / 'private' / 'security' / 'openai_api_key.txt').write_text('smoke-openai-key\n', encoding='utf-8')
(TEMP_ROOT / 'private' / 'security' / 'openai_base_url.txt').write_text('https://api.hydraai.ru/v1\n', encoding='utf-8')
(TEMP_ROOT / 'private' / 'security' / 'openai_model.txt').write_text('gpt-4o-mini\n', encoding='utf-8')
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

# Fresh imports after env bootstrap.
for mod_name in [
    'config', 'safe_files', 'tenant_manager', 'common', 'background_jobs',
    'history_sync_worker', 'history_service', 'complaint_core', 'question_core',
    'auth_core', 'browser_bot', 'storage_paths', 'automation_core',
    'automation_scheduler', 'price_pipeline', 'promo_calendar', 'automation_browser',
    'price_uploader', 'promo_executor', 'wb_price_api', 'ui_forensics', 'app', 'web_security', 'module_runs'
]:
    if mod_name in sys.modules:
        del sys.modules[mod_name]

import auth_core

flask_stub = types.ModuleType('flask')


class DummyFlask:
    def __init__(self, name: str):
        self.name = name
        self.routes: list[dict] = []
        self.before_request_funcs = []
        self.after_request_funcs = []
        self.teardown_request_funcs = []
        self.context_processors = []
        self.config = {}
        self.secret_key = ''

    def route(self, rule: str, methods=None, **kwargs):
        methods = set(methods or ['GET'])
        def decorator(func):
            self.routes.append({'rule': rule, 'methods': methods, 'endpoint': func.__name__})
            return func
        return decorator

    def before_request(self, func):
        self.before_request_funcs.append(func)
        return func

    def after_request(self, func):
        self.after_request_funcs.append(func)
        return func

    def teardown_request(self, func):
        self.teardown_request_funcs.append(func)
        return func

    def context_processor(self, func):
        self.context_processors.append(func)
        return func

    def run(self, *args, **kwargs):
        return None


flask_stub.Flask = DummyFlask
flask_stub.flash = lambda *args, **kwargs: None
flask_stub.g = types.SimpleNamespace()
flask_stub.jsonify = lambda payload=None, **kwargs: payload if payload is not None else kwargs
flask_stub.redirect = lambda url: ('redirect', url)
flask_stub.render_template = lambda template_name, **context: {'template': template_name, 'context': context}
flask_stub.request = types.SimpleNamespace(method='GET', args={}, form={}, values={}, headers={}, endpoint='', host_url='http://localhost/', full_path='/', path='/', query_string=b'')
flask_stub.send_file = lambda *args, **kwargs: {'send_file': True}
flask_stub.session = {}
flask_stub.url_for = lambda endpoint, **values: f'/{endpoint}'
sys.modules.setdefault('flask', flask_stub)

import app as appmod
import automation_browser
import automation_core
import automation_scheduler
import background_jobs
import browser_bot
import common
import config
import complaint_core
import history_service
import price_pipeline
import price_uploader
import promo_calendar
import promo_executor
import question_core
import safe_files
import storage_paths
import tenant_manager
import wb_price_api
import ui_forensics
import module_runs
import web_security


def _wait_job(job_id: str, timeout: float = 8.0) -> dict:
    started = time.time()
    while time.time() - started < timeout:
        current = background_jobs.get_job(job_id) or {}
        if current.get('status') in {'completed', 'error', 'abandoned'}:
            return current
        time.sleep(0.1)
    return background_jobs.get_job(job_id) or {}


# 0. Legacy private-root autodiscovery should prefer old C-drive-like storage and reuse security files.
legacy_root = TEMP_ROOT / 'legacy-root'
legacy_home = TEMP_ROOT / 'home-root'
legacy_home_root = legacy_home / 'wb-ai-private'
legacy_root.mkdir(parents=True, exist_ok=True)
legacy_home_root.mkdir(parents=True, exist_ok=True)
safe_files.write_json(legacy_root / 'tenants.json', [{'id': 'legacy', 'name': 'Legacy tenant'}])
safe_files.write_text(legacy_home_root / 'security' / 'admin_auth.json', json.dumps({'username': 'admin'}), encoding='utf-8')
resolved_root = storage_paths.resolve_private_root(env_value='', os_name='nt', home=legacy_home, windows_legacy_root=legacy_root)
assert resolved_root == legacy_root
copied_security = storage_paths.hydrate_security_files(resolved_root, storage_paths.sibling_private_roots(resolved_root, env_value='', os_name='nt', home=legacy_home, windows_legacy_root=legacy_root))
assert (resolved_root / 'security' / 'admin_auth.json').exists()
assert copied_security

# 1. Auth bootstrap / first run password setup.
assert auth_core.needs_bootstrap() is True
record = auth_core.bootstrap_admin_password('VeryStrongPass123', confirm_password='VeryStrongPass123')
assert record['username'] == 'admin'
assert auth_core.needs_bootstrap() is False
assert auth_core.verify_credentials('admin', 'VeryStrongPass123') is True
assert auth_core.verify_credentials('admin', 'WrongPass') is False

assert config.SESSION_COOKIE_SECURE is True
security_status = web_security.login_rate_status('admin', '127.0.0.1')
assert security_status['allowed'] is True
for _ in range(config.LOGIN_RATE_MAX_ATTEMPTS):
    security_status = web_security.record_login_failure('admin', '127.0.0.1')
blocked_status = web_security.login_rate_status('admin', '127.0.0.1')
assert blocked_status['allowed'] is False
assert blocked_status['retry_after_seconds'] > 0
web_security.clear_login_failures('admin', '127.0.0.1')
assert web_security.login_rate_status('admin', '127.0.0.1')['allowed'] is True
csp_header = web_security.build_csp_header()
assert "script-src 'self' 'nonce-" in csp_header
assert "script-src-attr 'none'" in csp_header


logger = module_runs.ModuleRunLogger('checks', 'smoke', tenant_id='alpha', request_id='req-check', job_id='job-check')
logger.write_input('sample', {'hello': 'world'})
logger.event('sample_event', stage='sample_stage', answer=42)
summary = logger.finalize('completed', message='ok')
assert summary['status'] == 'completed'
assert module_runs.BASE_DIR.exists()

assert config.OPENAI_API_KEY == 'smoke-openai-key'
assert config.OPENAI_API_KEY_SOURCE.startswith('file:')
ai_diag = common.get_ai_runtime_diagnostics()
assert ai_diag['key_present'] is True
assert ai_diag['key_source'].startswith('file:')

reply_draft = {'signature': 'review-sig', 'reply': 'fallback reply', 'source': 'fallback', 'ai_error_code': 'config_missing'}
assert appmod.is_draft_visible(reply_draft, 'review-sig') is True
assert appmod.is_draft_compatible(reply_draft, 'review-sig', 'prompt-sig') is False

question_draft = {
    'signature': 'question-sig',
    'reply': 'fallback reply',
    'source': 'fallback',
    'ai_error_code': 'config_missing',
    'cluster_key': 'other',
    'prompt_signature': 'prompt-sig',
    'rules_signature': 'rules-sig',
}
assert question_core.is_question_draft_compatible(question_draft, 'question-sig', 'prompt-sig', 'rules-sig', 'other') is False
assert question_core.is_question_draft_compatible({**question_draft, 'source': 'manual_edit'}, 'question-sig', 'prompt-sig', 'rules-sig', 'other') is True

helper_logger = module_runs.ModuleRunLogger('checks', 'finalize-helper', tenant_id='alpha')
helper_summary = appmod._logger_finalize_from_result(helper_logger, 'completed', {'message': 'done', 'sent': 1})
assert helper_summary['message'] == 'done'
assert helper_summary['sent'] == 1

# 2. Tenant isolation + malicious tenant id rejection.
alpha = tenant_manager.create_tenant('Alpha LLC', '+79990000001', 'token-alpha', tenant_slug='alpha')
beta = tenant_manager.create_tenant('Beta LLC', '+79990000002', 'token-beta', tenant_slug='beta')
for bad_tenant in ('../../evil', '<script>alert(1)</script>', 'a' * 81):
    try:
        tenant_manager.ensure_tenant_dirs(bad_tenant)
        raise AssertionError(f'Invalid tenant id was accepted: {bad_tenant!r}')
    except ValueError:
        pass

for tenant_id, text in [('alpha', 'draft-alpha'), ('beta', 'draft-beta')]:
    tenant = tenant_manager.get_tenant(tenant_id)
    paths = tenant_manager.ensure_tenant_dirs(tenant_id)
    tokens = common.bind_tenant_context(tenant_id, tenant=tenant, paths=paths)
    try:
        common.write_json(common.DRAFTS_FILE, {'row': {'reply': text}})
        common.write_json(common.REPLY_QUEUE_FILE, [{'id': tenant_id, 'status': 'queued'}])
    finally:
        common.reset_tenant_context(tokens)

for tenant_id, expected in [('alpha', 'draft-alpha'), ('beta', 'draft-beta')]:
    tenant = tenant_manager.get_tenant(tenant_id)
    paths = tenant_manager.ensure_tenant_dirs(tenant_id)
    tokens = common.bind_tenant_context(tenant_id, tenant=tenant, paths=paths)
    try:
        data = common.read_json(common.DRAFTS_FILE, {})
        assert data['row']['reply'] == expected
    finally:
        common.reset_tenant_context(tokens)

# 3. Safe file writes under concurrency.
atomic_target = TEMP_ROOT / 'atomic.json'


def _atomic_writer(name: str, rounds: int = 20):
    for index in range(rounds):
        safe_files.write_json(atomic_target, {'worker': name, 'round': index})


threads = [threading.Thread(target=_atomic_writer, args=('alpha',)), threading.Thread(target=_atomic_writer, args=('beta',))]
for thread in threads:
    thread.start()
for thread in threads:
    thread.join(timeout=5)
final_atomic = safe_files.read_json(atomic_target, {})
assert final_atomic.get('worker') in {'alpha', 'beta'}
assert isinstance(final_atomic.get('round'), int)

# 4. Background jobs: dedupe, completion, failure, restart recovery.

def slow_job(delay: float = 0.2):
    time.sleep(delay)
    return {'ok': True}


job1, created1 = background_jobs.submit_job(kind='questions_process', tenant_id='alpha', label='Questions process', target=slow_job, kwargs={'delay': 0.5}, unique_key='questions_process')
job2, created2 = background_jobs.submit_job(kind='questions_process', tenant_id='alpha', label='Questions process', target=slow_job, kwargs={'delay': 0.5}, unique_key='questions_process')
assert created1 is True
assert created2 is False
assert job1['job_id'] == job2['job_id']
assert _wait_job(job1['job_id']).get('status') == 'completed'


def bad_job():
    raise RuntimeError('boom')


job_err, created_err = background_jobs.submit_job(kind='complaints_process', tenant_id='alpha', label='Complaints process', target=bad_job, unique_key='complaints_process')
assert created_err is True
assert _wait_job(job_err['job_id']).get('status') == 'error'

recent_ts = common.utc_now_iso()
stale_jobs_path = tenant_manager.ensure_tenant_dirs('alpha')['jobs_file']
stale_jobs = tenant_manager.read_json(stale_jobs_path, [])
stale_jobs.append({
    'job_id': 'restart-job-1',
    'tenant_id': 'alpha',
    'kind': 'replies_process',
    'label': 'Restarted running job',
    'unique_key': 'replies_process',
    'singleton_key': 'replies_process::alpha::replies_process',
    'status': 'running',
    'created_at': recent_ts,
    'started_at': recent_ts,
    'updated_at': recent_ts,
    'last_heartbeat_at': recent_ts,
    'last_message': 'running before restart',
    'result': None,
    'error': '',
    'progress': {},
})
tenant_manager.write_json(stale_jobs_path, stale_jobs)
assert any(job.get('job_id') == 'restart-job-1' and job.get('status') == 'abandoned' for job in background_jobs.list_jobs('alpha', limit=50))

# 5. Concurrent background jobs across tenants should not mix context.

def tenant_write_job(label: str, delay: float = 0.2):
    common.write_json(common.DRAFTS_FILE, {'row': {'reply': label}})
    time.sleep(delay)
    return {'reply': label}


alpha_job, alpha_created = background_jobs.submit_job(kind='tenant_write', tenant_id='alpha', label='Alpha write', target=tenant_write_job, kwargs={'label': 'tenant-alpha', 'delay': 0.25}, unique_key='tenant_write')
beta_job, beta_created = background_jobs.submit_job(kind='tenant_write', tenant_id='beta', label='Beta write', target=tenant_write_job, kwargs={'label': 'tenant-beta', 'delay': 0.25}, unique_key='tenant_write')
assert alpha_created is True and beta_created is True
assert _wait_job(alpha_job['job_id']).get('status') == 'completed'
assert _wait_job(beta_job['job_id']).get('status') == 'completed'

for tenant_id, expected in [('alpha', 'tenant-alpha'), ('beta', 'tenant-beta')]:
    tenant = tenant_manager.get_tenant(tenant_id)
    paths = tenant_manager.ensure_tenant_dirs(tenant_id)
    tokens = common.bind_tenant_context(tenant_id, tenant=tenant, paths=paths)
    try:
        data = common.read_json(common.DRAFTS_FILE, {})
        assert data['row']['reply'] == expected
    finally:
        common.reset_tenant_context(tokens)

# 6. History sync database lifecycle.
history_service.ensure_db('alpha')
snapshot = {
    'fetched_at': common.utc_now_iso(),
    'count_unanswered': 1,
    'feedbacks': [
        {
            'id': 'review-1',
            'text': 'Товар хороший',
            'pros': 'качество',
            'cons': '',
            'productValuation': 5,
            'createdDate': '2025-01-01T00:00:00+00:00',
            'userName': 'Иван',
            'subjectName': 'Категория',
            'productDetails': {'productName': 'Товар', 'supplierArticle': 'ART-1', 'brandName': 'Brand', 'nmId': 101},
        }
    ],
}
inserted = history_service.upsert_active_snapshot(snapshot, 'alpha')
assert inserted == 1
counts = history_service.get_counts('alpha')
assert counts['total'] == 1 and counts['needs_reply'] == 1
history_service.mark_replied('review-1', 'Спасибо за отзыв', 'alpha')
counts = history_service.get_counts('alpha')
assert counts['needs_reply'] == 0

# 7. Complaints results analytics.
paths = tenant_manager.ensure_tenant_dirs('alpha')
tokens = common.bind_tenant_context('alpha', tenant=tenant_manager.get_tenant('alpha'), paths=paths)
try:
    complaint_core.append_result({'review_id': 'c1', 'status': 'accepted', 'category': 'Другое', 'review': {'product_name': 'Товар'}})
    complaint_core.append_result({'review_id': 'c2', 'status': 'rejected', 'category': 'Другое', 'review': {'product_name': 'Товар'}})
    complaint_core.append_result({'review_id': 'c3', 'outcome': 'pending', 'status': 'submitted', 'category': 'Другое', 'review': {'product_name': 'Товар'}})
    metrics = complaint_core.build_complaint_effectiveness(limit=100)
    assert metrics['accepted'] == 1
    assert metrics['rejected'] == 1
    assert metrics['pending'] >= 1
finally:
    common.reset_tenant_context(tokens)

# 8. Question queue prepare/process without live WB network.
paths = tenant_manager.ensure_tenant_dirs('alpha')
tokens = common.bind_tenant_context('alpha', tenant=tenant_manager.get_tenant('alpha'), paths=paths)
try:
    question = {
        'id': 'q-1',
        'text': 'Подойдет ли для зимы?',
        'createdDate': '2025-01-01T00:00:00+00:00',
        'state': 'none',
        'subjectName': 'Категория',
        'answer': None,
        'productDetails': {'productName': 'Куртка', 'supplierArticle': 'JACKET-1', 'brandName': 'Brand', 'nmId': 777, 'size': 'M', 'imtId': 555, 'supplierName': 'Seller'},
        'wasViewed': False,
        'isWarned': False,
    }
    common.write_json(common.QUESTION_SNAPSHOT_FILE, {'questions': [question], 'count_unanswered': 1})
    common.write_json(common.QUESTION_DRAFTS_FILE, {
        'q-1': {
            'reply': 'Да, модель рассчитана на зимний сезон и хорошо сохраняет тепло.',
            'action': 'answer',
            'manual_action': 'answer',
            'manager_comment': '',
            'confidence': 0.95,
            'auto_ready': True,
            'cluster_key': 'other',
            'cluster_title': 'Прочее',
            'signature': question_core.question_signature(common.normalize_question(question)),
            'source': 'manual',
            'needs_regeneration': False,
        }
    })

    class Form:
        def getlist(self, name: str):
            if name == 'selected_ids':
                return ['q-1']
            return []

        def get(self, name: str, default=None):
            values = {
                'action__q-1': 'answer',
                'reply__q-1': 'Да, модель рассчитана на зимний сезон и хорошо сохраняет тепло.',
                'manager_comment__q-1': '',
            }
            return values.get(name, default)

    original_get_snapshot = question_core.get_question_snapshot
    original_refresh_snapshot = question_core.refresh_question_snapshot
    original_patch = common.patch_question
    question_core.get_question_snapshot = lambda force_refresh=False: common.read_json(common.QUESTION_SNAPSHOT_FILE, {})
    question_core.refresh_question_snapshot = lambda: common.read_json(common.QUESTION_SNAPSHOT_FILE, {})
    added, notes = question_core.queue_questions_from_form(Form())
    assert added == 1
    sent_payloads = []
    common.patch_question = lambda payload: sent_payloads.append(payload) or {'ok': True}
    try:
        result = question_core.process_question_queue(max_items=10, auto_only=False)
    finally:
        question_core.get_question_snapshot = original_get_snapshot
        question_core.refresh_question_snapshot = original_refresh_snapshot
        common.patch_question = original_patch
    assert result['sent'] == 1
    assert sent_payloads and sent_payloads[0]['id'] == 'q-1'
finally:
    common.reset_tenant_context(tokens)

# 9. Browser security helpers.
assert browser_bot._is_allowed_navigation_url('https://seller.wildberries.ru/reviews') is True
assert browser_bot._is_allowed_navigation_url('https://example.com') is False
assert browser_bot._is_allowed_request_url('https://static-basket-01.wb.ru') is True
assert browser_bot._is_allowed_request_url('https://malicious.example.com/evil.js') is False
try:
    browser_bot._assert_allowed_navigation_url('https://example.com/phishing')
    raise AssertionError('Disallowed navigation URL was accepted')
except browser_bot.BrowserBotError:
    pass

# 10. Automation workspace + price build.
settings = automation_core.load_settings()
settings['prices']['enabled'] = True
settings['prices']['mode'] = 'manual'
settings['prices']['master_filename'] = 'master_prices.xlsx'
settings['tenants']['alpha']['template_filename'] = 'price_template__alpha.xlsx'
settings['tenants']['beta']['template_filename'] = 'price_template__beta.xlsx'
automation_core.save_settings(settings)

master_wb = Workbook()
master_ws = master_wb.active
master_ws['D1'] = 'Артикул'
master_ws['S1'] = 'Цена'
master_ws['U1'] = 'Скидка'
master_ws['D2'] = 'ART-1'
master_ws['S2'] = 1600
master_ws['U2'] = 12
master_ws['D3'] = 'ART-2'
master_ws['S3'] = 1700
master_ws['U3'] = 15
master_ws['D4'] = 'ART-3'
master_ws['S4'] = 1800
master_ws['U4'] = 20
master_path = automation_core.resolve_master_path(settings)
master_wb.save(master_path)

for tenant_id, rows in {
    'alpha': [('ART-1', 1000, 5), ('ART-2', 1000, 5)],
    'beta': [('ART-3', 1400, 10), ('ART-404', 999, 1)],
}.items():
    wb = Workbook()
    ws = wb.active
    ws['C1'] = 'Артикул'
    ws['J1'] = 'Цена'
    ws['L1'] = 'Скидка'
    row_index = 2
    for article, price, discount in rows:
        ws[f'C{row_index}'] = article
        ws[f'J{row_index}'] = price
        ws[f'L{row_index}'] = discount
        row_index += 1
    wb.save(automation_core.resolve_template_path(tenant_id, settings))

manifest = automation_core.build_workspace_manifest(settings)
assert manifest['master_exists'] is True
assert all(row['template_exists'] for row in manifest['tenants'])

# 10a. Master reader must stream rows and not use slow random-access .cell on read-only worksheets.
stream_master = automation_core.PRICE_WORKSPACE_DIR / 'stream_master.xlsx'
master_wb.save(stream_master)
stream_settings = {'prices': {'master_sheet_name': '', 'row_start': 2, 'master_article_col': 'D', 'master_price_col': 'S', 'master_discount_col': 'U'}}
_original_readonly_cell = ReadOnlyWorksheet.cell
def _readonly_cell_forbidden(self, *args, **kwargs):
    raise AssertionError('read_only_random_access_cell_used')
ReadOnlyWorksheet.cell = _readonly_cell_forbidden
try:
    streamed_rows, streamed_meta = price_pipeline._read_master_rows(stream_master, stream_settings, [])
finally:
    ReadOnlyWorksheet.cell = _original_readonly_cell
assert streamed_rows['ART-1']['price'] == 1600
assert streamed_rows['ART-3']['discount'] == 20
assert streamed_meta['rows_scanned'] == 3
assert price_pipeline._coerce_excel_value(2123.0502) == 2123
assert price_pipeline._coerce_excel_value(16.6667) == 17
assert price_pipeline._coerce_excel_value(510.5) == 511

price_result = price_pipeline.build_price_files(['alpha', 'beta'], run_source='smoke')
assert price_result['prepared'] == 2
assert price_result['failed'] == 0
assert Path(price_result['report_path']).exists()
assert Path(price_result['archive_path']).exists()
alpha_output = next(Path(item['output_path']) for item in price_result['results'] if item['tenant_id'] == 'alpha')
alpha_wb = load_workbook(alpha_output)
alpha_ws = alpha_wb.active
assert alpha_ws['J2'].value == 1600
assert alpha_ws['L2'].value == 12

# 11. Promo scan without live WB network.
original_list_promotions = promo_calendar._list_promotions
original_details_promotions = promo_calendar._details_promotions
try:
    future_start = (automation_core.now_local(settings) + timedelta(days=1)).astimezone(timezone.utc).isoformat().replace('+00:00', 'Z')
    future_end = (automation_core.now_local(settings) + timedelta(days=2)).astimezone(timezone.utc).isoformat().replace('+00:00', 'Z')
    promo_calendar._list_promotions = lambda *args, **kwargs: [
        {'id': 101, 'name': 'Auto future', 'startDateTime': future_start, 'endDateTime': future_end, 'type': 'auto'},
        {'id': 102, 'name': 'Regular empty', 'startDateTime': future_start, 'endDateTime': future_end, 'type': 'regular'},
    ]
    promo_calendar._details_promotions = lambda *args, **kwargs: [
        {'id': 101, 'name': 'Auto future', 'type': 'auto', 'startDateTime': future_start, 'endDateTime': future_end, 'inPromoActionTotal': 5, 'participationPercentage': 80},
        {'id': 102, 'name': 'Regular empty', 'type': 'regular', 'startDateTime': future_start, 'endDateTime': future_end, 'inPromoActionTotal': 0, 'participationPercentage': 0},
    ]
    promo_result = promo_calendar.scan_future_promotions(['alpha'], run_source='smoke')
finally:
    promo_calendar._list_promotions = original_list_promotions
    promo_calendar._details_promotions = original_details_promotions
assert promo_result['actionable_total'] == 1
assert Path(promo_result['report_path']).exists()
assert Path(promo_result['archive_path']).exists()

# 12. Automation browser helpers / upload-id extraction.
events = [
    {'json': {'data': {'uploadID': 12345}}},
    {'json': {'upload_id': 67890}},
]
assert automation_browser.extract_upload_ids(events) == [12345, 67890]
assert automation_browser.wait_for_upload_id(events, timeout_seconds=0.2, poll_interval=0.05) == 67890

price_profile = price_uploader._load_profile()
assert price_profile['upload_urls'] == [f"{common.WB_SELLER_BASE_URL}/discount-and-prices"]
assert "[data-testid='xlsx-action-open-test-id-button-interface']" in price_profile['open_upload_button_selectors']
open_dropdown_selectors = price_profile.get('open_upload_dropdown_selectors') or price_profile.get('upload_dropdown_menu_selectors') or []
open_option_selectors = price_profile.get('open_upload_dropdown_option_selectors') or price_profile.get('upload_dropdown_option_selectors') or []
open_option_texts = price_profile.get('open_upload_dropdown_option_texts') or price_profile.get('upload_dropdown_option_texts') or []
if open_dropdown_selectors:
    assert any('dropdown' in str(item).lower() for item in open_dropdown_selectors)
assert "[data-testid='xlsx-action-options-test-id-dropdown-option']" in open_option_selectors
assert 'Цены и скидки' in open_option_texts
assert "[data-testid='check-changes-warning-checkbox-test-id-checkbox-simple-input']" in price_profile['warning_checkbox_selectors']
assert price_uploader._body_indicates_prices_page('Цены и скидки Обновить через Excel', price_profile)
assert price_uploader._body_indicates_dashboard_redirect('Быстрый доступ к справочным материалам Баланс Задачи по магазину', price_profile)

# 13. Price upload cycle with fake browser + fake WB API verification.
settings = automation_core.load_settings()
settings['prices']['verify_via_api'] = True
settings['prices']['upload_via_browser'] = True
automation_core.save_settings(settings)

def fake_price_browser_handler(tenant_id: str, file_path: Path, settings: dict, run_dir: Path) -> dict:
    upload_id = 9000 + (1 if tenant_id == 'alpha' else 2)
    return {
        'tenant_id': tenant_id,
        'file_path': str(file_path),
        'upload_id': upload_id,
        'submitted': True,
        'success_hint': True,
        'network_events': [{'json': {'data': {'uploadID': upload_id}}}],
        'before': {},
        'after': {},
    }

original_poll_upload_until_processed = wb_price_api.poll_upload_until_processed
original_list_all_goods = wb_price_api.list_all_goods
original_list_quarantine_goods = wb_price_api.list_quarantine_goods
try:
    wb_price_api.poll_upload_until_processed = lambda api_key, upload_id, **kwargs: {
        'upload_id': upload_id,
        'status': 'done',
        'processing_data': {'status': 3},
        'processed_data': {'status': 3, 'successGoodsNumber': 1, 'overAllGoodsNumber': 1},
        'buffer_details': [],
        'history_details': [{'ok': True}],
    }
    wb_price_api.list_all_goods = lambda api_key, **kwargs: [
        {'vendorCode': 'ART-1', 'nmID': 101, 'sizes': [{'price': 1600}], 'discount': 12, 'editableSizePrice': False},
        {'vendorCode': 'ART-2', 'nmID': 102, 'sizes': [{'price': 1700}], 'discount': 15, 'editableSizePrice': False},
        {'vendorCode': 'ART-3', 'nmID': 103, 'sizes': [{'price': 1800}], 'discount': 20, 'editableSizePrice': False},
    ]
    wb_price_api.list_quarantine_goods = lambda api_key, **kwargs: []
    price_upload_result = price_uploader.run_price_upload_cycle(
        ['alpha', 'beta'],
        run_source='smoke',
        rebuild=False,
        build_summary=price_result,
        browser_handler=fake_price_browser_handler,
    )
finally:
    wb_price_api.poll_upload_until_processed = original_poll_upload_until_processed
    wb_price_api.list_all_goods = original_list_all_goods
    wb_price_api.list_quarantine_goods = original_list_quarantine_goods
assert price_upload_result['uploaded'] == 2
assert price_upload_result['failed'] == 0
assert price_upload_result['mismatched_total'] == 0
assert price_upload_result['quarantine_total'] == 0
assert Path(price_upload_result['report_path']).exists()
assert Path(price_upload_result['archive_path']).exists()

# 14. Promo execute with fake browser handler + post-check verification.
settings = automation_core.load_settings()
settings['promo']['verify_after_action'] = True
automation_core.save_settings(settings)
future_start = (automation_core.now_local(settings) + timedelta(days=1)).astimezone(timezone.utc).isoformat().replace('+00:00', 'Z')
future_end = (automation_core.now_local(settings) + timedelta(days=2)).astimezone(timezone.utc).isoformat().replace('+00:00', 'Z')
scan_before = {
    'rows': [
        {'tenant_id': 'alpha', 'actionable': [{'id': 501, 'name': 'Auto alpha', 'type': 'auto', 'startDateTime': future_start, 'endDateTime': future_end}], 'actionable_total': 1, 'new_actionable_ids': [501]},
        {'tenant_id': 'beta', 'actionable': [], 'actionable_total': 0, 'new_actionable_ids': []},
    ],
    'actionable_total': 1,
}
scan_after = {
    'rows': [
        {'tenant_id': 'alpha', 'actionable': [], 'actionable_total': 0, 'new_actionable_ids': []},
        {'tenant_id': 'beta', 'actionable': [], 'actionable_total': 0, 'new_actionable_ids': []},
    ],
    'actionable_total': 0,
}

def fake_promo_handler(tenant_id: str, tenant_scan: dict, run_dir: Path, settings: dict) -> dict:
    total = int(tenant_scan.get('actionable_total') or len(tenant_scan.get('actionable') or []))
    return {
        'tenant_id': tenant_id,
        'processed': total,
        'browser_success': total,
        'browser_failed': 0,
        'attempts': [{'promotion_id': item.get('id')} for item in (tenant_scan.get('actionable') or [])],
        'failures': [],
    }

original_scan_future_promotions = promo_calendar.scan_future_promotions
try:
    promo_calendar.scan_future_promotions = lambda tenant_ids=None, run_source='scheduler': scan_after if 'verify' in str(run_source) else scan_before
    promo_exec_result = promo_executor.execute_future_promotions(
        ['alpha', 'beta'],
        run_source='smoke',
        browser_handler=fake_promo_handler,
    )
finally:
    promo_calendar.scan_future_promotions = original_scan_future_promotions
assert promo_exec_result['browser_success_total'] == 1
assert promo_exec_result['browser_failed_total'] == 0
assert promo_exec_result['remaining_actionable_total'] == 0
assert Path(promo_exec_result['report_path']).exists()
assert Path(promo_exec_result['archive_path']).exists()


# 14. UI forensics bundle basics.
class _FakeWatcherPage:
    def wait_for_timeout(self, ms: int):
        return None

forensics_dir = TEMP_ROOT / 'forensics-run'
forensics = ui_forensics.UIForensics(forensics_dir, tenant_id='alpha', job='prices_upload', session_name='smoke')
forensics.set_flow_context(phase='smoke_phase', step_id='smoke_step', attempt=1, branch_id='unknown')
row = forensics.event('smoke_started', extra='ok')
assert row['run_id']
assert isinstance(row.get('seq'), int) and row['seq'] >= 1
assert row.get('phase') == 'smoke_phase'
probe = forensics.probe_locator('missing_control', None, note='not_found_expected')
assert probe['found'] is False
assert probe['actionability_status'] == 'missing'
timeline = forensics.watch_post_submit(
    _FakeWatcherPage(),
    'smoke_watch',
    lambda: {'state_name': 'page_ready_after_submit', 'terminal': True, 'confidence': 0.99, 'branch_id': 'without_checkbox'},
    duration_seconds=1.0,
    interval_seconds=0.2,
    capture_on_change=False,
    timing_points_seconds=(0.0,),
)
assert len(timeline) >= 1
summary_path = forensics.write_summary({'status': 'ok'})
assert Path(summary_path).exists()
assert forensics.events_path.exists()
assert safe_files.read_json(forensics.summary_path, {}).get('watch_ticks', 0) >= 1

# 15. Scheduler submits system promo_execute + prices_upload jobs.
settings = automation_core.load_settings()
settings['schedule_enabled'] = True
settings['promo']['enabled'] = True
settings['promo']['mode'] = 'auto'
settings['promo']['schedule_time'] = automation_core.now_local(settings).strftime('%H:%M')
settings['prices']['enabled'] = True
settings['prices']['mode'] = 'auto'
settings['prices']['schedule_time'] = automation_core.now_local(settings).strftime('%H:%M')
automation_core.save_settings(settings)
automation_core.save_state(automation_core.default_state())

original_execute_future_promotions = promo_executor.execute_future_promotions
original_run_price_upload_cycle = price_uploader.run_price_upload_cycle
try:
    promo_executor.execute_future_promotions = lambda tenant_ids=None, run_source='scheduler', **kwargs: {'message': 'scheduler-promo-ok', 'tenant_ids': tenant_ids or []}
    price_uploader.run_price_upload_cycle = lambda tenant_ids=None, run_source='scheduler', rebuild=True, **kwargs: {'message': 'scheduler-prices-ok', 'tenant_ids': tenant_ids or [], 'rebuild': rebuild}
    existing_ids = {job.get('job_id') for job in background_jobs.list_jobs('_system', limit=100)}
    automation_scheduler.scheduler_tick()
    system_jobs = background_jobs.list_jobs('_system', limit=50)
    promo_job = next((job for job in system_jobs if job.get('kind') == 'promo_execute' and job.get('job_id') not in existing_ids), None)
    prices_job = next((job for job in system_jobs if job.get('kind') == 'prices_upload' and job.get('job_id') not in existing_ids), None)
    assert promo_job is not None
    assert prices_job is not None
    assert _wait_job(promo_job['job_id']).get('status') == 'completed'
    assert _wait_job(prices_job['job_id']).get('status') == 'completed'
    automation_scheduler.scheduler_tick()
finally:
    promo_executor.execute_future_promotions = original_execute_future_promotions
    price_uploader.run_price_upload_cycle = original_run_price_upload_cycle

scheduler_state = automation_scheduler.scheduler_status()
assert scheduler_state['plans'].get('promo_daily', {}).get('last_job_id')
assert scheduler_state['plans'].get('prices_daily', {}).get('last_job_id')
print(json.dumps({
    'status': 'SMOKE_OK',
    'private_root': str(TEMP_ROOT),
    'auth_bootstrap': True,
    'tenant_isolation': True,
    'safe_file_writes': True,
    'background_jobs': True,
    'concurrency': True,
    'history_sync': True,
    'complaints': True,
    'questions': True,
    'browser_security': True,
    'automation': True,
}, ensure_ascii=False))

shutil.rmtree(TEMP_ROOT, ignore_errors=True)
