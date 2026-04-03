from __future__ import annotations

import time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import requests
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

import common

PRICES_UPLOAD_STATE_URL = "https://discounts-prices-api.wildberries.ru/api/v2/buffer/tasks"
PRICES_UPLOAD_DETAILS_URL = "https://discounts-prices-api.wildberries.ru/api/v2/buffer/goods/task"
PRICES_HISTORY_STATE_URL = "https://discounts-prices-api.wildberries.ru/api/v2/history/tasks"
PRICES_HISTORY_DETAILS_URL = "https://discounts-prices-api.wildberries.ru/api/v2/history/goods/task"
PRICES_LIST_URL = "https://discounts-prices-api.wildberries.ru/api/v2/list/goods/filter"
PRICES_QUARANTINE_URL = "https://discounts-prices-api.wildberries.ru/api/v2/quarantine/goods"
_REQUEST_DELAY_SECONDS = 0.65


class WBPriceApiError(RuntimeError):
    pass


def _clean(value: Any) -> str:
    return common.clean_text(value)


def _normalize_article(value: Any) -> str:
    return _clean(value).replace(" ", "").upper()


def _headers(api_key: str) -> Dict[str, str]:
    token = _clean(api_key)
    if not token:
        raise WBPriceApiError("Для проверки цен нужен WB API key у кабинета.")
    return {"Authorization": token}


def _request_json(
    method: str,
    url: str,
    api_key: str,
    *,
    params: Any = None,
    json_payload: Any = None,
    retries: int = 3,
    timeout: int = 25,
) -> Dict[str, Any]:
    session = requests.Session()
    last_error = ""
    for attempt in range(1, max(1, retries) + 1):
        try:
            response = session.request(
                method.upper(),
                url,
                headers=_headers(api_key),
                params=params,
                json=json_payload,
                timeout=max(10, int(timeout or 25)),
            )
            if response.status_code == 429:
                last_error = f"HTTP 429: {response.text[:200]}"
                time.sleep(_REQUEST_DELAY_SECONDS * attempt)
                continue
            if response.status_code == 404:
                return {"data": None, "error": False, "errorText": "", "http_status": 404}
            response.raise_for_status()
            payload = response.json()
            if isinstance(payload, dict) and payload.get("error"):
                raise WBPriceApiError(_clean(payload.get("errorText") or "WB API вернул ошибку"))
            if isinstance(payload, dict):
                payload["http_status"] = response.status_code
                return payload
            return {"data": payload, "error": False, "errorText": "", "http_status": response.status_code}
        except Exception as exc:
            last_error = _clean(exc)
            if attempt >= retries:
                break
            time.sleep(_REQUEST_DELAY_SECONDS * attempt)
    raise WBPriceApiError(last_error or f"Не удалось выполнить запрос WB API: {url}")


def _as_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except Exception:
            return None
    text = str(value).strip().replace("%", "").replace(" ", "")
    if not text:
        return None
    text = text.replace(",", ".")
    try:
        return float(text)
    except Exception:
        return None


def _sheet(workbook, name: str = ""):
    sheet_name = _clean(name)
    if sheet_name and sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    return workbook[workbook.sheetnames[0]]


def _column_index(value: str) -> int:
    try:
        return int(column_index_from_string((_clean(value) or "A").upper()))
    except Exception as exc:
        raise WBPriceApiError(f"Некорректная колонка Excel: {value!r}") from exc


def read_expected_prices(file_path: str | Path, settings: Dict[str, Any]) -> List[Dict[str, Any]]:
    path = Path(file_path)
    if not path.exists():
        raise WBPriceApiError(f"Файл для проверки не найден: {path}")
    wb = load_workbook(path, data_only=True, read_only=True, keep_vba=path.suffix.lower() == ".xlsm")
    prices_cfg = settings.get("prices") if isinstance(settings.get("prices"), dict) else {}
    sheet = _sheet(wb, prices_cfg.get("template_sheet_name"))
    row_start = max(1, int(prices_cfg.get("row_start") or 2))
    article_col = _column_index(prices_cfg.get("template_article_col") or "C")
    price_col = _column_index(prices_cfg.get("template_price_col") or "J")
    discount_col = _column_index(prices_cfg.get("template_discount_col") or "L")
    rows: List[Dict[str, Any]] = []
    for row_idx in range(row_start, sheet.max_row + 1):
        article = _normalize_article(sheet.cell(row=row_idx, column=article_col).value)
        if not article:
            continue
        rows.append(
            {
                "article": article,
                "price": _as_number(sheet.cell(row=row_idx, column=price_col).value),
                "discount": _as_number(sheet.cell(row=row_idx, column=discount_col).value),
                "row": row_idx,
            }
        )
    return rows


def list_all_goods(api_key: str, *, limit: int = 1000, max_pages: int = 100) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    offset = 0
    for _ in range(max_pages):
        payload = _request_json(
            "GET",
            PRICES_LIST_URL,
            api_key,
            params={"limit": limit, "offset": offset},
        )
        page = (((payload or {}).get("data") or {}).get("listGoods") or [])
        if not isinstance(page, list):
            page = []
        if not page:
            break
        rows.extend(item for item in page if isinstance(item, dict))
        if len(page) < limit:
            break
        offset += limit
        time.sleep(_REQUEST_DELAY_SECONDS)
    return rows


def list_quarantine_goods(api_key: str, *, limit: int = 1000, max_pages: int = 50) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    offset = 0
    for _ in range(max_pages):
        payload = _request_json(
            "GET",
            PRICES_QUARANTINE_URL,
            api_key,
            params={"limit": limit, "offset": offset},
        )
        page = (((payload or {}).get("data") or {}).get("quarantineGoods") or [])
        if not isinstance(page, list):
            page = []
        if not page:
            break
        rows.extend(item for item in page if isinstance(item, dict))
        if len(page) < limit:
            break
        offset += limit
        time.sleep(_REQUEST_DELAY_SECONDS)
    return rows


def _representative_price(item: Dict[str, Any]) -> Optional[float]:
    sizes = item.get("sizes") if isinstance(item.get("sizes"), list) else []
    for size in sizes:
        if not isinstance(size, dict):
            continue
        price = _as_number(size.get("price"))
        if price is not None:
            return price
    return _as_number(item.get("price"))


def verify_prices_against_file(api_key: str, file_path: str | Path, settings: Dict[str, Any]) -> Dict[str, Any]:
    expected_rows = read_expected_prices(file_path, settings)
    goods_rows = list_all_goods(api_key)
    goods_by_vendor = {_normalize_article(item.get("vendorCode")): item for item in goods_rows if _normalize_article(item.get("vendorCode"))}
    quarantine_rows = list_quarantine_goods(api_key)
    quarantine_by_nm = {int(item.get("nmID")): item for item in quarantine_rows if item.get("nmID") is not None}

    matched = 0
    exact_match = 0
    mismatched = 0
    missing = 0
    quarantined = 0
    samples: List[Dict[str, Any]] = []

    for row in expected_rows:
        article = row["article"]
        current = goods_by_vendor.get(article)
        if not current:
            missing += 1
            if len(samples) < 50:
                samples.append({"article": article, "status": "missing_in_api"})
            continue
        matched += 1
        current_price = _representative_price(current)
        current_discount = _as_number(current.get("discount"))
        nm_id = int(current.get("nmID") or 0)
        in_quarantine = nm_id in quarantine_by_nm
        if in_quarantine:
            quarantined += 1
        same_price = row["price"] is None or current_price == row["price"]
        same_discount = row["discount"] is None or current_discount == row["discount"]
        if same_price and same_discount and not in_quarantine:
            exact_match += 1
            continue
        mismatched += 1
        if len(samples) < 80:
            sample = {
                "article": article,
                "nmID": nm_id or None,
                "expected_price": row["price"],
                "expected_discount": row["discount"],
                "current_price": current_price,
                "current_discount": current_discount,
                "editable_size_price": bool(current.get("editableSizePrice")),
                "quarantine": in_quarantine,
            }
            if in_quarantine:
                sample["quarantine_detail"] = quarantine_by_nm.get(nm_id)
            samples.append(sample)

    return {
        "file_path": str(file_path),
        "rows_in_file": len(expected_rows),
        "goods_loaded": len(goods_rows),
        "matched": matched,
        "exact_match": exact_match,
        "mismatched": mismatched,
        "missing_in_api": missing,
        "quarantine_count": quarantined,
        "samples": samples,
    }


def _collect_upload_details(api_key: str, url: str, upload_id: int, *, limit: int = 1000, max_pages: int = 20) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    offset = 0
    for _ in range(max_pages):
        payload = _request_json(
            "GET",
            url,
            api_key,
            params={"uploadID": int(upload_id), "limit": limit, "offset": offset},
        )
        data = payload.get("data") if isinstance(payload, dict) else {}
        buffer_goods = []
        if isinstance(data, dict):
            buffer_goods = data.get("bufferGoods") or data.get("goods") or data.get("historyGoods") or []
        if not isinstance(buffer_goods, list):
            buffer_goods = []
        if not buffer_goods:
            break
        rows.extend(item for item in buffer_goods if isinstance(item, dict))
        if len(buffer_goods) < limit:
            break
        offset += limit
        time.sleep(_REQUEST_DELAY_SECONDS)
    return rows


def upload_snapshot_indicates_acceptance(snapshot: Dict[str, Any]) -> bool:
    if not isinstance(snapshot, dict):
        return False
    processing = snapshot.get("processing") if isinstance(snapshot.get("processing"), dict) else {}
    processed = snapshot.get("processed") if isinstance(snapshot.get("processed"), dict) else {}
    processing_data = snapshot.get("processing_data") if isinstance(snapshot.get("processing_data"), dict) else {}
    processed_data = snapshot.get("processed_data") if isinstance(snapshot.get("processed_data"), dict) else {}
    if snapshot.get("buffer_details") or snapshot.get("history_details"):
        return True
    if processing_data or processed_data:
        return True
    if int(snapshot.get("success_goods") or 0) > 0 or int(snapshot.get("overall_goods") or 0) > 0:
        return True
    for payload in (processing, processed):
        if not isinstance(payload, dict):
            continue
        if int(payload.get("http_status") or 0) not in {0, 404}:
            data = payload.get("data")
            if data not in (None, {}, []):
                return True
    return False


def poll_upload_acceptance(
    api_key: str,
    upload_id: int,
    *,
    timeout_seconds: int = 20,
    poll_interval_seconds: int = 4,
) -> Dict[str, Any]:
    started = time.time()
    snapshots: List[Dict[str, Any]] = []
    last_snapshot: Dict[str, Any] = {"upload_id": int(upload_id)}
    while True:
        snapshot = get_upload_snapshot(api_key, int(upload_id), include_details=False)
        last_snapshot = snapshot
        snapshot_row = {
            "ts": common.utc_now_iso(),
            "status": snapshot.get("status"),
            "processing_status": (snapshot.get("processing_data") or {}).get("status"),
            "processed_status": (snapshot.get("processed_data") or {}).get("status"),
            "success_goods": snapshot.get("success_goods"),
            "overall_goods": snapshot.get("overall_goods"),
        }
        snapshots.append(snapshot_row)
        if upload_snapshot_indicates_acceptance(snapshot):
            snapshot["poll_log"] = snapshots
            snapshot["accepted"] = True
            return snapshot
        if time.time() - started >= max(5, int(timeout_seconds or 20)):
            snapshot["poll_log"] = snapshots
            snapshot["accepted"] = False
            snapshot["timed_out"] = True
            return snapshot
        time.sleep(max(2, int(poll_interval_seconds or 4)))


def get_upload_snapshot(
    api_key: str,
    upload_id: int,
    *,
    include_details: bool = False,
    detail_limit: int = 200,
    detail_max_pages: int = 1,
) -> Dict[str, Any]:
    upload_id = int(upload_id)
    processing = _request_json("GET", PRICES_UPLOAD_STATE_URL, api_key, params={"uploadID": upload_id})
    processed = _request_json("GET", PRICES_HISTORY_STATE_URL, api_key, params={"uploadID": upload_id})
    buffer_details: List[Dict[str, Any]] = []
    history_details: List[Dict[str, Any]] = []
    if include_details:
        buffer_details = _collect_upload_details(api_key, PRICES_UPLOAD_DETAILS_URL, upload_id, limit=max(20, int(detail_limit or 200)), max_pages=max(1, int(detail_max_pages or 1)))
        history_details = _collect_upload_details(api_key, PRICES_HISTORY_DETAILS_URL, upload_id, limit=max(20, int(detail_limit or 200)), max_pages=max(1, int(detail_max_pages or 1)))
    processing_data = processing.get("data") if isinstance(processing.get("data"), dict) else {}
    processed_data = processed.get("data") if isinstance(processed.get("data"), dict) else {}
    status = _clean(processing_data.get("status") or processed_data.get("status"))
    return {
        "upload_id": upload_id,
        "processing": processing,
        "processed": processed,
        "processing_data": processing_data,
        "processed_data": processed_data,
        "buffer_details": buffer_details,
        "history_details": history_details,
        "status": status,
        "success_goods": int(processed_data.get("successGoodsNumber") or processing_data.get("successGoodsNumber") or 0),
        "overall_goods": int(processed_data.get("overAllGoodsNumber") or processing_data.get("overAllGoodsNumber") or 0),
    }


def poll_upload_until_processed(
    api_key: str,
    upload_id: int,
    *,
    timeout_seconds: int = 120,
    poll_interval_seconds: int = 8,
) -> Dict[str, Any]:
    started = time.time()
    snapshots: List[Dict[str, Any]] = []
    last_snapshot: Dict[str, Any] = {"upload_id": int(upload_id)}
    while True:
        snapshot = get_upload_snapshot(api_key, int(upload_id), include_details=False)
        last_snapshot = snapshot
        processing_status = snapshot.get("processing_data") if isinstance(snapshot.get("processing_data"), dict) else {}
        processed_status = snapshot.get("processed_data") if isinstance(snapshot.get("processed_data"), dict) else {}
        processing_state_raw = processing_status.get("status")
        processed_state_raw = processed_status.get("status")
        processing_state = int(processing_state_raw or 0) if str(processing_state_raw or "").isdigit() else 0
        has_history_state = bool(processed_status)
        success_goods = int(snapshot.get("success_goods") or 0)
        overall_goods = int(snapshot.get("overall_goods") or 0)
        snapshots.append(
            {
                "ts": common.utc_now_iso(),
                "status": snapshot.get("status"),
                "processing_status": processing_state_raw,
                "processed_status": processed_state_raw,
                "success_goods": success_goods,
                "overall_goods": overall_goods,
            }
        )
        if has_history_state and processing_state in {0, 3}:
            snapshot["poll_log"] = snapshots
            snapshot["processed"] = True
            return snapshot
        if time.time() - started >= max(5, int(timeout_seconds or 120)):
            snapshot["poll_log"] = snapshots
            snapshot["timed_out"] = True
            snapshot["processed"] = False
            return snapshot
        time.sleep(max(3, int(poll_interval_seconds or 8)))
